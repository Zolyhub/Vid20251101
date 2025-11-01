import os
import cv2
import threading
import subprocess
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import Tk, Label, Button, filedialog, StringVar, ttk, messagebox, Scale
import configparser
import shutil
import json
from tkinter import Toplevel, Label, Button
from datetime import timedelta
import winsound  # Csak Windows-on működik
import numpy as np
import pandas as pd
import re
import locale
from fpdf import FPDF, XPos, YPos
import pytesseract
from tkinter import Toplevel, Label, Frame, Button
import chardet
import time
from datetime import datetime


pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'  #
# -------------------------------------------
# -----------------------------------------------
# === PARAMÉTEREK ===
font_path_regular = "F:/__Panel/fonts/DejaVuSans.ttf"
font_path_bold = "F:/__Panel/fonts/DejaVuSans-Bold.ttf"
ffmpeg_path = r"C:/ffmpeg/bin/ffmpeg.exe"
ffprobe_path = r"C:/ffmpeg/bin/ffprobe.exe"
# Kompatibilitási aliasok (a feldolgozó motor FFMPEG_PATH / FFPROBE_PATH neveket használ)
FFMPEG_PATH = ffmpeg_path
FFPROBE_PATH = ffprobe_path
DETECTION_MODE_FAST = "fast_ffmpeg"          # Gyors (FFmpeg scene detect)
DETECTION_MODE_DETAILED = "detailed_opencv"  # Részletes (OpenCV frame diff)

pixel_threshold_default = 2400
min_motion_duration_default = 2.0  # sec
motion_end_buffer_default = 1.0  # sec
# A runtime_log.txt fájlnév dinamikus beállítása
# Létrehozunk egy dátumot tartalmazó fájlnevet
log_filename = f"runtime_log_{datetime.now().strftime('%Y-%m-%d')}.txt"
#----------------------------------------

def sec_to_hms(sec: float) -> str:
    """Másodperc -> HH:MM:SS szöveg."""
    try:
        sec = int(round(sec))
        h = sec // 3600
        m = (sec % 3600) // 60
        s = sec % 60
        return f"{h:02d}:{m:02d}:{s:02d}"
    except Exception:
        return "00:00:00"


def get_video_duration_seconds(input_path: str) -> float | None:
    """
    ffprobe segítségével visszaadja a videó hosszát mp-ben.
    Ha hiba van, None.
    """
    try:
        cmd = [
            FFPROBE_PATH,
            "-v", "error",
            "-show_entries", "format=duration",
            "-of", "default=noprint_wrappers=1:nokey=1",
            input_path
        ]
        out = subprocess.check_output(cmd, stderr=subprocess.STDOUT)
        dur = float(out.decode("utf-8", errors="ignore").strip())
        return dur
    except Exception:
        return None


def detect_motion_fast_ffmpeg(
    input_path: str,
    scene_threshold: float = 0.05,   # <<< ITT állítod az érzékenységet kézzel
    merge_window_sec: float = 2.0,
    pad_before_sec: float = 2.0,
    pad_after_sec: float = 2.0
):
    """
    Gyors mozgás/jelenet detektálás FFmpeg segítségével.
    -> NINCS automata küszöb, csak fix scene_threshold.
    -> Mindig 3 értéket ad vissza:
       (segments_list, segments_count, total_motion_time_sec)
    """

    import subprocess, os, re, datetime

    # --- app_ref csak a logoláshoz / log_folderhez kell ---
    app_ref = globals().get("app", None)
    # Biztonságos GUI-frissítő segédfüggvény
    def gui_set_threshold_label(value: float):
        try:
            if app_ref and hasattr(app_ref, "auto_threshold_value_var"):
                # Frissítés a Tkinter főszálban
                app_ref.root.after(0, lambda: app_ref.auto_threshold_value_var.set(f"Küszöb: {value:.2f} (fix)"))
        except Exception as e:
            dbg_write(f"[WARN] GUI küszöb label frissítés hiba: {e}")

    # --- Debug log könyvtár: a napi _Logok mappán belül legyen _debug_ffmpeg ---
    if app_ref and hasattr(app_ref, "log_folder"):
        debug_dir = os.path.join(app_ref.log_folder, "_debug_ffmpeg")
    else:
        debug_dir = os.path.join(os.path.dirname(input_path), "_debug_ffmpeg")
    os.makedirs(debug_dir, exist_ok=True)

    debug_file = os.path.join(
        debug_dir,
        f"ffmpeg_motion_{datetime.datetime.now().strftime('%H-%M-%S')}.txt"
    )

    def dbg_write(text: str):
        with open(debug_file, "a", encoding="utf-8") as f:
            f.write(text + "\n")

    # ---------------- DEBUG FEJLÉC ----------------
    dbg_write("=== FFmpeg motion detect START ===")
    dbg_write(f"Input: {input_path}")
    dbg_write(f"Scene_threshold FIX: {scene_threshold}")

    # --- Útvonal normalizálása FFmpeg-hez ---
    norm_input = os.path.normpath(input_path).replace("\\", "/")
    dbg_write(f"Normalized input path: {norm_input}")

    # --- Itt NINCS auto_threshold_var, nincs threshold_auto_tune ---
    # --- Küszöb meghatározása: Spinboxból vagy paraméterből ---
    thr = float(scene_threshold)
    if app_ref:
        try:
            if hasattr(app_ref, "ffmpeg_threshold_var"):
                thr = float(app_ref.ffmpeg_threshold_var.get())
                dbg_write(f"[SPINBOX] Küszöb a GUI-ból: {thr:.3f}")
            else:
                dbg_write(f"[DEFAULT] Küszöb paraméterből: {thr:.3f}")
        except Exception as e:
            dbg_write(f"[WARN] Küszöb lekérés hiba: {e}")

    gui_set_threshold_label(thr)
    dbg_write(f"[FINAL] Használt küszöb: {thr:.3f}")




    # --- FFmpeg parancs összeállítása ---
    # Fontos: escapelni kell a vesszőt Windows miatt \,
    filter_str = f"select=gt(scene\\,{thr:.3f}),showinfo,framestep=5"

    cmd = [
        ffmpeg_path,
        "-hide_banner",
        "-i", norm_input,
        "-filter:v", filter_str,
        "-f", "null",
        "-"
    ]

    dbg_write(f"Running command: {' '.join(cmd)}")

    try:
        result = subprocess.run(
            cmd,
            stderr=subprocess.PIPE,
            stdout=subprocess.DEVNULL,
            text=True,
            encoding="utf-8",
            errors="ignore"
        )

        # LOG: FFmpeg kimenet
        dbg_write("--- FFmpeg stderr BEGIN ---")
        dbg_write(result.stderr[:10000])  # max 10k karakter
        dbg_write("--- FFmpeg stderr END ---")
        dbg_write(f"Return code: {result.returncode}")

        # --- Képkocka-idők kigyűjtése FFmpeg showinfo-ból ---
        # Itt most azokra a frame-ekre számítunk, amiket a select átengedett,
        # tehát ezek "jelenetváltásként/mozgásként érdekes frame-ek".
        matches = re.findall(r"pts_time:([\d\.]+)", result.stderr)
        trigger_times = [float(m) for m in matches]
        dbg_write(f"Detected trigger frames: {len(trigger_times)}")

        # Ha nem talált semmit → fallback = teljes videó
        if not trigger_times:
            dur = get_video_duration_seconds(norm_input) or 0.0
            dbg_write(f"No motion found → full video fallback ({dur:.2f}s)")
            if app_ref:
                app_ref.add_log_entry(
                    "INFO",
                    f"Nincs FFmpeg-által jelzett mozgás, teljes videó átvéve ({dur:.2f} mp)."
                )
            # vissza: [ (0,dur) ], count=1, total_motion_time=dur
            return [(0.0, dur)], 1, dur

        # --- Szomszédos trigger frame-ekből "szegmensek" építése ---
        merged_segments = []
        last_start = max(0.0, trigger_times[0] - pad_before_sec)

        for i in range(1, len(trigger_times)):
            gap = trigger_times[i] - trigger_times[i - 1]
            # ha túl nagy a szünet 2 trigger között, lezárjuk az előző szegmenst
            if gap > merge_window_sec:
                merged_segments.append(
                    (last_start, trigger_times[i - 1] + pad_after_sec)
                )
                last_start = max(0.0, trigger_times[i] - pad_before_sec)

        # utolsó szegmens lezárása
        merged_segments.append(
            (last_start, trigger_times[-1] + pad_after_sec)
        )

        # --- Össz-idő és stat ---
        total_motion_time = sum(e - s for s, e in merged_segments)
        segment_count = len(merged_segments)

        dbg_write(f"Segments merged: {segment_count}")
        dbg_write(f"Total motion time: {total_motion_time:.2f} s")

        if app_ref:
            human_total = sec_to_hms(total_motion_time)
            app_ref.add_log_entry(
                "INFO",
                f"{os.path.basename(input_path)}: {segment_count} mozgásszegmens, összesen {human_total} ({total_motion_time:.2f} mp)."
            )

        dbg_write("=== FFmpeg motion detect END OK ===")

        # Mindig 3 érték
        return merged_segments, segment_count, total_motion_time

    except Exception as e:
        # Ha FFmpeg hiba van, ne álljon le a feldolgozás
        dbg_write(f"[ERROR] detect_motion_fast_ffmpeg exception: {e}")
        if app_ref:
            app_ref.add_log_entry(
                "ERROR",
                f"FFmpeg hiba mozgásdetektálás közben: {e}"
            )
        dur = get_video_duration_seconds(norm_input) or 0.0
        return [(0.0, dur)], 1, dur







def threshold_auto_tune(input_path: str, sample_seconds: int = 10) -> float:
    """
    Automatikus küszöb-ajánló FFmpeg 'scene_score' alapján.
    10 másodperces mintát elemez, majd visszaadja az optimális küszöbértéket.
    Példa: 0.28 vagy 0.45 – minél magasabb, annál szigorúbb a mozgásdetektálás.
    """
    import subprocess, re, statistics

    if not os.path.exists(input_path):
        print("[threshold_auto_tune] ⚠️ A bemeneti fájl nem található.")
        return 0.3

    try:
        cmd = [
            FFMPEG_PATH,
            "-hide_banner",
            "-ss", "0",
            "-t", str(sample_seconds),
            "-i", input_path,
            "-filter:v", f"select=gt(scene,{auto_thr}),showinfo",
            "-f", "null", "-"
        ]

        proc = subprocess.Popen(
            cmd,
            stderr=subprocess.PIPE,
            stdout=subprocess.DEVNULL,
            universal_newlines=True,
            encoding="utf-8",
            errors="ignore"
        )

        scene_scores = []
        for line in proc.stderr:
            if "scene_score" in line:
                m = re.search(r"scene_score:([0-9\.]+)", line)
                if m:
                    val = float(m.group(1))
                    scene_scores.append(val)

        proc.wait(timeout=30)

        if not scene_scores:
            print("[threshold_auto_tune] ⚪ Nincs mérhető scene_score érték – fallback 0.3")
            return 0.3

        mean_val = statistics.mean(scene_scores)
        p90 = statistics.quantiles(scene_scores, n=10)[-1]  # felső 10% átlag
        auto_thr = round(min(0.9, max(0.1, (mean_val + p90) / 2)), 3)

        print(f"[threshold_auto_tune] 📊 Átlag={mean_val:.3f}, 90p={p90:.3f} → ajánlott küszöb={auto_thr:.3f}")
        return auto_thr

    except Exception as e:
        print(f"[threshold_auto_tune] ⚠️ Hiba történt: {e}")
        return 0.3

def detect_motion_detailed_opencv(    input_path: str,    pixel_threshold: int = 5000,    frame_skip: int = 5,     resize_width: int = 640,     resize_height: int = 360,     merge_window_sec: float = 2.0,
    pad_before_sec: float = 1.0,     pad_after_sec: float = 1.0
):
    """
    Lassabb részletes mód, de még mindig többszörös gyorsítás:
    - csak minden 5. frame,
    - 640x360-ra lekicsinyítve,
    - pixeldiff számolás.

    Ha semmit nem talál, visszaadja a teljes videót.
    """
    cap = cv2.VideoCapture(input_path)
    if not cap.isOpened():
        raise RuntimeError(f"Nem nyitható meg: {input_path}")

    fps = cap.get(cv2.CAP_PROP_FPS) or 25.0
    frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    duration_sec = frame_count / fps if fps else 0.0

    ok, prev_frame = cap.read()
    if not ok or prev_frame is None:
        cap.release()
        return [(0.0, duration_sec if duration_sec > 0 else 60.0)]

    prev_gray = cv2.cvtColor(
        cv2.resize(prev_frame, (resize_width, resize_height)),
        cv2.COLOR_BGR2GRAY
    )

    motion_hits = []
    frame_idx = 0
    while True:
        cap.set(cv2.CAP_PROP_POS_FRAMES, frame_idx)
        ret, frame = cap.read()
        if not ret:
            break

        gray = cv2.cvtColor(
            cv2.resize(frame, (resize_width, resize_height)),
            cv2.COLOR_BGR2GRAY
        )
        diff = cv2.absdiff(prev_gray, gray)
        _, thresh = cv2.threshold(diff, 25, 255, cv2.THRESH_BINARY)
        motion_level = cv2.countNonZero(thresh)

        if motion_level > pixel_threshold:
            cur_t = frame_idx / fps
            motion_hits.append(cur_t)

        prev_gray = gray
        frame_idx += frame_skip

    cap.release()

    if not motion_hits:
        return [(0.0, duration_sec if duration_sec > 0 else 60.0)]

    raw_segments = []
    for t in motion_hits:
        s = max(0.0, t - pad_before_sec)
        e = (t + pad_after_sec) if duration_sec == 0 else min(duration_sec, t + pad_after_sec)
        raw_segments.append([s, e])

    raw_segments.sort(key=lambda x: x[0])
    merged = []
    for s, e in raw_segments:
        if not merged:
            merged.append([s, e])
        else:
            ls, le = merged[-1]
            if s <= le + merge_window_sec:
                merged[-1][1] = max(le, e)
            else:
                merged.append([s, e])

    final_segments = [(s, e) for (s, e) in merged if (e - s) >= 0.8]
    if not final_segments:
        return [(0.0, duration_sec if duration_sec > 0 else 60.0)]

    return final_segments


def cut_and_compress_segment(
    input_path: str,
    output_path: str,
    start_sec: float,
    end_sec: float,
    crf: int,
    preset: str,
    audio_bitrate: str,
):
    """
    FFmpeg kivágás + CRF tömörítés valós idejű (interpolált, stabil) GUI frissítéssel.
    Tartalmaz:
    - simított progress (aktuális + összes)
    - élő státusz: eltelt idő, sebesség, MB/s, fps, hátralévő idő
    - hibatűrő GUI frissítés (after_idle)
    """

    import subprocess, threading, re, time, os

    duration = max(0.01, end_sec - start_sec)
    cmd = [
        FFMPEG_PATH,
        "-hide_banner",
        "-ss", f"{start_sec:.3f}",
        "-t", f"{duration:.3f}",
        "-i", input_path,
        "-c:v", "libx264",
        "-crf", str(crf),
        "-preset", preset,
        "-c:a", "aac",
        "-b:a", audio_bitrate,
        "-movflags", "+faststart",
        "-y",
        output_path
    ]

    process = subprocess.Popen(
        cmd,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.PIPE,
        text=True,
        encoding="utf-8",
        errors="ignore"
    )

    time_pattern = re.compile(r"time=(\d+):(\d+):(\d+\.\d+)")
    start_time = time.time()
    last_update = 0.0
    stderr_buf = []

    app = globals().get("APP_INSTANCE", None)
    current_progress = 0.0
    target_progress = 0.0
    smooth_stop_flag = False

    # -------------------------------------------------------------
    # SMOOTH ANIMÁTOR: folyamatos progress interpoláció (0.1 s)
    # -------------------------------------------------------------
    def smooth_animator():
        nonlocal current_progress, target_progress, smooth_stop_flag
        while not smooth_stop_flag:
            if app:
                step = (target_progress - current_progress) * 0.25
                current_progress += step
                if current_progress > 100:
                    current_progress = 100

                try:
                    total_files = len(app.input_files)
                    done_files = app.processed_files_count
                    overall = ((done_files + current_progress / 100.0) / total_files * 100.0) if total_files > 0 else current_progress

                    app.root.after_idle(lambda p1=current_progress, p2=overall: (
                        app.file_progress.config(value=p1),
                        app.file_progress_label_text.set(f"Aktuális fájl feldolgozása: {p1:.1f}%"),
                        app.overall_progress.config(value=p2),
                        app.overall_progress_label_text.set(f"Összes fájl feldolgozása: {p2:.1f}%")
                    ))
                except Exception:
                    pass
            time.sleep(0.1)

    threading.Thread(target=smooth_animator, daemon=True).start()

    # -------------------------------------------------------------
    # FFMPEG WATCHER: stderr olvasás + sebesség/idő kijelzés
    # -------------------------------------------------------------
    def watcher():
        nonlocal target_progress, last_update
        for line in process.stderr:
            stderr_buf.append(line)
            if "time=" not in line:
                continue

            m = time_pattern.search(line)
            if not m:
                continue

            h, m_, s = map(float, m.groups())
            current_sec = h * 3600 + m_ * 60 + s
            progress = min(100.0, (current_sec / duration) * 100.0)
            target_progress = progress

            elapsed = time.time() - start_time
            remaining = (elapsed / progress * (100 - progress)) if progress > 0 else 0
            time_str = sec_to_hms(current_sec)
            rem_str = sec_to_hms(remaining)

            # sebesség + MB/s + fps
            input_mb = getattr(app, "current_input_size_mb", 0.0)
            fps = getattr(app, "current_input_fps", 25.0)
            processed_mb = (input_mb / duration) * current_sec if duration > 0 else 0.0
            mbps = processed_mb / elapsed if elapsed > 0 else 0.0
            speed_ratio = current_sec / elapsed if elapsed > 0 else 0.0
            est_fps = fps * speed_ratio

            if app and (time.time() - last_update > 0.3):
                try:
                    app.root.after_idle(lambda: (
                        app.status.set(f"FFmpeg: {time_str} | {progress:.1f}% | Hátralévő ~{rem_str}"),
                        app.runtime_info_var.set(
                            f"🕓 {sec_to_hms(elapsed)} | ⚙️ {speed_ratio:4.2f}× | "
                            f"💾 {mbps:5.2f} MB/s | 🎞️ {est_fps:5.1f} fps | "
                            f"📁 {app.processed_files_count+1}/{len(app.input_files)} | 💾 {progress:5.1f}%"
                        ),
                        setattr(app, "current_file_progress_input_duration", current_sec),
                        app.update_stats()
                    ))
                    last_update = time.time()
                except Exception:
                    pass

    threading.Thread(target=watcher, daemon=True).start()

    # -------------------------------------------------------------
    # FFmpeg befejezése
    # -------------------------------------------------------------
    process.wait()
    smooth_stop_flag = True
    end_time = time.time()

    success = (
        process.returncode == 0
        and os.path.exists(output_path)
        and os.path.getsize(output_path) > 0
    )

    out_size_mb = 0.0
    if os.path.exists(output_path):
        out_size_mb = os.path.getsize(output_path) / (1024 * 1024)

    stderr_tail = "".join(stderr_buf)[-2000:]

    # GUI frissítés a végén
    if app:
        app.root.after_idle(lambda: (
            app.file_progress.config(value=100),
            app.file_progress_label_text.set("Aktuális fájl feldolgozása: 100%"),
            app.overall_progress_label_text.set("Összes fájl feldolgozása: 100%"),
            app.status.set("✅ FFmpeg feldolgozás kész."),
            app.update_stats()
        ))

    return {
        "ok": success,
        "run_seconds": end_time - start_time,
        "out_size_mb": out_size_mb,
        "duration_hms": sec_to_hms(duration),
        "stderr_tail": stderr_tail,
    }










def calculate_file_stats(input_path: str,
                         merged_output_path: str,
                         motion_segments: list,
                         total_run_seconds: float):
    """
    Statisztikák Log1/TreeView-hoz:
    - Be MB
    - Ki MB
    - Megtakarítás %
    - Mozgás arány %
    - Feldolg. futásidő
    """
    try:
        input_mb = os.path.getsize(input_path) / (1024 * 1024)
    except Exception:
        input_mb = 0.0

    try:
        output_mb = os.path.getsize(merged_output_path) / (1024 * 1024)
    except Exception:
        output_mb = 0.0

    compression_pct = 0.0
    if input_mb > 0:
        compression_pct = (1 - (output_mb / input_mb)) * 100.0

    motion_time_sec = sum(max(0, e - s) for s, e in motion_segments)
    full_len = get_video_duration_seconds(input_path) or 1.0
    motion_pct = (motion_time_sec / full_len) * 100.0

    return {
        "input_mb": round(input_mb, 2),
        "output_mb": round(output_mb, 2),
        "compression_pct": round(compression_pct, 2),
        "motion_pct": round(motion_pct, 2),
        "motion_time_sec": round(motion_time_sec, 2),
        "run_hms": sec_to_hms(total_run_seconds),
        "run_seconds": total_run_seconds
    }

#-------------------------------------

class MotionExtractorApp:
    def __init__(self, root):
        # A globális app referencia beállítása
        globals()["app"] = self

        # Lokalizáció beállítása csak a GUI megjelenítéshez (dátumok, stb.), de számokhoz 'C' lokalizáció
        locale.setlocale(locale.LC_TIME, 'hu_HU.UTF-8')  # Magyar dátumformátumok
        locale.setlocale(locale.LC_NUMERIC, 'C')  # Angol tizedes tört (0.5)

        self.root = root
        script_name = os.path.basename(__file__)
        script_dir = os.path.dirname(os.path.abspath(__file__))

        self.script_name = script_name
        self.script_dir = script_dir
        self.play_sound = True  # Alapértelmezett érték: bekapcsolt hangjelzés
        self.errors = []  # A hibák gyűjtésére szolgáló lista

        self.root.title(f"Video Mozgásérzékelő és Feldolgozó - {self.script_name}")
        self.root.geometry("1400x900")
        self.root.resizable(True, True)
        self.is_paused = False
        self.stop_processing_flag = False
        self.processing_thread = None
        self.detection_mode_var = tk.StringVar(value=DETECTION_MODE_FAST)
        self.log_line_number = 0
        self.runtime_log = []  # <<< Ezt a sort adja hozzá
        self.current_log_file = None  # <<< EZT A SORT KELL HOZZÁADNI

        self.current_file_index = -1
        self.processed_files_count = 0
        self.current_threshold_value = 0.005  # Alapértelmezett FFmpeg scene küszöb

        self.processed_size_mb = 0
        self.processed_duration_sec = 0
        self.current_file_progress_duration = 0
        self.current_file_progress_size = 0
        self.start_time = None
        self.end_time = None
        self.input_directory = ""
        self.output_folder = ""
        self.input_files = []
        self.total_size_mb = 0
        self.total_duration_sec = 0
        self.start_time = None
        self.calculated_end_time = None
        self.remaining_time = None
        self.end_time = None

        self.processed_input_size_mb = 0
        self.processed_output_size_mb = 0
        self.processed_input_duration_sec = 0
        self.processed_output_duration_sec = 0
        self.current_file_progress_input_size = 0
        self.current_file_progress_input_duration = 0
        self.tree_items = {}
        self.settings_file = "settings.ini"
        self.processing_state_file = "processing_state.json"


        self.log1_data = []
        self.log2_data = {}
        self.log3_data = []

        self.status = StringVar()
        self.status.set("Válassz ki egy bemeneti mappát és kimeneti mappát.")

        # Log mappa változó inicializálása
        self.log_folder = ""  # Alapértelmezett üres érték
        self.log_dir_var = StringVar(value="")  # GUI-ból olvasható log mappa

        # --- GUI felépítése ---
        main_frame = ttk.Frame(root, padding="10")
        main_frame.pack(fill="both", expand=True)

        main_frame.grid_columnconfigure(0, weight=1)
        main_frame.grid_columnconfigure(1, weight=0)
        main_frame.grid_rowconfigure(0, weight=0)
        main_frame.grid_rowconfigure(1, weight=1)
        main_frame.grid_rowconfigure(2, weight=0)

        # Bal oldali top panel (Napló és Statisztika)
        left_top_panel = ttk.Frame(main_frame)
        left_top_panel.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        left_top_panel.grid_columnconfigure(0, weight=1)
        left_top_panel.grid_rowconfigure(0, weight=0)
        left_top_panel.grid_rowconfigure(1, weight=0)

        log_frame = ttk.Frame(left_top_panel)
        log_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        log_frame.grid_columnconfigure(0, weight=1)
        log_frame.grid_rowconfigure(0, weight=1)

        ttk.Label(log_frame, text="Napló", font=("Helvetica", 14, "bold")).grid(row=0, column=0, sticky="w")

        self.log_text = ttk.Treeview(log_frame, columns=("time", "level", "message"), show="headings", height=5)
        self.log_text.heading("time", text="Idő")
        self.log_text.heading("level", text="Szint")
        self.log_text.heading("message", text="Üzenet")
        self.log_text.column("time", width=150)
        self.log_text.column("level", width=80)
        self.log_text.column("message", width=800)
        self.log_text.grid(row=1, column=0, sticky="nsew", pady=5)

        self.log_scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=self.log_scrollbar.set)
        self.log_scrollbar.grid(row=1, column=1, sticky="ns")

        self.check_dependencies()

        self.stats_frame = ttk.LabelFrame(left_top_panel, text="Statisztika", padding="10")
        self.stats_frame.grid(row=1, column=0, sticky="ew", pady=(0, 10))
        # --- statisztika oszlopbeállítások, hogy ne ugráljon ---
        self.stats_frame.grid_columnconfigure(0, weight=0, minsize=110)  # címkék (pl. "Fájlok:")
        self.stats_frame.grid_columnconfigure(1, weight=0, minsize=160)  # Összesen
        self.stats_frame.grid_columnconfigure(2, weight=0, minsize=160)  # Feldolgozva
        self.stats_frame.grid_columnconfigure(3, weight=0, minsize=160)  # Hátralévő
        self.stats_frame.grid_columnconfigure(4, weight=1)  # a jobb oldali töltés

        self.total_files_var = StringVar(value="Összesen: N/A")
        self.processed_files_var = StringVar(value="Feldolgozva: N/A")
        self.remaining_files_var = StringVar(value="Hátralévő: N/A")

        self.total_size_var = StringVar(value="Összesen: N/A MB")
        self.processed_size_var = StringVar(value="Feldolgozva: N/A MB")
        self.remaining_size_var = StringVar(value="Hátralévő: N/A MB")

        self.total_duration_var = StringVar(value="Összesen: N/A")
        self.processed_duration_var = StringVar(value="Feldolgozva: N/A")
        self.remaining_duration_var = StringVar(value="Hátralévő: N/A")

        self.start_time_var = StringVar(value="Kezdés: N/A")
        self.elapsed_time_var = StringVar(value="Eltelt: N/A")
        self.remaining_time_var = StringVar(value="Hátralévő: N/A")
        self.end_time_var = StringVar(value="Várható zárás: N/A")
        self.total_processing_time_var = StringVar(value="Össz. futásidő: N/A")
        self.summary_text_var = StringVar(value="Összesített: N/A")

        self.total_output_mb_var = StringVar(value="Összes kimenő MB: N/A")
        self.saving_percent_var = StringVar(value="Megtakarítás: N/A %")

        self.custom_resolution_var = StringVar(value="0")
        self.output_width_var = StringVar(value="1920")
        self.output_height_var = StringVar(value="1080")

        stats_labels = [
            ("Fájlok:", self.total_files_var, self.processed_files_var, self.remaining_files_var),
            ("Méret:", self.total_size_var, self.processed_size_var, self.remaining_size_var),
            ("Idő:", self.total_duration_var, self.processed_duration_var, self.remaining_duration_var),
            ("Időpontok:", self.start_time_var, self.elapsed_time_var, self.remaining_time_var, self.end_time_var),
            ("Futásidő:", self.total_processing_time_var)
        ]

        for r, (label_text, *vars) in enumerate(stats_labels):
            ttk.Label(
                self.stats_frame,
                text=label_text,
                font=("Segoe UI", 9, "bold"),
                anchor="w"
            ).grid(row=r, column=0, padx=(10, 5), pady=2, sticky="w")

            for c, var in enumerate(vars):
                ttk.Label(
                    self.stats_frame,
                    textvariable=var,
                    font=("Consolas", 9),
                    anchor="e",  # jobbra igazítás, hogy a számok oszlopban álljanak
                    width=18  # fix szélesség minden cellának
                ).grid(row=r, column=c + 1, padx=5, pady=2, sticky="ew")

        # --- Alsó sor: részletes fájlstatisztika (kisebb/nagyobb fájlok) ---
        self.summary_files_detail_var = StringVar(value="Részletes statisztika: N/A")
        self.summary_files_detail_label = ttk.Label(
            self.stats_frame,
            textvariable=self.summary_files_detail_var,
            font=("Consolas", 9),
            foreground="#004080"  # alap: kék
        )
        self.summary_files_detail_label.grid(row=r + 2, column=0, columnspan=4, padx=10, pady=(4, 6), sticky="w")

        # --- Összesített / megtakarítás sorok (külön, a dinamikus statisztika alatt) ---
        r += 1  # következő sor

        # Összesített sor (ez tartalmazza a kisebb/nagyobb fájl statisztikát is)
        ttk.Label(
            self.stats_frame,
            textvariable=self.summary_text_var,
            font=("Segoe UI", 10, "bold"),
            justify="left",
        ).grid(row=r, column=0, columnspan=5, padx=10, pady=(4, 0), sticky="w")

        # Kimenő MB + Megtakarítás % egy sorban
        r += 1
        ttk.Label(
            self.stats_frame,
            textvariable=self.total_output_mb_var,
            font=("Segoe UI", 10),
            anchor="w",
        ).grid(row=r, column=0, columnspan=2, padx=10, pady=(0, 4), sticky="w")

        ttk.Label(
            self.stats_frame,
            textvariable=self.saving_percent_var,
            font=("Segoe UI", 10, "bold"),
            foreground="#0055cc",
            anchor="e",
        ).grid(row=r, column=2, columnspan=3, padx=10, pady=(0, 4), sticky="e")

        # ------------------------------------------------------------
        # --- Separator + Összesített eredmény szekció ---
        # A separator közvetlenül a statisztika utolsó sora (r) után kerül
        ttk.Separator(self.stats_frame, orient="horizontal").grid(
            row=r + 1, column=0, columnspan=5, sticky="ew", pady=(6, 8)
        )

        # Az összesített blokk az elválasztó alá kerül
        summary_row = r + 2

        # --- Első sor: Összesített / Be / Ki / Megtakarítás ---
        ttk.Label(
            self.stats_frame,
            text="Összesített:",
            font=("Segoe UI", 10, "bold"),
            anchor="w"
        ).grid(row=summary_row, column=0, padx=(10, 5), pady=(4, 2), sticky="w")

        self.summary_input_var = StringVar(value="Be: N/A MB")
        ttk.Label(
            self.stats_frame,
            textvariable=self.summary_input_var,
            font=("Consolas", 9),
            anchor="center"
        ).grid(row=summary_row, column=1, padx=5, pady=(4, 2), sticky="ew")

        self.summary_output_var = StringVar(value="Ki: N/A MB")
        ttk.Label(
            self.stats_frame,
            textvariable=self.summary_output_var,
            font=("Consolas", 9),
            anchor="center"
        ).grid(row=summary_row, column=2, padx=5, pady=(4, 2), sticky="ew")

        self.summary_saving_var = StringVar(value="Megtakarítás: N/A %")
        self.summary_saving_label = ttk.Label(
            self.stats_frame,
            textvariable=self.summary_saving_var,
            font=("Segoe UI", 10, "bold"),
            foreground="#0055cc",
            anchor="e"
        )
        self.summary_saving_label.grid(
            row=summary_row, column=3, columnspan=2, padx=10, pady=(4, 2), sticky="e"
        )

        # --- Második sor: Összes kimenő MB és részletes statisztika ---
        summary_row += 1
        ttk.Label(
            self.stats_frame,
            textvariable=self.total_output_mb_var,
            font=("Consolas", 9),
            anchor="w"
        ).grid(row=summary_row, column=0, columnspan=2, padx=10, pady=(2, 6), sticky="w")

        self.summary_files_detail_label.grid(
            row=summary_row,
            column=2,
            columnspan=3,
            padx=10,
            pady=(2, 6),
            sticky="e"
        )
        # ------------------------------------------------------------

        # Jobb oldali panel (Feldolgozási beállítások)
        right_panel = ttk.Frame(main_frame, padding="10")
        right_panel.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)
        right_panel.grid_columnconfigure(0, weight=1)
        right_panel.grid_rowconfigure(0, weight=1)

        settings_frame = ttk.LabelFrame(right_panel, text="Feldolgozási beállítások", padding="10")
        settings_frame.grid(row=0, column=0, sticky='nsew')
        settings_frame.grid_columnconfigure(1, weight=1)

        row_index = 0
        ttk.Label(settings_frame, text="Videó minőség (CRF):").grid(row=row_index, column=0, sticky="w", pady=2)
        self.crf_var = StringVar(value="23")
        self.crf_scale = Scale(settings_frame, from_=0, to=51, orient='horizontal',
                               command=lambda v: self.update_scale_label(self.crf_var, v))
        self.crf_scale.set(23)
        self.crf_scale.grid(row=row_index, column=1, sticky="we")
        ttk.Label(settings_frame, textvariable=self.crf_var).grid(row=row_index, column=2, padx=5, sticky="w")
        row_index += 1

        ttk.Label(settings_frame, text="Kódolási sebesség (Preset):").grid(row=row_index, column=0, sticky="w", pady=2)
        self.preset_var = StringVar(value="medium")
        presets = ["ultrafast", "superfast", "veryfast", "faster", "fast", "medium", "slow", "slower", "veryslow"]
        self.preset_menu = ttk.Combobox(settings_frame, textvariable=self.preset_var, values=presets, state="readonly")
        self.preset_menu.grid(row=row_index, column=1, sticky="we")
        row_index += 1

        # --- Mozgásdetektálási mód választása (Gyors / Részletes) ---
        motion_mode_frame = ttk.LabelFrame(settings_frame, text="Mozgásdetektálás módja")
        motion_mode_frame.grid(row=row_index, column=0, columnspan=3, sticky="ew", pady=(5, 5), padx=(5, 5))
        row_index += 1

        self.detection_mode_var = tk.StringVar(value=DETECTION_MODE_FAST)

        ttk.Radiobutton(
            motion_mode_frame,
            text="Gyors (FFmpeg)",
            variable=self.detection_mode_var,
            value=DETECTION_MODE_FAST
        ).pack(anchor="w", padx=10, pady=2)

        ttk.Radiobutton(
            motion_mode_frame,
            text="Részletes (OpenCV)",
            variable=self.detection_mode_var,
            value=DETECTION_MODE_DETAILED
        ).pack(anchor="w", padx=10, pady=2)

        ttk.Label(settings_frame, text="Mozgásérzékenységi küszöb (pixel):").grid(row=row_index, column=0, sticky="w",
                                                                                  pady=2)
        self.pixel_threshold_var = StringVar(value=str(pixel_threshold_default))
        self.pixel_threshold_scale = Scale(settings_frame, from_=1000, to=50000, orient='horizontal', resolution=100,
                                           command=lambda v: self.update_scale_label(self.pixel_threshold_var, v))
        self.pixel_threshold_scale.set(pixel_threshold_default)
        self.pixel_threshold_scale.grid(row=row_index, column=1, sticky="we")
        ttk.Label(settings_frame, textvariable=self.pixel_threshold_var).grid(row=row_index, column=2, padx=5,
                                                                              sticky="w")
        row_index += 1

        # --- FFmpeg mozgásküszöb Spinbox ---
        threshold_frame = ttk.Frame(settings_frame)

        threshold_frame.grid(row=motion_mode_frame.grid_info()["row"], column=1, sticky="ne", padx=(10, 10))

        ttk.Label(
            threshold_frame,
            text="FFmpeg küszöb:",
            font=("Segoe UI", 9)
        ).pack(side="left", padx=(0, 5))

        self.ffmpeg_threshold_var = tk.DoubleVar(value=0.30)
        self.ffmpeg_threshold_spin = ttk.Spinbox(
            threshold_frame,
            from_=0.01,
            to=0.50,
            increment=0.01,
            format="%.2f",
            textvariable=self.ffmpeg_threshold_var,
            width=5,
            command=lambda: self.update_ffmpeg_threshold()
        )
        self.ffmpeg_threshold_spin.pack(side="left")

        ttk.Label(
            threshold_frame,
            text=" (0.01–0.50)",
            font=("Segoe UI", 8, "italic"),
            foreground="#666666"
        ).pack(side="left", padx=(3, 0))


        # --- Automatikus FFmpeg scene threshold ---
        row_index += 1
        self.auto_threshold_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            settings_frame,
            text="Automatikus mozgásküszöb (FFmpeg)",
            variable=self.auto_threshold_var
        ).grid(row=row_index, column=0, columnspan=2, sticky="w", pady=(5, 2))

        # 🔹 Ha az automatikus mód be van pipálva, tiltsuk le a Spinboxot
        self.auto_threshold_var.trace_add(
            "write",
            lambda *a: self.ffmpeg_threshold_spin.config(
                state=("disabled" if self.auto_threshold_var.get() else "normal")
            )
        )

        # aktuális érték megjelenítése
        self.auto_threshold_value_var = tk.StringVar(value="Küszöb: 0.30 (fix)")
        ttk.Label(
            settings_frame,
            textvariable=self.auto_threshold_value_var,
            font=("Consolas", 9),
            foreground="#0078D7"
        ).grid(row=row_index, column=2, sticky="e", padx=(10, 5))
        row_index += 1

        ttk.Label(settings_frame, text="Min. mozgásklip hossza (mp):").grid(row=row_index, column=0, sticky="w", pady=2)
        self.min_motion_duration_var = StringVar(value=f"{min_motion_duration_default:.1f}".replace('.', ','))
        self.min_motion_duration_scale = Scale(settings_frame, from_=0.1, to=10, orient='horizontal', resolution=0.1,
                                               command=lambda v: self.update_scale_label(self.min_motion_duration_var,
                                                                                         v, decimals=1))
        self.min_motion_duration_scale.set(min_motion_duration_default)
        self.min_motion_duration_scale.grid(row=row_index, column=1, sticky="we")
        ttk.Label(settings_frame, textvariable=self.min_motion_duration_var).grid(row=row_index, column=2, padx=5,
                                                                                  sticky="w")
        row_index += 1

        ttk.Label(settings_frame, text="Üresjárat hossza (mp):").grid(row=row_index, column=0, sticky="w", pady=2)
        self.idle_duration_var = StringVar(value="5,0")
        self.idle_duration_scale = Scale(settings_frame, from_=0.1, to=10, orient='horizontal', resolution=0.1,
                                         command=lambda v: self.update_scale_label(self.idle_duration_var, v,
                                                                                   decimals=1))
        self.idle_duration_scale.set(5.0)
        self.idle_duration_scale.grid(row=row_index, column=1, sticky="we")
        ttk.Label(settings_frame, textvariable=self.idle_duration_var).grid(row=row_index, column=2, padx=5, sticky="w")
        row_index += 1

        ttk.Label(settings_frame, text="Elő-mozgás puffer (mp):").grid(row=row_index, column=0, sticky="w", pady=2)
        self.pre_motion_buffer_var = StringVar(value="1,0")
        self.pre_motion_buffer_scale = Scale(settings_frame, from_=0, to=5, orient='horizontal', resolution=0.1,
                                             command=lambda v: self.update_scale_label(self.pre_motion_buffer_var, v,
                                                                                       decimals=1))
        self.pre_motion_buffer_scale.set(1.0)
        self.pre_motion_buffer_scale.grid(row=row_index, column=1, sticky="we")
        ttk.Label(settings_frame, textvariable=self.pre_motion_buffer_var).grid(row=row_index, column=2, padx=5,
                                                                                sticky="w")
        row_index += 1

        ttk.Label(settings_frame, text="Utó-mozgás puffer (mp):").grid(row=row_index, column=0, sticky="w", pady=2)
        self.motion_end_buffer_var = StringVar(value=f"{motion_end_buffer_default:.1f}".replace('.', ','))
        self.motion_end_buffer_scale = Scale(settings_frame, from_=0, to=5, orient='horizontal', resolution=0.1,
                                             command=lambda v: self.update_scale_label(self.motion_end_buffer_var, v,
                                                                                       decimals=1))
        self.motion_end_buffer_scale.set(motion_end_buffer_default)
        self.motion_end_buffer_scale.grid(row=row_index, column=1, sticky="we")
        ttk.Label(settings_frame, textvariable=self.motion_end_buffer_var).grid(row=row_index, column=2, padx=5,
                                                                                sticky="w")
        row_index += 1

        ttk.Label(settings_frame, text="Áttűnés hossza (mp):").grid(row=row_index, column=0, sticky="w", pady=2)
        self.crossfade_duration_var = StringVar(value="0,5")
        self.crossfade_duration_scale = Scale(settings_frame, from_=0, to=2, orient='horizontal', resolution=0.1,
                                              command=lambda v: self.update_scale_label(self.crossfade_duration_var, v,
                                                                                        decimals=1))
        self.crossfade_duration_scale.set(0.5)
        self.crossfade_duration_scale.grid(row=row_index, column=1, sticky="we")
        ttk.Label(settings_frame, textvariable=self.crossfade_duration_var).grid(row=row_index, column=2, padx=5,
                                                                                 sticky="w")
        row_index += 1



        # Fájllista panel
        filelist_frame = ttk.Frame(main_frame)
        filelist_frame.grid(row=1, column=0, columnspan=2, sticky="nsew", padx=5, pady=5)
        filelist_frame.grid_columnconfigure(0, weight=1)
        filelist_frame.grid_columnconfigure(1, weight=1)
        filelist_frame.grid_rowconfigure(5, weight=1)

        ttk.Label(filelist_frame, text="Bemeneti mappa:").grid(row=0, column=0, sticky="w")
        self.input_dir_var = StringVar()
        ttk.Entry(filelist_frame, textvariable=self.input_dir_var, state='readonly').grid(row=0, column=1, sticky="we",
                                                                                          padx=(5, 0))
        ttk.Button(filelist_frame, text="Tallózás...", command=self.select_input_directory).grid(row=0, column=2,
                                                                                                 sticky="e",
                                                                                                 padx=(5, 0))

        ttk.Label(filelist_frame, text="Kimeneti mappa:").grid(row=1, column=0, sticky="w", pady=(5, 0))
        self.output_dir_var = StringVar()
        ttk.Entry(filelist_frame, textvariable=self.output_dir_var, state='readonly').grid(row=1, column=1, sticky="we",
                                                                                           padx=(5, 0), pady=(5, 0))
        ttk.Button(filelist_frame, text="Tallózás...", command=self.select_output_folder).grid(row=1, column=2,
                                                                                               sticky="e",
                                                                                               padx=(5, 0), pady=(5, 0))

        ttk.Label(filelist_frame, text="Log mappa:").grid(row=2, column=0, sticky="w", pady=2)
        self.log_dir_var = StringVar()
        ttk.Entry(filelist_frame, textvariable=self.log_dir_var, state='readonly').grid(row=2, column=1, sticky="we",
                                                                                        padx=(5, 0))
        ttk.Button(filelist_frame, text="Tallózás...", command=self.select_log_folder).grid(row=2, column=2, sticky="e",
                                                                                            padx=(5, 0))

        control_and_delete_frame = ttk.Frame(filelist_frame)
        control_and_delete_frame.grid(row=3, column=0, columnspan=3, pady=(5, 10), sticky="w")

        self.delete_selected_button = ttk.Button(control_and_delete_frame, text="Kiválasztott törlése",
                                                 command=self.delete_selected_file, state="disabled")
        self.delete_selected_button.pack(side="left", padx=5)

        self.clear_all_button = ttk.Button(control_and_delete_frame, text="Összes törlése",
                                           command=self.clear_file_list, state="disabled")
        self.clear_all_button.pack(side="left", padx=5)

        self.start_button = ttk.Button(control_and_delete_frame, text="Feldolgozás indítása",
                                       command=self.toggle_processing)
        self.start_button.pack(side="left", padx=5)
        self.stop_button = ttk.Button(control_and_delete_frame, text="Feldolgozás leállítása",
                                      command=self.stop_processing, state="disabled")
        self.stop_button.pack(side="left", padx=5)

        self.resume_button = ttk.Button(control_and_delete_frame, text="Folytatás",
                                        command=self.resume_processing, state="disabled")
        self.resume_button.pack(side="left", padx=5)

        self.adjust_columns_button = ttk.Button(control_and_delete_frame, text="Oszlopszélességek állítása", command=self.adjust_column_widths)
        self.adjust_columns_button.pack(side="left", padx=5)

        self.skip_processed_var = StringVar(value="0")
        ttk.Checkbutton(control_and_delete_frame, text="Kihagyja a már feldolgozott videókat",
                        variable=self.skip_processed_var,
                        onvalue="1", offvalue="0").pack(side="left", padx=5)

        self.delete_temp_var = StringVar(value="1")
        ttk.Checkbutton(control_and_delete_frame, text="Ideiglenes fájlok törlése", variable=self.delete_temp_var,
                        onvalue="1",
                        offvalue="0").pack(side="left", padx=5)

        control_and_delete_frame_frame = ttk.Frame(filelist_frame)
        control_and_delete_frame_frame.grid(row=4, column=0, columnspan=3, pady=10, sticky="w")
        ttk.Button(control_and_delete_frame_frame, text="Beállítások betöltése", command=self.load_settings).pack(
            side="left",
            padx=5)
        ttk.Button(control_and_delete_frame_frame, text="Beállítások mentése", command=self.save_settings).pack(
            side="left",
            padx=5)

        columns = (
            "index", "file_in", "file_size_mb", "duration_sec",
            "output_file", "output_size_mb", "output_duration_sec", "compression_percent",
            "motion_duration_sec", "motion_percent",
            "processing_start_time", "processing_end_time", "processing_time",
            "status", "method", "profile", "input_full_path", "output_full_path",
            "script_name", "script_dir"
        )

        # --- Egyedi felbontás beállítások (áthelyezve ide) ---
        self.custom_resolution_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(control_and_delete_frame_frame, text="Egyedi kimeneti felbontás",
                        variable=self.custom_resolution_var).pack(side="left", padx=10)

        ttk.Label(control_and_delete_frame_frame, text="Szélesség:").pack(side="left", padx=(10, 2))
        self.output_width_entry = ttk.Entry(control_and_delete_frame_frame, width=6)
        self.output_width_entry.pack(side="left", padx=(0, 5))

        ttk.Label(control_and_delete_frame_frame, text="Magasság:").pack(side="left", padx=(10, 2))
        self.output_height_entry = ttk.Entry(control_and_delete_frame_frame, width=6)
        self.output_height_entry.pack(side="left", padx=(0, 10))

        # --- Új mező: Popup gyakoriság (fájlonként) ---
        ttk.Label(control_and_delete_frame_frame, text="Popup gyakoriság (fájlonként):").pack(side="left", padx=5)
        self.popup_interval_var = StringVar(value="10")
        self.popup_interval_entry = ttk.Entry(control_and_delete_frame_frame, textvariable=self.popup_interval_var,
                                              width=6)
        self.popup_interval_entry.pack(side="left", padx=5)

        self.file_tree = ttk.Treeview(filelist_frame, columns=columns, show="headings", height=15)

        self.file_tree.heading("index", text="Index")
        self.file_tree.heading("file_in", text="Bemeneti fájl")
        self.file_tree.heading("file_size_mb", text="Be MB")
        self.file_tree.heading("duration_sec", text="Be Idő")
        self.file_tree.heading("output_file", text="Ki név")
        self.file_tree.heading("output_size_mb", text="Ki MB")
        self.file_tree.heading("output_duration_sec", text="Ki Idő")
        self.file_tree.heading("compression_percent", text="Tömörítés (%)")
        self.file_tree.heading("motion_duration_sec", text="Mozgás Idő (s)")
        self.file_tree.heading("motion_percent", text="Mozgás (%)")
        self.file_tree.heading("processing_start_time", text="Feld. Kezdés")
        self.file_tree.heading("processing_end_time", text="Feld. Végzés")
        self.file_tree.heading("processing_time", text="Futásidő")
        self.file_tree.heading("status", text="Státusz")
        self.file_tree.heading("method", text="Eljárás")
        self.file_tree.heading("profile", text="Profil")
        self.file_tree.heading("input_full_path", text="Bemeneti útvonal")
        self.file_tree.heading("output_full_path", text="Kimeneti útvonal")
        self.file_tree.heading("script_name", text="Script név")
        self.file_tree.heading("script_dir", text="Script könyvtár")

        self.file_tree.column("index", width=50, stretch=False)
        self.file_tree.column("file_in", width=150)
        self.file_tree.column("file_size_mb", width=80)
        self.file_tree.column("duration_sec", width=80)
        self.file_tree.column("output_file", width=150)
        self.file_tree.column("output_size_mb", width=80)
        self.file_tree.column("output_duration_sec", width=80)
        self.file_tree.column("compression_percent", width=100)
        self.file_tree.column("motion_duration_sec", width=100)
        self.file_tree.column("motion_percent", width=80)
        self.file_tree.column("processing_start_time", width=100)
        self.file_tree.column("processing_end_time", width=100)
        self.file_tree.column("processing_time", width=80)
        self.file_tree.column("status", width=80)
        self.file_tree.column("method", width=80)
        self.file_tree.column("profile", width=120)
        self.file_tree.column("input_full_path", width=200)
        self.file_tree.column("output_full_path", width=200)
        self.file_tree.column("script_name", width=120)
        self.file_tree.column("script_dir", width=150)
        # --- Színezett tagek a profil oszlophoz (thr értékhez)
        self.file_tree.tag_configure("thr_auto", foreground="#0078D7")  # kék
        self.file_tree.tag_configure("thr_fix", foreground="#555555")  # szürke

        self.file_tree.grid(row=5, column=0, columnspan=3, sticky="nsew", pady=(0, 5))

        self.tree_scrollbar = ttk.Scrollbar(filelist_frame, orient="vertical", command=self.file_tree.yview)
        self.file_tree.configure(yscrollcommand=self.tree_scrollbar.set)
        self.tree_scrollbar.grid(row=5, column=3, sticky="ns")


        bottom_frame = ttk.Frame(main_frame, padding="10")
        bottom_frame.grid(row=2, column=0, columnspan=2, sticky="nsew")
        bottom_frame.grid_columnconfigure(0, weight=1)

        self.file_loading_progress_label_text = StringVar(value="Fájlok betöltése: 0/0")
        ttk.Label(bottom_frame, textvariable=self.file_loading_progress_label_text).grid(row=0, column=0, sticky="w",
                                                                                         pady=(5, 0))
        self.file_loading_progress = ttk.Progressbar(bottom_frame, mode='determinate')
        self.file_loading_progress.grid(row=1, column=0, sticky="ew")

        self.overall_progress_label_text = StringVar(value="Összes fájl feldolgozása: 0%")
        ttk.Label(bottom_frame, textvariable=self.overall_progress_label_text).grid(row=2, column=0, sticky="w",
                                                                                    pady=(5, 0))
        self.overall_progress = ttk.Progressbar(bottom_frame, length=600, mode='determinate')
        self.overall_progress.grid(row=3, column=0, sticky="ew")

        self.file_progress_label_text = StringVar(value="Aktuális fájl feldolgozása: 0%")
        ttk.Label(bottom_frame, textvariable=self.file_progress_label_text).grid(row=4, column=0, sticky="w",
                                                                                 pady=(5, 0))
        self.file_progress = ttk.Progressbar(bottom_frame, length=600, mode='determinate')
        self.file_progress.grid(row=5, column=0, sticky="ew")

        self.status = StringVar()
        self.status.set("Válassz ki egy bemeneti mappát és kimeneti mappát.")
        ttk.Label(bottom_frame, textvariable=self.status, wraplength=1380, justify="center").grid(row=6, column=0,sticky="ew", pady=(5, 0))

        # --- Valós idejű digitális kijelző (sebesség / idő / százalék) ---
        self.runtime_info_var = tk.StringVar(value="Feldolgozásra kész.")
        self.runtime_info_label = ttk.Label(
            self.root,
            textvariable=self.runtime_info_var,
            font=("Consolas", 9),
            anchor="center",
            foreground="#0078D7"
        )
        # a progress bar-ok és státuszsor alá tesszük
        self.runtime_info_label.pack(fill="x", pady=(4, 6))

        self.file_tree.bind("<<TreeviewSelect>>", self.on_file_select)

        self.load_settings()
        self.update_stats()
        self.check_for_interrupted_processing()
        self.update_resume_button_state()

    def update_scale_label(self, var, value, decimals=0):
        try:
            # Biztosítjuk, hogy a bemenet string legyen, és konvertáljuk angol tizedes tört formátumba
            value = str(value).replace(',', '.')
            float_value = float(value)
            # A magyar lokalizációhoz igazítjuk a kimeneti formátumot (vessző tizedes tört jelölőként)
            formatted_value = f"{float_value:.{decimals}f}".replace('.', ',')
            var.set(formatted_value)
        except ValueError as e:
            self.add_log_entry("ERROR", f"Hiba a csúszka értékének konvertálásakor: {e}")
            self.status.set("Hiba a csúszka értékének konvertálásakor.")
            var.set("N/A")

    def read_settings_file(path):
        with open(path, 'rb') as f:
            raw_data = f.read()
        encoding = chardet.detect(raw_data)['encoding'] or 'utf-8'
        try:
            return raw_data.decode(encoding)
        except Exception:
            # végső fallback
            return raw_data.decode('latin2', errors='replace')

    def setup_log_directory(self):
        """Napváltáskor új log könyvtár és logfile létrehozása"""

        today_str = datetime.now().strftime("%Y-%m-%d")

        # Fallback a script könyvtárába, ha nincs beállítva kimeneti vagy log mappa
        base_log_dir = self.log_dir_var.get() or self.output_folder or self.script_dir

        # dátumos mappa
        daily_dir = os.path.join(base_log_dir, today_str)

        # azon belül _Logok almappa
        log_dir = os.path.join(daily_dir, "_Logok")

        # A mappa létrehozása
        os.makedirs(log_dir, exist_ok=True)

        # Mentse el a log mappa elérési útját egy osztályszintű változóba
        self.log_dir = log_dir

        # runtime log fájl neve
        log_filename = f"runtime_log_{today_str}.txt"
        log_path = os.path.join(self.log_dir, log_filename)

        # ha már létezik, és folytatni akarjuk, akkor _2, _3, stb.
        counter = 2
        while os.path.exists(log_path):
            log_filename = f"runtime_log_{today_str}_{counter}.txt"
            log_path = os.path.join(self.log_dir, log_filename)
            counter += 1

        # állapot mentése
        self.current_log_date = today_str
        self.current_log_file = log_path
        self.log_line_number = 0

        # fejléc a log fájlban
        with open(self.current_log_file, "a", encoding="utf-8") as f:
            f.write(f"\nRuntime log - Kezdés: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Napi log könyvtár: {self.log_dir}\n")
            f.write(f"Script: {self.script_name}\n\n")

        self.add_log_entry("INFO", f"Log könyvtár beállítva: {self.log_dir}, logfájl: {log_filename}")

    def update_ffmpeg_threshold(self):
        """Spinbox érték változásakor frissíti a küszöb-feliratot és a logot."""
        try:
            new_thr = float(self.ffmpeg_threshold_var.get())
            self.current_threshold_value = new_thr
            if hasattr(self, "auto_threshold_value_var"):
                self.auto_threshold_value_var.set(f"Küszöb: {new_thr:.2f} (fix)")
            self.add_log_entry("INFO", f"FFmpeg küszöb beállítva: {new_thr:.2f}")
        except Exception as e:
            self.add_log_entry("ERROR", f"Küszöb frissítés hiba: {e}")



    def add_log_entry(self, level, message, process_log_path=None):
        """
        Naplóbejegyzés hozzáadása a futásidőhöz és a megfelelő logfájlba.

        Args:
            level (str): A napló szintje (pl. "INFO", "ERROR").
            message (str): A napló üzenet szövege.
            process_log_path (str, optional): A specifikus feldolgozási naplófájl elérési útja.
                                              Ha nincs megadva, a fő runtime logba ír.
        """
        # A fő, folyamatos sorszámozású log
        self.log_line_number += 1
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_entry_text = f"[{self.log_line_number:04d}] [{timestamp}] [{level}] {message}"

        # Hozzáadás a memóriához és a fő logfájlhoz
        self.runtime_log.append(log_entry_text)
        print(log_entry_text)
        if self.current_log_file:
            try:
                with open(self.current_log_file, "a", encoding="utf-8") as f:
                    f.write(log_entry_text + "\n")
            except Exception as e:
                print(f"Hiba a fő log mentése közben: {e}")

        # Speciális fájl-szintű naplózás (sorszámozás újraindul)
        if process_log_path:
            # Létrehozunk egy új számlálót a specifikus fájlhoz
            if not hasattr(self, 'process_log_line_number'):
                self.process_log_line_number = 0

            self.process_log_line_number += 1
            process_log_entry_text = f"[{self.process_log_line_number:04d}] [{timestamp}] [{level}] {message}"
            try:
                with open(process_log_path, "a", encoding="utf-8") as f:
                    f.write(process_log_entry_text + "\n")
            except Exception as e:
                print(f"Hiba a feldolgozási log mentése közben: {e}")

        # Naplóbejegyzés beszúrása a Treeview-ba
        try:
            if hasattr(self, 'log_text'):
                self.log_text.insert("", "end", values=(timestamp, level, message))
                self.log_text.yview_moveto(1)
        except Exception as e:
            print(f"Hiba a Treeview frissítése közben: {e}")

    def check_dependencies(self):
        missing_dependencies = []
        try:
            import cv2
        except ImportError:
            missing_dependencies.append("opencv-python")
        try:
            import pandas
        except ImportError:
            missing_dependencies.append("pandas")
        try:
            import numpy
        except ImportError:
            missing_dependencies.append("numpy")
        try:
            import pytesseract
        except ImportError:
            missing_dependencies.append("pytesseract")
        try:
            import fpdf
        except ImportError:
            missing_dependencies.append("fpdf")

        if missing_dependencies:
            messagebox.showwarning("Hiányzó függőségek",
                                   f"A következő Python könyvtárak hiányoznak: {', '.join(missing_dependencies)}. Kérlek telepítsd őket a 'pip install <könyvtárnév>' paranccsal a terminálban.")

        if not os.path.exists(ffmpeg_path):
            messagebox.showwarning("FFmpeg hiányzik",
                                   f"Az FFmpeg végrehajtható fájl nem található a megadott útvonalon: {ffmpeg_path}. Kérlek ellenőrizd, hogy telepítve van-e és a helyes útvonal van-e megadva.")

        try:
            subprocess.run([ffmpeg_path, "-version"], capture_output=True, text=True, check=True)
            self.add_log_entry("INFO", "FFmpeg sikeresen megtalálva")
        except Exception as e:
            self.add_log_entry("ERROR", f"FFmpeg nem található vagy hibás: {e}")
            self.status.set("Hiba: FFmpeg nem található. Ellenőrizd az elérési utat.")

        try:
            subprocess.run([pytesseract.pytesseract.tesseract_cmd, "--version"], capture_output=True, text=True,
                           check=True)
            self.add_log_entry("INFO", "Tesseract sikeresen megtalálva")
        except Exception as e:
            self.add_log_entry("ERROR", f"Tesseract nem található vagy hibás: {e}")
            self.status.set("Hiba: Tesseract nem található. Ellenőrizd az elérési utat.")

    def show_completion_popup(self, processed_count, total_duration, errors):
        """
        Megjelenít egy felugró ablakot a feldolgozás összefoglalójával.

        Args:
            processed_count (int): A sikeresen feldolgozott fájlok száma.
            total_duration (float): A teljes futási idő másodpercben.
            errors (list): Hibák listája.
        """
        # Hozzáadjuk a hangjelzést, ha a beállítás engedélyezi
        if self.play_sound:
            self.play_completion_sound()

        # Létrehozzuk a felugró ablakot
        popup = Toplevel(self.root)
        popup.title("Feldolgozás Befejezve")
        popup.geometry("400x300")

        # A pop-up ablak bezárásának letiltása
        #popup.protocol("WM_DELETE_WINDOW", self.disable_event)

        # Információk megjelenítése
        duration_formatted = str(timedelta(seconds=total_duration)).split('.')[0]

        Label(popup, text="Feldolgozás Befejezve!", font=("Helvetica", 16, "bold")).pack(pady=10)
        Label(popup, text=f"Feldolgozott fájlok: {processed_count}", font=("Helvetica", 12)).pack()
        Label(popup, text=f"Összes futási idő: {duration_formatted}", font=("Helvetica", 12)).pack()

        # Hibák megjelenítése, ha vannak
        if errors:
            Label(popup, text="A következő hibák léptek fel:", font=("Helvetica", 12, "bold"), fg="red").pack(
                pady=(15, 5))
            for error in errors:
                Label(popup, text=f"- {error}", wraplength=380, justify="left").pack(anchor="w", padx=10)

        # Hangjelzés ki/be kapcsoló gomb
        sound_text = "Hangjelzés kikapcsolása" if self.play_sound else "Hangjelzés bekapcsolása"
        sound_button = Button(popup, text=sound_text, command=lambda: self.toggle_sound(sound_button))
        sound_button.pack(pady=10)

        # Bezárás gomb
        Button(popup, text="Bezárás", command=popup.destroy).pack(pady=10)

    def play_completion_sound(self):
        """Lejátssza a hangjelzést a feldolgozás befejezésekor."""
        try:
            # Lejátsza a rendszerszintű "exclamation" hangot
            winsound.MessageBeep(winsound.MB_ICONEXCLAMATION)
        except Exception as e:
            self.add_log_entry("ERROR", f"Hiba a hangjelzés lejátszása közben: {e}")

    def save_processing_state(self):
        state = {
            'processed_files_count': self.processed_files_count,
            'processed_size_mb': self.processed_size_mb,
            'processed_duration_sec': self.processed_duration_sec,
            'current_file_index': self.current_file_index,
            'processed_files': [f['filepath'] for f in self.input_files[:self.processed_files_count]],
            'input_directory': self.input_directory,
            'output_folder': self.output_folder
        }
        with open(self.processing_state_file, 'w') as f:
            json.dump(state, f)

    def load_processing_state(self):
        if os.path.exists(self.processing_state_file):
            try:
                with open(self.processing_state_file, 'r') as f:
                    state = json.load(f)
                    self.processed_files_count = state.get('processed_files_count', 0)
                    self.processed_size_mb = state.get('processed_size_mb', 0)
                    self.processed_duration_sec = state.get('processed_duration_sec', 0)
                    self.current_file_index = state.get('current_file_index', 0)
                    self.input_directory = state.get('input_directory', "")
                    self.output_folder = state.get('output_folder', "")
                    self.add_log_entry("INFO", "Korábbi feldolgozási állapot betöltve.")
                    self.update_stats()
            except Exception as e:
                self.add_log_entry("ERROR", f"Hiba a feldolgozási állapot betöltése közben: {e}")
        else:
            self.add_log_entry("INFO", "Nincs korábbi feldolgozási állapot.")

    def clear_log(self):
        self.log_text.delete(*self.log_text.get_children())

    def play_success_sound(self):
        if self.play_sound:
            try:
                winsound.Beep(1000, 200)  # 1000 Hz, 200 ms
                winsound.Beep(1500, 200)  # 1500 Hz, 200 ms
            except:
                pass  # Csak Windows-on működik, más platformokon ne generáljon hibát

    def toggle_sound(self, button):
        """Ki- és bekapcsolja a hangjelzést és frissíti a gomb szövegét."""
        self.play_sound = not self.play_sound
        if self.play_sound:
            button.config(text="Hangjelzés kikapcsolása")
            self.add_log_entry("INFO", "Hangjelzés bekapcsolva.")
        else:
            button.config(text="Hangjelzés bekapcsolása")
            self.add_log_entry("INFO", "Hangjelzés kikapcsolva.")

    def disable_event(self):
        """Letiltja a pop-up ablak bezárását a 'X' gombbal."""
        pass

    def on_file_select(self, event):
        selected_items = self.file_tree.selection()
        if selected_items:
            self.delete_selected_button.config(state="normal")
        else:
            self.delete_selected_button.config(state="disabled")

    def toggle_resolution_fields(self):
        if self.custom_resolution_var.get() == "1":
            self.output_width_entry.config(state="normal")
            self.output_height_entry.config(state="normal")
        else:
            self.output_width_entry.config(state="disabled")
            self.output_height_entry.config(state="disabled")



    def select_input_directory(self):
        self.input_directory = filedialog.askdirectory()
        if self.input_directory:
            self.input_dir_var.set(self.input_directory)
            self.add_log_entry("INFO", f"Bemeneti mappa kiválasztva: {self.input_directory}")
            self.status.set(f"Bemeneti mappa kiválasztva: {self.input_directory}")
            self.save_settings()
            self.update_file_list()

    def select_output_folder(self):
        self.output_folder = filedialog.askdirectory()
        if self.output_folder:
            self.output_dir_var.set(self.output_folder)
            self.add_log_entry("INFO", f"Kimeneti mappa kiválasztva: {self.output_folder}")
            self.status.set(f"Kimeneti mappa kiválasztva: {self.output_folder}")
            self.save_settings()
            self.setup_log_directory()  # Hívja meg itt
            self.update_button_states()

    def select_log_folder(self):
        self.log_folder = filedialog.askdirectory()
        if self.log_folder:
            self.log_dir_var.set(self.log_folder)
            self.add_log_entry("INFO", f"Naplózási mappa kiválasztva: {self.log_folder}")
            self.status.set(f"Naplózási mappa kiválasztva: {self.log_folder}")
            self.save_settings()
            self.setup_log_directory()  # Hívja meg itt

    def update_file_list(self):
        self.input_files = []
        self.total_size_mb = 0
        self.total_duration_sec = 0
        self.file_tree.delete(*self.file_tree.get_children())
        self.tree_items = {}

        # Ellenőrizzük, hogy a bemeneti mappa létezik-e
        if not os.path.isdir(self.input_directory):
            self.add_log_entry("ERROR", f"Érvénytelen bemeneti mappa: {self.input_directory}")
            self.status.set(f"Érvénytelen bemeneti mappa: {self.input_directory}")
            self.file_loading_progress['value'] = 0
            self.file_loading_progress_label_text.set("Fájlok betöltése: 0/0")
            self.update_stats()
            self.update_button_states()
            return

        all_files = [f for f in os.listdir(self.input_directory) if
                     os.path.isfile(os.path.join(self.input_directory, f))]
        total_files = len(all_files)
        self.file_loading_progress['maximum'] = total_files
        self.file_loading_progress['value'] = 0
        loaded_count = 0

        for i, filename in enumerate(all_files):
            if self.stop_processing_flag:
                self.add_log_entry("WARNING", "Fájlbetöltés megszakítva a felhasználó által.")
                self.status.set("Fájlbetöltés megszakítva.")
                break

            file_path = os.path.join(self.input_directory, filename)
            if filename.lower().endswith(('.mp4', '.avi', '.mov', '.mkv')):
                try:
                    file_size_mb = os.path.getsize(file_path) / (1024 * 1024)  # MB-ra konvertálás
                    duration_sec = self.get_video_duration(file_path)
                    if duration_sec is not None:
                        self.input_files.append(file_path)
                        loaded_count += 1
                        self.total_size_mb += file_size_mb
                        self.total_duration_sec += duration_sec
                        item_id = self.file_tree.insert("", "end", values=(
                            f"{len(self.input_files):04d}",
                            filename,
                            f"{file_size_mb:.2f} MB",
                            self.format_time(duration_sec),
                            "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A",
                            "Várakozás", "N/A", "N/A", file_path, "N/A",
                            self.script_name, self.script_dir
                        ))
                        self.tree_items[file_path] = item_id
                        self.add_log_entry("INFO",
                                           f"Fájl betöltve: {filename}, Méret: {file_size_mb:.2f} MB, Időtartam: {self.format_time(duration_sec)}")
                    else:
                        self.add_log_entry("WARNING", f"Érvénytelen időtartam a fájlhoz: {filename}")
                except Exception as e:
                    self.add_log_entry("ERROR", f"Hiba a fájl betöltésekor: {filename} - {str(e)}")
            self.file_loading_progress['value'] = i + 1
            self.file_loading_progress_label_text.set(f"Fájlok betöltése: {loaded_count}/{total_files}")
            self.root.update_idletasks()

        self.file_loading_progress['value'] = total_files
        self.file_loading_progress_label_text.set(f"Fájlok betöltése: {loaded_count}/{total_files}")
        self.update_stats()
        self.update_button_states()
        if loaded_count == 0:
            self.add_log_entry("WARNING", f"Nincs betölthető videófájl a bemeneti mappában: {self.input_directory}")
            self.status.set("Nincs betölthető videófájl a bemeneti mappában.")
        else:
            self.add_log_entry("INFO", f"{loaded_count} videófájl betöltve a bemeneti mappából: {self.input_directory}")
            self.status.set(f"{loaded_count} videófájl betöltve.")

    def update_button_states(self):
        """Frissíti a gombok állapotát a fájlok és a kimeneti mappa megléte alapján."""
        has_files = bool(self.input_files)
        has_output_folder = bool(self.output_folder)

        self.start_button.config(state="normal" if has_files and has_output_folder else "disabled")
        self.delete_selected_button.config(state="normal" if has_files else "disabled")
        self.clear_all_button.config(state="normal" if has_files else "disabled")

        if self.file_tree.get_children():
            # A fa bejegyzéseinek van valamilyen tartalma, tehát van mit törölni
            self.clear_all_button.config(state="normal")
            self.delete_selected_button.config(state="normal")

    def clear_file_list(self):
        self.file_tree.delete(*self.file_tree.get_children())
        self.input_files = []
        self.total_size_mb = 0
        self.total_duration_sec = 0
        self.update_stats()
        self.status.set("Fájllista törölve.")
        self.add_log_entry("INFO", "Összes fájl törölve a listából.")
        self.update_button_states()

    def delete_selected_file(self):
        selected_items = self.file_tree.selection()
        if not selected_items:
            return

        for item in selected_items:
            values = self.file_tree.item(item, 'values')
            if not values:
                continue

            file_path_to_delete = values[16]

            if file_path_to_delete in self.input_files:
                file_size_mb = os.path.getsize(file_path_to_delete) / (1024 * 1024)
                file_duration_sec = self.get_video_duration(file_path_to_delete)

                self.input_files.remove(file_path_to_delete)
                if file_path_to_delete in self.tree_items:
                    del self.tree_items[file_path_to_delete]

                self.total_size_mb -= file_size_mb
                self.total_duration_sec -= file_duration_sec
                self.file_tree.delete(item)
                self.add_log_entry("INFO", f"Fájl törölve a listából: {os.path.basename(file_path_to_delete)}")
            else:
                self.add_log_entry("WARNING",
                                   f"Nem található a fájl a listában a törléshez: {os.path.basename(file_path_to_delete)}")

        self.update_stats()
        self.clear_all_button.config(state="normal" if self.input_files else "disabled")
        self.delete_selected_button.config(state="disabled")
        self.status.set("Kiválasztott fájl(ok) törölve.")

    def get_video_duration(self, file_path):
        try:
            cap = cv2.VideoCapture(file_path)
            if not cap.isOpened():
                self.add_log_entry("ERROR", f"A videó nem olvasható: {file_path}")
                return 0
            fps = cap.get(cv2.CAP_PROP_FPS)
            frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            cap.release()
            if fps > 0 and frame_count > 0:
                self.add_log_entry("INFO", f"Videó időtartama sikeresen lekérve: {file_path}")
                return frame_count / fps
            else:
                self.add_log_entry("ERROR", f"Érvénytelen FPS vagy képkockaszám: {file_path}")
                return 0
        except Exception as e:
            self.add_log_entry("ERROR", f"Hiba a videó időtartamának lekérésekor: {file_path} - {e}")
            return 0

    def get_video_info_ffprobe(self, file_path):
        try:
            cmd = [
                ffmpeg_path.replace('ffmpeg.exe', 'ffprobe.exe'),
                "-v", "error",
                "-show_entries", "format=duration,size",
                "-of", "default=noprint_wrappers=1:nokey=1",
                file_path
            ]
            result = subprocess.run(cmd, capture_output=True, text=True, check=True,
                                    creationflags=subprocess.CREATE_NO_WINDOW)
            output = result.stdout.strip().split('\n')
            duration_sec = float(output[0]) if len(output) > 0 and output[0].replace('.', '', 1).isdigit() else 0.0
            size_bytes = int(output[1]) if len(output) > 1 and output[1].isdigit() else 0

            return duration_sec, size_bytes / (1024 * 1024)
        except (subprocess.CalledProcessError, ValueError, IndexError) as e:
            self.add_log_entry("ERROR",
                               f"Hiba az ffprobe futtatása során a fájlhoz: {os.path.basename(file_path)}. Hiba: {e}")
            return 0.0, 0.0


    def update_stats(self):
        total_files = len(self.input_files)
        remaining_files = total_files - self.processed_files_count

        current_processed_input_size = self.processed_input_size_mb + self.current_file_progress_input_size
        current_processed_input_duration = self.processed_input_duration_sec + self.current_file_progress_input_duration

        remaining_size_mb = self.total_size_mb - current_processed_input_size
        remaining_duration_sec = self.total_duration_sec - current_processed_input_duration

        self.total_files_var.set(f"Összesen: {total_files}")
        self.processed_files_var.set(f"Feldolgozva: {self.processed_files_count}")
        self.remaining_files_var.set(f"Hátralévő: {remaining_files}")

        self.total_size_var.set(f"Összesen: {self.total_size_mb:.2f} MB")
        self.processed_size_var.set(f"Feldolgozva: {current_processed_input_size:.2f} MB")
        self.remaining_size_var.set(f"Hátralévő: {remaining_size_mb:.2f} MB")

        self.total_duration_var.set(f"Összesen: {self.format_time(self.total_duration_sec)}")
        self.processed_duration_var.set(f"Feldolgozva: {self.format_time(current_processed_input_duration)}")
        self.remaining_duration_var.set(f"Hátralévő: {self.format_time(remaining_duration_sec)}")

        elapsed_time = "N/A"
        remaining_time = "N/A"
        calculated_end_time = "N/A"
        total_processing_time = "N/A"

        if self.start_time:
            time_elapsed_seconds = (datetime.now() - self.start_time).total_seconds()
            elapsed_time = self.format_time(time_elapsed_seconds)
            total_processing_time = self.format_time(time_elapsed_seconds)

            if current_processed_input_duration > 0:
                estimated_total_time = (time_elapsed_seconds / current_processed_input_duration) * self.total_duration_sec
                self.calculated_end_time = datetime.now() + timedelta(seconds=(estimated_total_time - time_elapsed_seconds))
                remaining_seconds = (self.calculated_end_time - datetime.now()).total_seconds()
                remaining_time = self.format_time(remaining_seconds)
                calculated_end_time = self.calculated_end_time.strftime('%Y-%m-%d %H:%M:%S')

        self.start_time_var.set(f"Kezdés: {self.start_time.strftime('%H:%M:%S') if self.start_time else 'N/A'}")
        self.elapsed_time_var.set(f"Eltelt: {elapsed_time}")
        self.remaining_time_var.set(f"Hátralévő: {remaining_time}")
        self.end_time_var.set(f"Várható zárás: {calculated_end_time}")
        self.total_processing_time_var.set(f"Össz. futásidő: {total_processing_time}")

        overall_progress_percent = (
            current_processed_input_duration / self.total_duration_sec) * 100 if self.total_duration_sec > 0 else 0
        self.overall_progress['value'] = overall_progress_percent
        self.overall_progress_label_text.set(f"Összes fájl feldolgozása: {overall_progress_percent:.2f}%")

        # --- Kimeneti összesített értékek ---
        if self.processed_output_size_mb > 0:
            self.total_output_mb_var.set(f"Összes kimenő MB: {self.processed_output_size_mb:.2f} MB")
        else:
            self.total_output_mb_var.set("Összes kimenő MB: N/A")

        if self.processed_input_size_mb > 0:
            saving_percent = 100 * (1 - (self.processed_output_size_mb / self.processed_input_size_mb))
            self.saving_percent_var.set(f"Megtakarítás: {saving_percent:+.2f} %")
        else:
            self.saving_percent_var.set("Megtakarítás: N/A %")

        # --- Összesített eredmény frissítése ---
        if self.processed_input_size_mb > 0 and self.processed_output_size_mb > 0:
            total_in = self.processed_input_size_mb
            total_out = self.processed_output_size_mb
            saving_percent = 100 - (total_out / total_in) * 100

            self.summary_input_var.set(f"Be: {total_in:.2f} MB")
            self.summary_output_var.set(f"Ki: {total_out:.2f} MB")
            self.summary_saving_var.set(f"Megtakarítás: {saving_percent:+.2f} %")

            # --- Részletes statisztika ---
            try:
                smaller_files = getattr(self, "smaller_files_count", 0)
                larger_files = getattr(self, "larger_files_count", 0)
                smaller_diff = getattr(self, "smaller_total_diff_mb", 0.0)
                larger_diff = getattr(self, "larger_total_diff_mb", 0.0)
                total_files = getattr(self, "processed_files_count", 0)

                # Színezés
                if self.processed_input_size_mb > 0:
                    net_diff = self.processed_output_size_mb - self.processed_input_size_mb
                    if net_diff < 0:
                        color = "#0078D7"
                        emoji = "💾"
                    elif net_diff > 0:
                        color = "#C00000"
                        emoji = "📈"
                    else:
                        color = "#004080"
                        emoji = "⚪"
                else:
                    color = "#004080"
                    emoji = "⚪"

                # Futásidő kiszámítása
                runtime_str = "N/A"
                if getattr(self, "start_time", None):
                    elapsed = datetime.now() - self.start_time
                    total_seconds = int(elapsed.total_seconds())
                    h, rem = divmod(total_seconds, 3600)
                    m, s = divmod(rem, 60)
                    runtime_str = f"{h:02d}:{m:02d}:{s:02d}"

                detail_text = (
                    f"{emoji} Összes fájl: {total_files} db | "
                    f"💾 Kisebb: {smaller_files} db (−{smaller_diff:.2f} MB) | "
                    f"📈 Nagyobb: {larger_files} db (+{larger_diff:.2f} MB) | "
                    f"⏱️ Futásidő: {runtime_str}"
                )

                self.summary_files_detail_var.set(detail_text)
                self.summary_files_detail_label.configure(foreground=color)

            except Exception as e:
                self.summary_files_detail_var.set("Részletes statisztika: N/A")
                self.summary_files_detail_label.configure(foreground="#004080")
                self.add_log_entry("ERROR", f"Hiba a részletes statisztika frissítésénél: {e}")

            if saving_percent >= 0:
                self.summary_saving_label.configure(foreground="#007800")
            else:
                self.summary_saving_label.configure(foreground="#C00000")

        else:
            self.summary_input_var.set("Be: N/A MB")
            self.summary_output_var.set("Ki: N/A MB")
            self.summary_saving_var.set("Megtakarítás: N/A %")
            self.summary_saving_label.configure(foreground="#004080")

        # --- Megtakarítás / növekedés ikon és szín ---
        try:
            total_in = getattr(self, "processed_input_size_mb", 0.0)
            total_out = getattr(self, "processed_output_size_mb", 0.0)
            if total_in > 0:
                diff_mb = total_out - total_in
                change_percent = abs((diff_mb / total_in) * 100)
                if total_out < total_in:
                    icon, color, text = "💾", "blue", f"💾 Megtakarítás: {change_percent:.2f} %"
                elif total_out > total_in:
                    icon, color, text = "📈", "red", f"📈 Növekedés: {change_percent:.2f} %"
                else:
                    icon, color, text = "⚪", "black", "⚪ Nincs változás"
                self.summary_saving_var.set(text)
                self.summary_saving_label.config(foreground=color)
            else:
                self.summary_saving_var.set("Megtakarítás: N/A %")
                self.summary_saving_label.config(foreground="black")
        except Exception as e:
            self.add_log_entry("ERROR", f"Hiba a megtakarítás statisztika frissítésekor: {e}")

        self.root.update_idletasks()


    def format_time(self, seconds):
        if seconds < 0:
            seconds = 0
        minutes, seconds = divmod(seconds, 60)
        hours, minutes = divmod(minutes, 60)
        return f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"


    # --------------------------------------------------------------
    # PERIODIKUS GUI FRISSÍTÉS (1 mp-enként)
    # --------------------------------------------------------------
    def _periodic_gui_update(self):
        """Másodpercenként frissíti a GUI-t (statisztika + progress bar)."""
        try:
            if self.processing_thread and self.processing_thread.is_alive():
                self.update_stats()
                self.root.update_idletasks()
                self.root.after(1000, self._periodic_gui_update)
        except Exception as e:
            self.add_log_entry("ERROR", f"GUI frissítési hiba: {e}")




    def toggle_processing(self):
        self.add_log_entry("DEBUG", "toggle_processing metódus meghívva")

        # --- Ha már fut a feldolgozó szál, akkor szüneteltetés / folytatás ---
        if self.processing_thread and self.processing_thread.is_alive():
            self.add_log_entry("DEBUG", "Feldolgozó szál fut, szüneteltetés/folytatás")
            self.is_paused = not self.is_paused
            if self.is_paused:
                self.start_button.config(text="Folytatás")
                self.add_log_entry("INFO", "Feldolgozás szüneteltetve.")
                self.status.set("Feldolgozás szüneteltetve.")
            else:
                self.start_button.config(text="Szünet")
                self.add_log_entry("INFO", "Feldolgozás folytatva.")
                self.status.set("Feldolgozás folytatva.")
            return  # ne fusson tovább

        # --- Ha nem fut, új feldolgozás indul ---
        self.add_log_entry("DEBUG", "Új feldolgozás indítása")

        # Ellenőrzések
        if not self.input_directory or not self.output_folder:
            self.add_log_entry("WARNING", "Hiányzó bemeneti vagy kimeneti mappa")
            messagebox.showwarning("Hiányzó adatok", "Kérlek válaszd ki a bemeneti és kimeneti mappát.")
            return

        if not self.input_files:
            self.add_log_entry("WARNING", "Nincsenek videófájlok a bemeneti mappában")
            messagebox.showwarning("Hiányzó adatok", "A kiválasztott mappában nem található videófájl.")
            return

        # --- Gombok és állapotok előkészítése ---
        self.start_button.config(text="Szünet")
        self.stop_button.config(state="normal")
        self.resume_button.config(state="disabled")
        self.stop_processing_flag = False
        self.is_paused = False

        # --- Számlálók reset ---
        self.processed_files_count = 0
        self.processed_size_mb = 0
        self.processed_duration_sec = 0
        self.current_file_progress_duration = 0
        self.current_file_progress_size = 0
        self.log1_data = []
        self.log3_data = []

        # --- Feldolgozás szál létrehozása és indítása ---
        self.add_log_entry("DEBUG", "Szál létrehozása előtt")
        self.processing_thread = threading.Thread(target=self.process_all_files, daemon=True)
        self.add_log_entry("DEBUG", "Szál létrehozva, indítás")
        self.processing_thread.start()
        self.add_log_entry("DEBUG", "Szál elindítva")

        # --- GUI frissítés elindítása ---
        self._periodic_gui_update()  # <<< EZ biztosítja a valós idejű GUI-frissítést


    def stop_processing(self):
        self.stop_processing_flag = True
        self.is_paused = False
        if self.processing_thread and self.processing_thread.is_alive():
            self.add_log_entry("WARNING", "Feldolgozás leállítva, várjon a szál befejezésére.")
            self.status.set("Feldolgozás leállítva, várjon a szál befejezésére.")
        else:
            self.reset_state()

    def reset_state(self):
        """
        Feldolgozási állapot visszaállítása — csak akkor, ha a futás megszakadt vagy hibásan zárult.
        Ha a feldolgozás normálisan befejeződött (end_time már létezik), nem töröljük az adatokat.
        """

        # --- Ha a feldolgozás normálisan befejeződött, ne nullázzuk az adatokat ---
        if hasattr(self, "end_time") and self.end_time is not None:
            self.add_log_entry("DEBUG", "reset_state kihagyva, mert a feldolgozás befejeződött.")
            return

        self.start_button.config(text="Feldolgozás indítása")
        self.stop_button.config(state="disabled")
        self.resume_button.config(state="normal" if os.path.exists(self.processing_state_file) else "disabled")
        self.add_log_entry("INFO", "A feldolgozás leállítva.")
        self.status.set("A feldolgozás leállítva.")
        self.current_file_index = -1
        self.processed_files_count = 0
        self.processed_input_size_mb = 0
        self.processed_output_size_mb = 0
        self.processed_input_duration_sec = 0
        self.processed_output_duration_sec = 0
        self.current_file_progress_input_size = 0
        self.current_file_progress_input_duration = 0
        self.start_time = None
        self.end_time = None
        self.calculated_end_time = None
        self.update_stats()
        self.file_progress['value'] = 0
        self.overall_progress['value'] = 0
        self.overall_progress_label_text.set("Összes fájl feldolgozása: 0%")
        self.file_progress_label_text.set("Aktuális fájl feldolgozása: 0%")

    # ÚJ: Folytatás gomb logikája
    def resume_processing(self):
        if not os.path.exists(self.processing_state_file):
            messagebox.showinfo("Nincs folytatható feldolgozás",
                                "Nincs félbeszakadt feldolgozás, amit folytatni lehetne.")
            self.add_log_entry("INFO", "Nincs folytatható feldolgozás.")
            self.status.set("Nincs folytatható feldolgozás.")
            return

        with open(self.processing_state_file, 'r') as f:
            state = json.load(f)

        if state.get('status') != 'processing':
            messagebox.showinfo("Nincs folytatható feldolgozás", "A korábbi feldolgozás befejeződött vagy érvénytelen.")
            self.add_log_entry("INFO", "Nincs folytatható feldolgozás.")
            self.status.set("Nincs folytatható feldolgozás.")
            return

        interrupted_index = state.get('current_index', -1)
        interrupted_file = state.get('interrupted_file', 'N/A')
        interruption_time = state.get('interruption_time', datetime.now().isoformat())

        msg = (f"Félbeszakadt feldolgozást észleltünk!\n"
               f"Utolsó sikeres fájl index: {interrupted_index - 1 if interrupted_index > 0 else -1}\n"
               f"Megszakadt fájl: {interrupted_file}\n"
               f"Leállás ideje: {interruption_time}\n"
               f"Folytatni szeretné a feldolgozást a megszakadt fájltól?")
        if messagebox.askyesno("Folytatás félbeszakadt feldolgozásból", msg):
            self.current_file_index = interrupted_index - 1
            self.processed_files_count = interrupted_index
            self.start_button.config(text="Szünet")
            self.stop_button.config(state="normal")
            self.resume_button.config(state="disabled")
            self.stop_processing_flag = False
            self.is_paused = False
            self.processing_thread = threading.Thread(target=self.process_all_files, daemon=True)
            self.processing_thread.start()
            self.add_log_entry("INFO", f"Folytatás a félbeszakadt feldolgozásból az index {interrupted_index}-től.")
            self.status.set(f"Folytatás a félbeszakadt feldolgozásból.")
        else:
            os.remove(self.processing_state_file)
            self.add_log_entry("INFO", "Félbeszakadt feldolgozás nem folytatva.")
            self.status.set("Félbeszakadt feldolgozás nem folytatva.")
            self.resume_button.config(state="disabled")

    # ÚJ: A "Folytatás" gomb állapotának frissítése
    def update_resume_button_state(self):
        if os.path.exists(self.processing_state_file):
            self.resume_button.config(state="normal")
        else:
            self.resume_button.config(state="disabled")

    def format_runtime(self, seconds):
        minutes = int(seconds // 60)
        remaining_seconds = int(seconds % 60)
        return f"{minutes:02d}:{remaining_seconds:02d}"

    def generate_log_files(self):
        """Összefoglaló log fájlok (Log1, Log2, Log3) mentése a napi _Logok mappába"""

        import json
        from openpyxl import Workbook
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        # --- DejaVuSans font regisztrálása ---
        try:
            pdfmetrics.registerFont(TTFont("DejaVuSans", "F:/__Panel/fonts/DejaVuSans.ttf"))
            custom_style = ParagraphStyle("Custom", fontName="DejaVuSans", fontSize=10, leading=12)
        except Exception as e:
            # Ha valamiért nem találja a fontot, fallback
            styles = getSampleStyleSheet()
            custom_style = styles["Normal"]

        # Ha nincs aktuális log könyvtár, állítsuk be
        if not hasattr(self, "current_log_file") or not hasattr(self, "current_log_date"):
            self.setup_log_directory()

        # az aktuális runtime log mappa (_Logok)
        log_dir = os.path.dirname(self.current_log_file)
        today_str = self.current_log_date
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

        # ===================================================================
        # LOG1 - fájlok részletes feldolgozási listája
        # ===================================================================
        log1_filename = f"Log1-files-{today_str}_{timestamp}"

        # TXT
        with open(os.path.join(log_dir, f"{log1_filename}.txt"), "w", encoding="utf-8") as f:
            for entry in self.log1_data:
                for key, value in entry.items():
                    f.write(f"{key}: {value}\n")
                f.write("\n")

        # JSON
        with open(os.path.join(log_dir, f"{log1_filename}.json"), "w", encoding="utf-8") as f:
            json.dump(self.log1_data, f, ensure_ascii=False, indent=4)

        # XLSX
        wb = Workbook()
        ws = wb.active
        if self.log1_data:
            ws.append(list(self.log1_data[0].keys()))
            for entry in self.log1_data:
                ws.append(list(entry.values()))
        wb.save(os.path.join(log_dir, f"{log1_filename}.xlsx"))

        # PDF
        pdf = SimpleDocTemplate(os.path.join(log_dir, f"{log1_filename}.pdf"), pagesize=A4)
        story = []
        for entry in self.log1_data:
            for key, value in entry.items():
                story.append(Paragraph(f"<b>{key}:</b> {value}", custom_style))
            story.append(Spacer(1, 12))
        pdf.build(story)

        # ===================================================================
        # LOG2 - összesített statisztika
        # ===================================================================
        log2_filename = f"Log2-full-{today_str}_{timestamp}"

        total_files = self.processed_files_count
        total_input_size = self.processed_input_size_mb
        total_output_size = self.processed_output_size_mb
        total_time = (self.end_time - self.start_time).total_seconds() if hasattr(self, "end_time") else 0

        saving_mb = total_input_size - total_output_size
        saving_percent = 0
        if total_input_size > 0 and total_output_size > 0:
            saving_percent = 100 - (total_output_size / total_input_size) * 100

        # Szöveges összesítés generálása
        if total_input_size > 0:
            summary_text = (f"A teljes feldolgozás során "
                            f"{total_input_size:.2f} MB → {total_output_size:.2f} MB "
                            f"({saving_percent:+.2f} %) méretváltozás történt.")
        else:
            summary_text = "Nem áll rendelkezésre elegendő adat a megtakarítás kiszámításához."

        log2_data = {
            "01. A feldolgozott fájlok száma": f"{total_files} db",
            "02. A feldolgozott fájlok mérete megabájtban": f"{total_input_size:.2f} MB",
            "03. A feldolgozott fájlok időtartama": self.format_time(self.processed_input_duration_sec),
            "04. Összes kimenő fájlméret MB": f"{total_output_size:.2f} MB",
            "05. A feldolgozott fájlok megtakarítása MB": f"{saving_mb:.2f} MB",
            "06. A feldolgozott fájlok megtakarítása %": f"{saving_percent:+.2f} %",
            "07. Összesített eredmény szövegesen": summary_text,
            "08. A program indításának időbélyege": self.start_time.strftime('%Y-%m-%d %H:%M:%S'),
            "09. A program zárásának időbélyege": self.end_time.strftime('%Y-%m-%d %H:%M:%S') if hasattr(self,
                                                                                                         'end_time') else 'N/A',
            "10. A konvertálás kezdetének időbélyege": self.start_time.strftime('%H:%M:%S'),
            "11. A konvertálás zárásának időbélyege": self.end_time.strftime('%H:%M:%S') if hasattr(self,
                                                                                                    'end_time') else 'N/A',
            "12. A konvertálás futásideje összes file": self.format_time(total_time),
            "13. Bemeneti könyvtár": self.input_directory,
            "14. Kimeneti könyvtár": self.output_folder,
            "15. Script neve": self.script_name,
            "16. Script könyvtára": os.path.dirname(os.path.abspath(__file__)),
            "17. Feldolgozott fájlok listája": self.processed_files_list

        }
        thr_val = f"{getattr(self, 'current_threshold_value', 0.3):.2f}".replace(".", ",")
        thr_mode = "auto" if getattr(self, "auto_threshold_var", None) and self.auto_threshold_var.get() else "fix"
        log2_data["18. FFmpeg küszöb (threshold)"] = f"{thr_val} ({thr_mode})"

        # TXT
        with open(os.path.join(log_dir, f"{log2_filename}.txt"), "w", encoding="utf-8") as f:
            for k, v in log2_data.items():
                if isinstance(v, list):
                    f.write(f"{k}:\n")
                    for fname in v:
                        f.write(f"   {fname}\n")
                else:
                    f.write(f"{k}: {v}\n")

        # JSON
        with open(os.path.join(log_dir, f"{log2_filename}.json"), "w", encoding="utf-8") as f:
            json.dump(log2_data, f, ensure_ascii=False, indent=4)

        # XLSX
        wb = Workbook()
        ws = wb.active
        ws.append(["Kulcs", "Érték"])
        for k, v in log2_data.items():
            if isinstance(v, list):
                ws.append([k])  # fejléc (pl. "15. Feldolgozott fájlok listája")
                for fname in v:
                    ws.append(["", fname])  # fájlok külön sorban, 2. oszlopban
            else:
                ws.append([k, v])
        wb.save(os.path.join(log_dir, f"{log2_filename}.xlsx"))

        # PDF
        pdf = SimpleDocTemplate(os.path.join(log_dir, f"{log2_filename}.pdf"), pagesize=A4)
        story = []
        for k, v in log2_data.items():
            if isinstance(v, list):
                story.append(Paragraph(f"<b>{k}:</b>", custom_style))
                for fname in v:
                    story.append(Paragraph(f"   {fname}", custom_style))
            else:
                story.append(Paragraph(f"<b>{k}:</b> {v}", custom_style))
        pdf.build(story)

        # LOG3 - hibák listája (sorszámozva, fájlnévvel együtt)
        # ===================================================================
        log3_filename = f"Log3-errors-{today_str}_{timestamp}"

        log3_data_numbered = {}
        line_no = 1

        error_count = len(self.log3_data)
        log3_data_numbered[f"{line_no:02d}. Hibás fájlok száma"] = f"{error_count} db"
        line_no += 1

        if error_count == 0:
            log3_data_numbered[f"{line_no:02d}. Üzenet"] = "Nem történt hiba."
        else:
            for idx, entry in enumerate(self.log3_data, start=1):
                if isinstance(entry, dict):
                    file_name = entry.get("file", "Ismeretlen fájl")
                    error_msg = entry.get("error", "Ismeretlen hiba")
                else:
                    # ha csak sima string volt eltárolva
                    file_name = "Ismeretlen fájl"
                    error_msg = str(entry)
                log3_data_numbered[f"{line_no:02d}. Hiba {idx}"] = f"{file_name} – {error_msg}"
                line_no += 1

        # TXT
        with open(os.path.join(log_dir, f"{log3_filename}.txt"), "w", encoding="utf-8") as f:
            for key, value in log3_data_numbered.items():
                f.write(f"{key}: {value}\n")

        # JSON
        with open(os.path.join(log_dir, f"{log3_filename}.json"), "w", encoding="utf-8") as f:
            json.dump(log3_data_numbered, f, ensure_ascii=False, indent=4)

        # XLSX
        wb = Workbook()
        ws = wb.active
        ws.append(["Sorszám", "Hiba"])
        for key, value in log3_data_numbered.items():
            ws.append([key, value])
        wb.save(os.path.join(log_dir, f"{log3_filename}.xlsx"))

        # PDF
        pdf = SimpleDocTemplate(os.path.join(log_dir, f"{log3_filename}.pdf"), pagesize=A4)
        story = []
        for key, value in log3_data_numbered.items():
            story.append(Paragraph(f"<b>{key}:</b> {value}", custom_style))
        pdf.build(story)

        # dátumos főmappa (_Logok szülőkönyvtára)
        daily_dir = os.path.dirname(self.log_dir)

        # alkönyvtárak
        ocr_dir = os.path.join(daily_dir, "OCR")
        logs_dir = os.path.join(daily_dir, "LOGS")
        runtime_dir = os.path.join(daily_dir, "Runtime")

        os.makedirs(ocr_dir, exist_ok=True)
        os.makedirs(logs_dir, exist_ok=True)
        os.makedirs(runtime_dir, exist_ok=True)

        import glob, shutil

        # OCR képek mozgatása
        for f in glob.glob(os.path.join(self.log_dir, "debug_ocr_*.png")):
            shutil.move(f, os.path.join(ocr_dir, os.path.basename(f)))

        # LOG_Ch* fájlok mozgatása
        for f in glob.glob(os.path.join(self.log_dir, "LOG_Ch*.txt")):
            shutil.move(f, os.path.join(logs_dir, os.path.basename(f)))

        # runtime_log* fájlok mozgatása
        for f in glob.glob(os.path.join(self.log_dir, "runtime_log*.txt")):
            shutil.move(f, os.path.join(runtime_dir, os.path.basename(f)))

        # Log1, Log2, Log3 → fő dátum mappába
        for f in glob.glob(os.path.join(self.log_dir, "Log[123]*.*")):
            shutil.move(f, os.path.join(daily_dir, os.path.basename(f)))

        #-----------------------------------------------
        self.add_log_entry("INFO", f"Összes log mentve a mappába: {log_dir}")

    def time_to_seconds(self, time_str):
        time_str = time_str.rstrip('s')
        components = time_str.split(':')
        if len(components) == 3:
            h, m, s = map(int, components)
            return h * 3600 + m * 60 + s
        elif len(components) == 2:
            m, s = map(int, components)
            return m * 60 + s
        else:
            raise ValueError(f"Érvénytelen időformátum: {time_str}")

    def detect_motion_fast_ffmpeg(input_path: str, threshold: float = 0.3) -> list:
        """
        Gyors mozgásdetektálás FFmpeg scene detection segítségével (diagnosztikai naplózással).
        Visszatér [(start_sec, end_sec), ...] szegmensekkel.
        """

        import subprocess, re, time

        if not os.path.exists(input_path):
            return []

        start_dt = datetime.now()
        base = os.path.basename(input_path)
        log_prefix = f"[detect_motion_fast_ffmpeg] {base}"

        try:
            cmd = [
                FFMPEG_PATH,
                "-hide_banner",
                "-i", input_path,
                "-filter:v", f"select=gt(scene,{auto_thr}),showinfo,framestep=2",
                "-f", "null", "-"
            ]

            print(f"{log_prefix} → FFmpeg indítva threshold={threshold}")
            frame_times = []
            scene_scores = []

            # --- FFmpeg indítása ---
            proc = subprocess.Popen(
                cmd,
                stderr=subprocess.PIPE,
                stdout=subprocess.DEVNULL,
                universal_newlines=True,
                encoding="utf-8",
                errors="ignore"
            )

            # --- stderr soronként olvasás ---
            for line in proc.stderr:
                if "scene_score" in line and "pts_time:" in line:
                    try:
                        t_match = re.search(r"pts_time:([0-9\.]+)", line)
                        s_match = re.search(r"scene_score:([0-9\.]+)", line)
                        if t_match and s_match:
                            t = float(t_match.group(1))
                            s = float(s_match.group(1))
                            scene_scores.append((t, s))
                            # Naplózzuk a detektált értéket, ha van self.add_log_entry
                            try:
                                if hasattr(globals().get("app", None), "add_log_entry"):
                                    app.add_log_entry("DEBUG", f"Scene detect: t={t:.2f}s, score={s:.3f}")
                            except Exception:
                                pass
                            if s > threshold:
                                frame_times.append(t)
                    except Exception:
                        continue

            proc.wait(timeout=120)

            if not frame_times:
                print(f"{log_prefix} → Nem talált mozgást (scene_score < {threshold})")
                try:
                    if hasattr(globals().get("app", None), "add_log_entry"):
                        app.add_log_entry("INFO", f"FFmpeg: Nincs mozgás {base}-ben, fallback teljes videóra.")
                except Exception:
                    pass
                dur_full = get_video_duration_seconds(input_path) or 60.0
                return [(0.0, dur_full)]

            # --- Szegmenscsoportosítás ---
            segments = []
            seg_start = max(0.0, frame_times[0] - 1.0)

            for i in range(1, len(frame_times)):
                if frame_times[i] - frame_times[i - 1] > 5.0:
                    seg_end = frame_times[i - 1] + 1.0
                    segments.append((seg_start, seg_end))
                    seg_start = max(0.0, frame_times[i] - 1.0)

            segments.append((seg_start, frame_times[-1] + 1.0))

            end_dt = datetime.now()
            elapsed = (end_dt - start_dt).total_seconds()

            print(f"{log_prefix} → {len(segments)} szegmens detektálva, futásidő: {elapsed:.1f} s")
            try:
                if hasattr(globals().get("app", None), "add_log_entry"):
                    app.add_log_entry("INFO", f"FFmpeg mozgásdetektálás: {len(segments)} szegmens ({elapsed:.1f}s)")
            except Exception:
                pass

            return segments

        except Exception as e:
            print(f"{log_prefix} → Hiba: {e}")
            try:
                if hasattr(globals().get("app", None), "add_log_entry"):
                    app.add_log_entry("ERROR", f"FFmpeg detect hiba: {e}")
            except Exception:
                pass

            dur_full = get_video_duration_seconds(input_path) or 60.0
            return [(0.0, dur_full)]

    def process_single_file_fast(self, file_info: dict):
        """
        Egyetlen videófájl feldolgozása GYORS módban.
        - Mozgás-szegmensek detektálása (FFmpeg, fallback OpenCV-vel)
        - Szegmensek kivágása+tömörítése FFmpeg-preset/CRF szerint
        - Klippek összefűzése egy végső output fájlba
        - Statisztikák számítása (MB, tömörítés %, mozgás %)
        - TreeView és GUI státusz frissítése
        - Logolás

        Visszatér egy dict-tel (file_stats), amit a process_all_files() használ:
        {
            'status': 'Kész' vagy 'Hiba',
            'output_file': <kimeneti teljes fájl elérési út>,
            'motion_time_sec': float,
            'motion_pct': float,
            ...
        }
        """

        import os
        import shutil
        import subprocess
        import time


        # ------------------------------------------------------------------
        # 0) Előkészítés, bemeneti adatok
        # ------------------------------------------------------------------
        input_path = file_info.get("path", "")
        basename = os.path.basename(input_path)

        # file_stats MOSTANTÓL létezik, hogy ne dobjon hibát
        file_stats = {
            "status": "INIT",
            "output_file": None,
            "motion_segments_count": 0,
            "motion_time_sec": 0.0,
            "motion_pct": 0.0,
            "input_mb": 0.0,
            "output_mb": 0.0,
            "compression_pct": 0.0,
            "run_seconds": 0.0,
        }

        start_global_dt = datetime.now()
        start_global_ts = start_global_dt.strftime("%Y-%m-%d %H:%M:%S")

        # státusz + log a GUI-n
        self.add_log_entry("INFO", f"Feldolgozás indul: {basename} ({start_global_ts})")
        self.start_time = datetime.now()  # valós időmérés indítása GUI-hoz
        self.status.set(f"Feldolgozás: {basename}")
        self.root.update_idletasks()

        # próbáljuk megtalálni a TreeView sorát ehhez a fájlhoz
        try:
            item_id = self.tree_items.get(input_path)
        except Exception:
            item_id = None

        # frissítjük a TreeView állapotot, ha van
        try:
            if item_id is not None:
                current_vals = list(self.file_tree.item(item_id, "values"))
                # mezők: Index, Be név, Be MB, Be Idő, Ki név, Ki MB, Ki Idő, Tömörítés (%),
                #        Mozgás Idő (s), Mozgás (%), Feld. Kezdés, Feld. Végzés, Futásidő,
                #        Státusz, Eljárás, Profil, Bemeneti út, Kimeneti út, Script neve, Script dir
                # itt csak státuszt és eljárást frissítünk induláskor
                if len(current_vals) >= 20:
                    current_vals[13] = "Feldolgozás alatt"
                    current_vals[14] = "FFmpeg (gyors)"
                    current_vals[15] = self.preset_var.get() if hasattr(self, "preset_var") else "N/A"
                    self.file_tree.item(item_id, values=current_vals)
                    self.root.update_idletasks()
        except Exception:
            pass

        # ------------------------------------------------------------------
        # 1) Mozgás-szegmensek meghatározása (FFmpeg/OpenCV)
        # ------------------------------------------------------------------
        requested_mode = getattr(self, "detection_mode_var", None)
        if requested_mode is None:
            # ha valamiért nincs ilyen GUI elem (legacy kód), legyen default a gyors FFmpeg
            mode_selected = DETECTION_MODE_FAST
        else:
            mode_selected = requested_mode.get().strip().lower()

        segments = []
        used_mode = DETECTION_MODE_FAST
        motion_segments_count = 0
        total_motion_time = 0.0

        try:
            if mode_selected == DETECTION_MODE_FAST:
                # --- DEBUG bejegyzés: indítás ---
                self.add_log_entry("DEBUG", f"FFmpeg mozgásdetektálás indítása: {basename}")

                dm_result = detect_motion_fast_ffmpeg(input_path)

                # --- DEBUG bejegyzés: visszatérés ---
                self.add_log_entry("DEBUG", f"detect_motion_fast_ffmpeg() visszatért: {type(dm_result)}")

                # Biztonságos bontás
                if isinstance(dm_result, (list, tuple)) and len(dm_result) == 3:
                    segments, motion_segments_count, total_motion_time = dm_result
                else:
                    segments = dm_result
                    motion_segments_count = len(dm_result) if dm_result else 0
                    total_motion_time = sum(e - s for (s, e) in dm_result) if dm_result else 0.0

                # --- DEBUG bejegyzés: detektált adatok ---
                self.add_log_entry(
                    "DEBUG",
                    f"FFmpeg detektálás eredménye: {motion_segments_count} szegmens, {total_motion_time:.2f} mp"
                )

                # file_stats itt még NEM létezik → nem írunk bele!
                # A motion adatok majd később kerülnek a stats dict-be.

                # Ha az FFmpeg szerint nincs értelmes mozgás, fallback OpenCV
                if (not segments) or (
                        len(segments) == 1
                        and (segments[0][1] - segments[0][0]) < 1.0
                ):
                    self.add_log_entry(
                        "WARN",
                        f"Nincs értelmezhető FFmpeg mozgásdetektálás, fallback OpenCV: {basename}"
                    )
                    used_mode = DETECTION_MODE_DETAILED
                    segments = detect_motion_detailed_opencv(input_path)
                else:
                    used_mode = DETECTION_MODE_FAST

            else:
                # direkt a részletes módot választotta a felhasználó
                used_mode = DETECTION_MODE_DETAILED
                self.add_log_entry("DEBUG", f"Részletes OpenCV mód használata: {basename}")
                segments = detect_motion_detailed_opencv(input_path)

        except Exception as motion_err:
            self.add_log_entry("ERROR", f"Mozgásdetektálási hiba ({basename}): {motion_err}")
            dur_full = get_video_duration_seconds(input_path) or 60.0
            segments = [(0.0, dur_full)]
            motion_segments_count = 1
            total_motion_time = dur_full
            used_mode = "ERROR_FALLBACK_FULL"

        # ha még mindig nincs semmi (nagyon extrém eset), fallback teljes videóra
        if not segments:
            dur_full = get_video_duration_seconds(input_path) or 60.0
            segments = [(0.0, dur_full)]
            used_mode = "FALLBACK_FULL"

        self.add_log_entry(
            "INFO",
            f"{basename}: {len(segments)} detektált szegmens (mód: {used_mode})."
        )

        # ------------------------------------------------------------------
        # 2) Klippek legenerálása és tömörítése
        # ------------------------------------------------------------------
        try:
            crf_value = int(self.crf_scale.get())
        except Exception:
            crf_value = getattr(self, "crf_value", 28)

        ffmpeg_preset = (
            self.preset_var.get()
            if hasattr(self, "preset_var") else
            getattr(self, "ffmpeg_preset", "veryfast")
        )
        audio_bitrate = getattr(self, "audio_bitrate", "64k")

        # output mappa
        os.makedirs(self.output_folder, exist_ok=True)

        temp_clips = []         # ide gyűjtjük az egyes kivágott/tömörített klip fájlok elérési útját
        total_encode_run = 0.0  # az összes klip tömörítésének ideje összesen (mp)

        clip_index = 1

        # a belső GUI "Aktuális fájl feldolgozása: X%" frissítéséhez
        total_segments = max(1, len(segments))

        for (seg_start, seg_end) in segments:

            # Ha van saját névképződ, akkor:
            # clip_name = self.generate_output_filename(input_path, seg_start, seg_end, clip_index, used_mode)
            # Most a te jelenlegi verziódnak megfelelő fallback:
            clip_name = self.generate_output_filename(
                input_path,
                profile_suffix="_1080",
                index=clip_index
            )

            clip_path = os.path.join(self.output_folder, clip_name)

            self.add_log_entry(
                "INFO",
                f"Klip #{clip_index}: {sec_to_hms(seg_start)} → {sec_to_hms(seg_end)} → {clip_name}"
            )

            cut_info = cut_and_compress_segment(
                input_path=input_path,
                output_path=clip_path,
                start_sec=seg_start,
                end_sec=seg_end,
                crf=crf_value,
                preset=ffmpeg_preset,
                audio_bitrate=audio_bitrate,
            )

            total_encode_run += cut_info.get("run_seconds", 0.0)
            temp_clips.append(clip_path)

            # GUI helyzetfrissítés az adott fájlon belül
            try:
                pct_local = (clip_index / total_segments) * 100.0
                self.file_progress['value'] = pct_local
                self.file_progress_label_text.set(
                    f"Aktuális fájl feldolgozása: {pct_local:.0f}%"
                )
                self.root.update_idletasks()
            except Exception:
                pass

            # részstátusz visszairása a TreeView-be futás közben
            try:
                if item_id is not None:
                    vals_now = list(self.file_tree.item(item_id, "values"))
                    if len(vals_now) >= 20:
                        # Mozgó értékek:
                        vals_now[13] = f"Rész {clip_index}/{total_segments} kész"
                        vals_now[14] = "FFmpeg (gyors)"
                        vals_now[15] = (
                            self.preset_var.get() if hasattr(self, "preset_var") else "N/A"
                        )
                        # Kimeneti útvonal pillanatnyi klipre:
                        vals_now[17] = clip_path
                        self.file_tree.item(item_id, values=vals_now)
                        self.root.update_idletasks()
            except Exception:
                pass

            # felhasználói leállítás / szünet
            if getattr(self, "stop_processing_flag", False):
                self.add_log_entry(
                    "WARN",
                    "Felhasználó leállította a feldolgozást (közben)."
                )
                break

            while getattr(self, "is_paused", False) and not getattr(self, "stop_processing_flag", False):
                # Pause alatt engedjük a GUI frissülést
                self.root.update()

            if getattr(self, "stop_processing_flag", False):
                break

            clip_index += 1

        # ------------------------------------------------------------------
        # 3) Klippek összefűzése egyetlen végső MP4-be
        # ------------------------------------------------------------------
        merged_output_name = f"{os.path.splitext(basename)[0]}_merged.mp4"
        merged_output_path = os.path.join(self.output_folder, merged_output_name)

        if len(temp_clips) > 1:
            concat_list_file = os.path.join(self.output_folder, "_concat_list.txt")
            with open(concat_list_file, "w", encoding="utf-8") as f:
                for p in temp_clips:
                    # ffmpeg concat file formátum – idézőjelezve
                    f.write(f"file '{p}'\n")

            cmd_concat = [
                ffmpeg_path,  # <-- figyelem: globális ffmpeg_path, nem FFMPEG_PATH
                "-f", "concat",
                "-safe", "0",
                "-i", concat_list_file,
                "-c", "copy",
                "-y",
                merged_output_path
            ]

            concat_proc = subprocess.run(
                cmd_concat,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                encoding="utf-8",
                errors="ignore"
            )

            if concat_proc.returncode != 0:
                self.add_log_entry(
                    "WARNING",
                    f"Összefűzés figyelmeztetés ({basename}), ffmpeg rc={concat_proc.returncode}"
                )

        elif len(temp_clips) == 1:
            # csak 1 klip volt -> az lesz a végső output
            try:
                shutil.move(temp_clips[0], merged_output_path)
            except Exception:
                # ha nem tudjuk mozgatni, próbáljuk másolni
                try:
                    shutil.copy2(temp_clips[0], merged_output_path)
                except Exception as copy_err:
                    self.add_log_entry(
                        "ERROR",
                        f"Nem sikerült áthelyezni az egyklipes outputot: {copy_err}"
                    )
        else:
            # Ha semmi klip nem jött létre (stop közben leállítva?)
            merged_output_path = None

        # ------------------------------------------------------------------
        # 4) Statisztikák számítása
        # ------------------------------------------------------------------
        stats = {
            "status": "Hiba",
            "output_file": merged_output_path,
            "motion_time_sec": 0.0,
            "motion_pct": 0.0,
        }

        if merged_output_path and os.path.exists(merged_output_path):
            calc = calculate_file_stats(
                input_path=input_path,
                merged_output_path=merged_output_path,
                motion_segments=segments,
                total_run_seconds=total_encode_run
            )

            # illesszük be, amit a későbbi log/TreeView kód használ
            stats.update({
                "status": "Kész",
                "input_mb":        calc.get("input_mb", 0.0),
                "output_mb":       calc.get("output_mb", 0.0),
                "compression_pct": calc.get("compression_pct", 0.0),
                "motion_time_sec": calc.get("motion_time_sec", 0.0),
                "motion_pct":      calc.get("motion_pct", 0.0),
                "run_seconds":     calc.get("run_seconds", 0.0),
                "output_file":     merged_output_path,
            })
        else:
            # Nincs érvényes output => Hiba
            stats["status"] = "Hiba"
            stats["output_file"] = None

        # ------------------------------------------------------------------
        # 5) TreeView végső frissítés erre a fájlra
        # ------------------------------------------------------------------
        end_global_dt = datetime.now()
        end_global_ts = end_global_dt.strftime("%Y-%m-%d %H:%M:%S")
        elapsed_total = (end_global_dt - start_global_dt).total_seconds()
        elapsed_hms = sec_to_hms(elapsed_total)

        try:
            if item_id is not None:
                vals_done = list(self.file_tree.item(item_id, "values"))
                if len(vals_done) >= 20:
                    # Index (0), Be név (1), Be MB (2), Be Idő (3),
                    # Ki név (4), Ki MB (5), Ki Idő (6),
                    # Tömörítés % (7),
                    # Mozgás Idő (s) (8),
                    # Mozgás % (9),
                    # Feld. Kezdés (10),
                    # Feld. Végzés (11),
                    # Futásidő (12),
                    # Státusz (13),
                    # Eljárás (14),
                    # Profil (15),
                    # Bemenet (16),
                    # Kimenet (17),
                    # Script név (18),
                    # Script dir (19)

                    # Ki név / Ki MB / Ki Idő
                    if stats["output_file"]:
                        out_base = os.path.basename(stats["output_file"])
                        try:
                            out_size_mb = os.path.getsize(stats["output_file"]) / (1024 * 1024)
                        except Exception:
                            out_size_mb = 0.0
                        out_dur_sec = self.get_video_duration(stats["output_file"])
                    else:
                        out_base = "N/A"
                        out_size_mb = 0.0
                        out_dur_sec = 0.0

                    # compression %
                    if out_size_mb > 0:
                        try:
                            in_mb = os.path.getsize(input_path) / (1024 * 1024)
                        except Exception:
                            in_mb = 0.0
                        if in_mb > 0:
                            compr_local = 100 - (out_size_mb / in_mb) * 100
                        else:
                            compr_local = 0.0
                    else:
                        compr_local = 0.0

                    vals_done[4]  = out_base
                    vals_done[5]  = f"{out_size_mb:.2f} MB" if out_size_mb > 0 else "N/A"
                    vals_done[6]  = self.format_time(out_dur_sec) if out_dur_sec > 0 else "N/A"
                    vals_done[7]  = f"{compr_local:.2f} %" if compr_local > 0 else "N/A"

                    vals_done[8]  = (
                        f"{stats.get('motion_time_sec',0.0):.2f}"
                        if stats.get("motion_time_sec",0.0) > 0
                        else "N/A"
                    )
                    vals_done[9]  = (
                        f"{stats.get('motion_pct',0.0):.2f}"
                        if stats.get("motion_pct",0.0) > 0
                        else "N/A"
                    )

                    vals_done[10] = start_global_dt.strftime("%H:%M:%S")
                    vals_done[11] = end_global_dt.strftime("%H:%M:%S")
                    vals_done[12] = elapsed_hms
                    vals_done[13] = stats.get("status","N/A")
                    vals_done[14] = "FFmpeg (gyors)"

                    # Profil oszlop: Preset + automatikus FFmpeg küszöbérték
                    preset_text = (
                        self.preset_var.get()
                        if hasattr(self, "preset_var")
                        else "N/A"
                    )
                    thr_value = getattr(self, "current_threshold_value", 0.3)
                    vals_done[15] = f"{preset_text}, thr={thr_value:.2f}"

                    vals_done[16] = input_path
                    vals_done[17] = stats["output_file"] if stats["output_file"] else "N/A"
                    vals_done[18] = self.script_name
                    vals_done[19] = self.script_dir

                    # --- Színezés (auto / fix threshold szerint)
                    thr_mode_auto = (
                        getattr(self, "auto_threshold_var", None)
                        and self.auto_threshold_var.get()
                    )
                    if thr_mode_auto:
                        self.file_tree.item(item_id, values=vals_done, tags=("thr_auto",))
                    else:
                        self.file_tree.item(item_id, values=vals_done, tags=("thr_fix",))

                    # végső értékek visszaírása
                    self.file_tree.item(item_id, values=vals_done)
                    self.root.update_idletasks()
        except Exception as tree_err:
            self.add_log_entry("WARNING", f"Nem sikerült TreeView frissítés: {tree_err}")

        # ------------------------------------------------------------------
        # 6) végállapot log / status
        # ------------------------------------------------------------------
        if stats.get("status") == "Kész":
            self.add_log_entry(
                "INFO",
                f"KÉSZ: {basename} -> "
                f"{os.path.basename(stats['output_file']) if stats['output_file'] else 'N/A'} "
                f"(mozgás {stats.get('motion_pct',0.0):.2f}%, futás {elapsed_hms})"
            )
            self.status.set(f"Kész: {basename} ({elapsed_hms})")
        else:
            self.add_log_entry(
                "ERROR",
                f"HIBA: {basename} feldolgozása sikertelen."
            )
            self.status.set(f"Hiba: {basename}")

        # ------------------------------------------------------------------
        # 7) stat-panel biztonsági frissítés
        # ------------------------------------------------------------------
        try:
            self.update_stats()
        except Exception:
            pass

        # ------------------------------------------------------------------
        # 8) visszatérés a process_all_files() részére
        # ------------------------------------------------------------------
        return {
            "status": stats.get("status", "N/A"),
            "output_file": stats.get("output_file"),
            "motion_time_sec": stats.get("motion_time_sec", 0.0),
            "motion_pct": stats.get("motion_pct", 0.0),
            "input_mb": stats.get("input_mb", 0.0),
            "output_mb": stats.get("output_mb", 0.0),
            "compression_pct": stats.get("compression_pct", 0.0),
            "run_seconds": stats.get("run_seconds", 0.0)
        }


    def generate_output_filename(self, input_path, profile_suffix="_1080", index=1):
        """
        Kimeneti fájlnév generálása (Ch / other / egyéb típushoz).
        Példa: Ch2_20200818141909.mp4 → Ch2_2020-08-18-Sze_14-19-09_1080_01.mp4
        """

        import os
        from datetime import datetime

        basename = os.path.basename(input_path)
        name, _ = os.path.splitext(basename)
        folder = os.path.basename(os.path.dirname(input_path))

        # --- 1. Ha "Ch" kezdetű fájl ---
        if name.startswith("Ch") and len(name) >= 17:
            cam = name[:3]
            date_part = name[4:12]  # 20200818
            time_part = name[12:18] # 141909

            dt_str = f"{date_part[:4]}-{date_part[4:6]}-{date_part[6:8]}"
            hh, mm, ss = time_part[:2], time_part[2:4], time_part[4:6]
            weekday = self.get_hungarian_weekday(dt_str)  # pl. "Sze"

            out_name = f"{cam}_{dt_str}-{weekday}_{hh}-{mm}-{ss}{profile_suffix}_{index:02d}.mp4"
            return out_name

        # --- 2. "other" típus (mappa alapján) ---
        elif "Y" in folder and "M" in folder and "D" in folder and "H" in folder:
            try:
                y = folder.split("Y")[0]
                m = folder.split("Y")[1].split("M")[0]
                d = folder.split("M")[1].split("D")[0]
                h = folder.split("D")[1].split("H")[0]
                weekday = self.get_hungarian_weekday(f"{y}-{m}-{d}")
                out_name = f"{y}-{m}-{d}-{weekday}_{h}-{name.replace('M','-').replace('S','')}{profile_suffix}_{index:02d}.mp4"
                return out_name
            except Exception:
                pass

        # --- 3. Egyéb típus ---
        out_name = f"{name}{profile_suffix}-{index:02d}.mp4"
        return out_name

    def get_hungarian_weekday(self, date_str):
        """
        YYYY-MM-DD formátumú dátumból visszaadja a magyar rövidített napnevet.
        """
        import datetime
        days = ["H", "K", "Sze", "Cs", "P", "Szo", "V"]
        try:
            dt = datetime.datetime.strptime(date_str, "%Y-%m-%d")
            return days[dt.weekday()]
        except Exception:
            return "?"

    def process_single_file_fast(self, file_info: dict):
        """
        Egyetlen videófájl feldolgozása GYORS módban.
        - Mozgás-szegmensek detektálása (FFmpeg, fallback OpenCV-vel)
        - Szegmensek kivágása+tömörítése FFmpeg-preset/CRF szerint
        - Klippek összefűzése egy végső output fájlba
        - Statisztikák számítása (MB, tömörítés %, mozgás %)
        - TreeView és GUI státusz frissítése
        - Logolás

        Visszatér egy dict-tel (file_stats), amit a process_all_files() használ:
        {
            'status': 'Kész' vagy 'Hiba',
            'output_file': <kimeneti teljes fájl elérési út>,
            'motion_time_sec': float,
            'motion_pct': float,
            ...
        }
        """

        import os
        import shutil
        import subprocess
        import time

        # ------------------------------------------------------------------
        # 0) Előkészítés, bemeneti adatok
        # ------------------------------------------------------------------
        input_path = file_info.get("path", "")
        basename = os.path.basename(input_path)

        # file_stats MOSTANTÓL létezik, hogy ne dobjon hibát
        file_stats = {
            "status": "INIT",
            "output_file": None,
            "motion_segments_count": 0,
            "motion_time_sec": 0.0,
            "motion_pct": 0.0,
            "input_mb": 0.0,
            "output_mb": 0.0,
            "compression_pct": 0.0,
            "run_seconds": 0.0,
        }

        start_global_dt = datetime.now()
        start_global_ts = start_global_dt.strftime("%Y-%m-%d %H:%M:%S")

        # státusz + log a GUI-n
        self.add_log_entry("INFO", f"Feldolgozás indul: {basename} ({start_global_ts})")
        self.status.set(f"Feldolgozás: {basename}")
        self.root.update_idletasks()

        # próbáljuk megtalálni a TreeView sorát ehhez a fájlhoz
        try:
            item_id = self.tree_items.get(input_path)
        except Exception:
            item_id = None

        # frissítjük a TreeView állapotot, ha van
        try:
            if item_id is not None:
                current_vals = list(self.file_tree.item(item_id, "values"))
                # mezők: Index, Be név, Be MB, Be Idő, Ki név, Ki MB, Ki Idő, Tömörítés (%),
                #        Mozgás Idő (s), Mozgás (%), Feld. Kezdés, Feld. Végzés, Futásidő,
                #        Státusz, Eljárás, Profil, Bemeneti út, Kimeneti út, Script neve, Script dir
                # itt csak státuszt és eljárást frissítünk induláskor
                if len(current_vals) >= 20:
                    current_vals[13] = "Feldolgozás alatt"
                    current_vals[14] = "FFmpeg (gyors)"
                    current_vals[15] = self.preset_var.get() if hasattr(self, "preset_var") else "N/A"
                    self.file_tree.item(item_id, values=current_vals)
                    self.root.update_idletasks()
        except Exception:
            pass

        # ------------------------------------------------------------------
        # 1) Mozgás-szegmensek meghatározása (FFmpeg/OpenCV)
        # ------------------------------------------------------------------
        requested_mode = getattr(self, "detection_mode_var", None)
        if requested_mode is None:
            # ha valamiért nincs ilyen GUI elem (legacy kód), legyen default a gyors FFmpeg
            mode_selected = DETECTION_MODE_FAST
        else:
            mode_selected = requested_mode.get().strip().lower()

        segments = []
        used_mode = DETECTION_MODE_FAST
        motion_segments_count = 0
        total_motion_time = 0.0

        try:
            if mode_selected == DETECTION_MODE_FAST:
                # --- DEBUG bejegyzés: indítás ---
                self.add_log_entry("DEBUG", f"FFmpeg mozgásdetektálás indítása: {basename}")

                dm_result = detect_motion_fast_ffmpeg(input_path)

                # --- DEBUG bejegyzés: visszatérés ---
                self.add_log_entry("DEBUG", f"detect_motion_fast_ffmpeg() visszatért: {type(dm_result)}")

                # Biztonságos bontás
                if isinstance(dm_result, (list, tuple)) and len(dm_result) == 3:
                    segments, motion_segments_count, total_motion_time = dm_result
                else:
                    segments = dm_result
                    motion_segments_count = len(dm_result) if dm_result else 0
                    total_motion_time = sum(e - s for (s, e) in dm_result) if dm_result else 0.0

                # --- DEBUG bejegyzés: detektált adatok ---
                self.add_log_entry(
                    "DEBUG",
                    f"FFmpeg detektálás eredménye: {motion_segments_count} szegmens, {total_motion_time:.2f} mp"
                )

                # file_stats itt még NEM létezik → nem írunk bele!
                # A motion adatok majd később kerülnek a stats dict-be.

                # Ha az FFmpeg szerint nincs értelmes mozgás, fallback OpenCV
                if (not segments) or (
                        len(segments) == 1
                        and (segments[0][1] - segments[0][0]) < 1.0
                ):
                    self.add_log_entry(
                        "WARN",
                        f"Nincs értelmezhető FFmpeg mozgásdetektálás, fallback OpenCV: {basename}"
                    )
                    used_mode = DETECTION_MODE_DETAILED
                    segments = detect_motion_detailed_opencv(input_path)
                else:
                    used_mode = DETECTION_MODE_FAST

            else:
                # direkt a részletes módot választotta a felhasználó
                used_mode = DETECTION_MODE_DETAILED
                self.add_log_entry("DEBUG", f"Részletes OpenCV mód használata: {basename}")
                segments = detect_motion_detailed_opencv(input_path)

        except Exception as motion_err:
            self.add_log_entry("ERROR", f"Mozgásdetektálási hiba ({basename}): {motion_err}")
            dur_full = get_video_duration_seconds(input_path) or 60.0
            segments = [(0.0, dur_full)]
            motion_segments_count = 1
            total_motion_time = dur_full
            used_mode = "ERROR_FALLBACK_FULL"

        # ha még mindig nincs semmi (nagyon extrém eset), fallback teljes videóra
        if not segments:
            dur_full = get_video_duration_seconds(input_path) or 60.0
            segments = [(0.0, dur_full)]
            used_mode = "FALLBACK_FULL"

        self.add_log_entry(
            "INFO",
            f"{basename}: {len(segments)} detektált szegmens (mód: {used_mode})."
        )

        # ------------------------------------------------------------------
        # 2) Klippek legenerálása és tömörítése
        # ------------------------------------------------------------------
        try:
            crf_value = int(self.crf_scale.get())
        except Exception:
            crf_value = getattr(self, "crf_value", 28)

        ffmpeg_preset = (
            self.preset_var.get()
            if hasattr(self, "preset_var") else
            getattr(self, "ffmpeg_preset", "veryfast")
        )
        audio_bitrate = getattr(self, "audio_bitrate", "64k")

        # output mappa
        os.makedirs(self.output_folder, exist_ok=True)

        temp_clips = []  # ide gyűjtjük az egyes kivágott/tömörített klip fájlok elérési útját
        total_encode_run = 0.0  # az összes klip tömörítésének ideje összesen (mp)

        clip_index = 1

        # a belső GUI "Aktuális fájl feldolgozása: X%" frissítéséhez
        total_segments = max(1, len(segments))

        for (seg_start, seg_end) in segments:

            # Ha van saját névképződ, akkor:
            # clip_name = self.generate_output_filename(input_path, seg_start, seg_end, clip_index, used_mode)
            # Most a te jelenlegi verziódnak megfelelő fallback:
            clip_name = self.generate_output_filename(
                input_path,
                profile_suffix="_1080",
                index=clip_index
            )

            clip_path = os.path.join(self.output_folder, clip_name)

            self.add_log_entry(
                "INFO",
                f"Klip #{clip_index}: {sec_to_hms(seg_start)} → {sec_to_hms(seg_end)} → {clip_name}"
            )

            cut_info = cut_and_compress_segment(
                input_path=input_path,
                output_path=clip_path,
                start_sec=seg_start,
                end_sec=seg_end,
                crf=crf_value,
                preset=ffmpeg_preset,
                audio_bitrate=audio_bitrate,
            )

            total_encode_run += cut_info.get("run_seconds", 0.0)
            temp_clips.append(clip_path)

            # GUI helyzetfrissítés az adott fájlon belül
            try:
                pct_local = (clip_index / total_segments) * 100.0
                self.file_progress['value'] = pct_local
                self.file_progress_label_text.set(
                    f"Aktuális fájl feldolgozása: {pct_local:.0f}%"
                )
                self.root.update_idletasks()
            except Exception:
                pass

            # részstátusz visszairása a TreeView-be futás közben
            try:
                if item_id is not None:
                    vals_now = list(self.file_tree.item(item_id, "values"))
                    if len(vals_now) >= 20:
                        # Mozgó értékek:
                        vals_now[13] = f"Rész {clip_index}/{total_segments} kész"
                        vals_now[14] = "FFmpeg (gyors)"
                        vals_now[15] = (
                            self.preset_var.get() if hasattr(self, "preset_var") else "N/A"
                        )
                        # Kimeneti útvonal pillanatnyi klipre:
                        vals_now[17] = clip_path
                        self.file_tree.item(item_id, values=vals_now)
                        self.root.update_idletasks()
            except Exception:
                pass

            # felhasználói leállítás / szünet
            if getattr(self, "stop_processing_flag", False):
                self.add_log_entry(
                    "WARN",
                    "Felhasználó leállította a feldolgozást (közben)."
                )
                break

            while getattr(self, "is_paused", False) and not getattr(self, "stop_processing_flag", False):
                # Pause alatt engedjük a GUI frissülést
                self.root.update()

            if getattr(self, "stop_processing_flag", False):
                break

            clip_index += 1

        # ------------------------------------------------------------------
        # 3) Klippek összefűzése egyetlen végső MP4-be
        # ------------------------------------------------------------------
        merged_output_name = f"{os.path.splitext(basename)[0]}_merged.mp4"
        merged_output_path = os.path.join(self.output_folder, merged_output_name)

        if len(temp_clips) > 1:
            concat_list_file = os.path.join(self.output_folder, "_concat_list.txt")
            with open(concat_list_file, "w", encoding="utf-8") as f:
                for p in temp_clips:
                    # ffmpeg concat file formátum – idézőjelezve
                    f.write(f"file '{p}'\n")

            cmd_concat = [
                ffmpeg_path,  # <-- figyelem: globális ffmpeg_path, nem FFMPEG_PATH
                "-f", "concat",
                "-safe", "0",
                "-i", concat_list_file,
                "-c", "copy",
                "-y",
                merged_output_path
            ]

            concat_proc = subprocess.run(
                cmd_concat,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                encoding="utf-8",
                errors="ignore"
            )

            if concat_proc.returncode != 0:
                self.add_log_entry(
                    "WARNING",
                    f"Összefűzés figyelmeztetés ({basename}), ffmpeg rc={concat_proc.returncode}"
                )

        elif len(temp_clips) == 1:
            # csak 1 klip volt -> az lesz a végső output
            try:
                shutil.move(temp_clips[0], merged_output_path)
            except Exception:
                # ha nem tudjuk mozgatni, próbáljuk másolni
                try:
                    shutil.copy2(temp_clips[0], merged_output_path)
                except Exception as copy_err:
                    self.add_log_entry(
                        "ERROR",
                        f"Nem sikerült áthelyezni az egyklipes outputot: {copy_err}"
                    )
        else:
            # Ha semmi klip nem jött létre (stop közben leállítva?)
            merged_output_path = None

        # ------------------------------------------------------------------
        # 4) Statisztikák számítása
        # ------------------------------------------------------------------
        stats = {
            "status": "Hiba",
            "output_file": merged_output_path,
            "motion_time_sec": 0.0,
            "motion_pct": 0.0,
        }

        if merged_output_path and os.path.exists(merged_output_path):
            calc = calculate_file_stats(
                input_path=input_path,
                merged_output_path=merged_output_path,
                motion_segments=segments,
                total_run_seconds=total_encode_run
            )

            # illesszük be, amit a későbbi log/TreeView kód használ
            stats.update({
                "status": "Kész",
                "input_mb": calc.get("input_mb", 0.0),
                "output_mb": calc.get("output_mb", 0.0),
                "compression_pct": calc.get("compression_pct", 0.0),
                "motion_time_sec": calc.get("motion_time_sec", 0.0),
                "motion_pct": calc.get("motion_pct", 0.0),
                "run_seconds": calc.get("run_seconds", 0.0),
                "output_file": merged_output_path,
            })
        else:
            # Nincs érvényes output => Hiba
            stats["status"] = "Hiba"
            stats["output_file"] = None

        # ------------------------------------------------------------------
        # 5) TreeView végső frissítés erre a fájlra
        # ------------------------------------------------------------------
        end_global_dt = datetime.now()
        end_global_ts = end_global_dt.strftime("%Y-%m-%d %H:%M:%S")
        elapsed_total = (end_global_dt - start_global_dt).total_seconds()
        elapsed_hms = sec_to_hms(elapsed_total)

        try:
            if item_id is not None:
                vals_done = list(self.file_tree.item(item_id, "values"))
                if len(vals_done) >= 20:
                    # Index (0), Be név (1), Be MB (2), Be Idő (3),
                    # Ki név (4), Ki MB (5), Ki Idő (6),
                    # Tömörítés % (7),
                    # Mozgás Idő (s) (8),
                    # Mozgás % (9),
                    # Feld. Kezdés (10),
                    # Feld. Végzés (11),
                    # Futásidő (12),
                    # Státusz (13),
                    # Eljárás (14),
                    # Profil (15),
                    # Bemenet (16),
                    # Kimenet (17),
                    # Script név (18),
                    # Script dir (19)

                    # Ki név / Ki MB / Ki Idő
                    if stats["output_file"]:
                        out_base = os.path.basename(stats["output_file"])
                        try:
                            out_size_mb = os.path.getsize(stats["output_file"]) / (1024 * 1024)
                        except Exception:
                            out_size_mb = 0.0
                        out_dur_sec = self.get_video_duration(stats["output_file"])
                    else:
                        out_base = "N/A"
                        out_size_mb = 0.0
                        out_dur_sec = 0.0

                    # compression %
                    if out_size_mb > 0:
                        try:
                            in_mb = os.path.getsize(input_path) / (1024 * 1024)
                        except Exception:
                            in_mb = 0.0
                        if in_mb > 0:
                            compr_local = 100 - (out_size_mb / in_mb) * 100
                        else:
                            compr_local = 0.0
                    else:
                        compr_local = 0.0

                    vals_done[4] = out_base
                    vals_done[5] = f"{out_size_mb:.2f} MB" if out_size_mb > 0 else "N/A"
                    vals_done[6] = self.format_time(out_dur_sec) if out_dur_sec > 0 else "N/A"
                    vals_done[7] = f"{compr_local:.2f} %" if compr_local > 0 else "N/A"

                    vals_done[8] = (
                        f"{stats.get('motion_time_sec', 0.0):.2f}"
                        if stats.get("motion_time_sec", 0.0) > 0
                        else "N/A"
                    )
                    vals_done[9] = (
                        f"{stats.get('motion_pct', 0.0):.2f}"
                        if stats.get("motion_pct", 0.0) > 0
                        else "N/A"
                    )

                    vals_done[10] = start_global_dt.strftime("%H:%M:%S")
                    vals_done[11] = end_global_dt.strftime("%H:%M:%S")
                    vals_done[12] = elapsed_hms
                    vals_done[13] = stats.get("status", "N/A")
                    vals_done[14] = "FFmpeg (gyors)"

                    # Profil oszlop: Preset + automatikus FFmpeg küszöbérték
                    preset_text = (
                        self.preset_var.get()
                        if hasattr(self, "preset_var")
                        else "N/A"
                    )
                    thr_value = getattr(self, "current_threshold_value", 0.3)
                    vals_done[15] = f"{preset_text}, thr={thr_value:.2f}"

                    vals_done[16] = input_path
                    vals_done[17] = stats["output_file"] if stats["output_file"] else "N/A"
                    vals_done[18] = self.script_name
                    vals_done[19] = self.script_dir

                    # --- Színezés (auto / fix threshold szerint)
                    thr_mode_auto = (
                            getattr(self, "auto_threshold_var", None)
                            and self.auto_threshold_var.get()
                    )
                    if thr_mode_auto:
                        self.file_tree.item(item_id, values=vals_done, tags=("thr_auto",))
                    else:
                        self.file_tree.item(item_id, values=vals_done, tags=("thr_fix",))

                    # végső értékek visszaírása
                    self.file_tree.item(item_id, values=vals_done)
                    self.root.update_idletasks()
        except Exception as tree_err:
            self.add_log_entry("WARNING", f"Nem sikerült TreeView frissítés: {tree_err}")

        # ------------------------------------------------------------------
        # 6) végállapot log / status
        # ------------------------------------------------------------------
        if stats.get("status") == "Kész":
            self.add_log_entry(
                "INFO",
                f"KÉSZ: {basename} -> "
                f"{os.path.basename(stats['output_file']) if stats['output_file'] else 'N/A'} "
                f"(mozgás {stats.get('motion_pct', 0.0):.2f}%, futás {elapsed_hms})"
            )
            self.status.set(f"Kész: {basename} ({elapsed_hms})")
        else:
            self.add_log_entry(
                "ERROR",
                f"HIBA: {basename} feldolgozása sikertelen."
            )
            self.status.set(f"Hiba: {basename}")

        # ------------------------------------------------------------------
        # 7) stat-panel biztonsági frissítés
        # ------------------------------------------------------------------
        try:
            self.update_stats()
        except Exception:
            pass

        # ------------------------------------------------------------------
        # 8) visszatérés a process_all_files() részére
        # ------------------------------------------------------------------
        return {
            "status": stats.get("status", "N/A"),
            "output_file": stats.get("output_file"),
            "motion_time_sec": stats.get("motion_time_sec", 0.0),
            "motion_pct": stats.get("motion_pct", 0.0),
            "input_mb": stats.get("input_mb", 0.0),
            "output_mb": stats.get("output_mb", 0.0),
            "compression_pct": stats.get("compression_pct", 0.0),
            "run_seconds": stats.get("run_seconds", 0.0)
        }

    def process_all_files(self):
        """
        Teljes feldolgozási ciklus (GYORSÍTOTT VERZIÓ).
        - inicializálja a számlálókat, statisztikát, logot
        - opcionálisan kihagyja a már feldolgozott fájlokat
        - végigmegy az összes videón (FFmpeg-alapú gyors motor)
        - minden N-edik fájl után popupot dob, az N-et a self.popup_interval_var tartalmazza
        """

        # --- 0) Reset minden futás előtt ---
        self.processed_files_list = []
        self.processed_input_size_mb = 0.0
        self.processed_output_size_mb = 0.0
        self.processed_input_duration_sec = 0.0
        self.processed_output_duration_sec = 0.0
        self.processed_files_count = 0
        self.errors = []
        self.current_file_index = -1
        self.smaller_files_count = 0
        self.larger_files_count = 0
        self.smaller_total_diff_mb = 0.0
        self.larger_total_diff_mb = 0.0

        # Új log mappa / napi _Logok előkészítése
        self.setup_log_directory()

        # Időadatok induláskor
        self.start_time = datetime.now()
        self.calculated_end_time = self.start_time + timedelta(seconds=self.total_duration_sec)
        self.end_time = None

        # GUI stat reset
        self.update_stats()
        self.overall_progress['value'] = 0
        self.overall_progress_label_text.set("Összes fájl feldolgozása: 0%")
        self.file_progress['value'] = 0
        self.file_progress_label_text.set("Aktuális fájl feldolgozása: 0%")
        self.root.update_idletasks()

        # --- 1) Fájllista és opcionális kihagyás már feldolgozott output alapján ---
        total_files = len(self.input_files)

        if self.skip_processed_var.get() == "1":
            skipped_files = []
            for file_path in self.input_files[:]:
                base_name = os.path.splitext(os.path.basename(file_path))[0]
                date_part = base_name[4:12]
                time_part = base_name[12:18]
                formatted_date = f"{date_part[:4]}-{date_part[4:6]}-{date_part[6:8]}"
                formatted_time = f"{time_part[:2]}-{time_part[2:4]}-{time_part[4:6]}"

                cap = cv2.VideoCapture(file_path)
                if not cap.isOpened():
                    self.add_log_entry("ERROR", f"Hiba: Nem sikerült megnyitni a videót: {file_path}")
                    continue

                frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
                cap.set(cv2.CAP_PROP_POS_FRAMES, frame_count - 1)
                ret, last_frame = cap.read()
                end_time_str = None

                if ret:
                    try:
                        crop = last_frame[0:150, 0:1000]
                        gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
                        _, thresh = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY)
                        for psm in [6, 7, 8, 10, 11, 13]:
                            text = pytesseract.image_to_string(
                                thresh,
                                config=f'--psm {psm} --oem 3 -c tessedit_char_whitelist=0123456789: '
                            )
                            text = ' '.join(text.strip().split())
                            match = re.search(r'(\d{2}):(\d{2}):(\d{2})', text)
                            if match:
                                hours, minutes, seconds = map(int, match.groups())
                                if 0 <= hours <= 23 and 0 <= minutes <= 59 and 0 <= seconds <= 59:
                                    end_time_str = f"{hours:02d}-{minutes:02d}-{seconds:02d}"
                                    break
                    except Exception as ocr_err:
                        self.add_log_entry("WARNING", f"OCR hiba (end_time): {ocr_err}")

                cap.release()

                if not end_time_str:
                    file_duration_sec = self.get_video_duration(file_path)
                    total_seconds = int(file_duration_sec)
                    hh = total_seconds // 3600
                    mm = (total_seconds % 3600) // 60
                    ss = total_seconds % 60
                    end_time_str = f"{hh:02d}-{mm:02d}-{ss:02d}"

                potential_output_base = f"{base_name[:3]}_{formatted_date}-P_{formatted_time}__{end_time_str}_1920"

                counter = 0
                found_existing = False
                while counter <= 100:
                    if counter == 0:
                        candidate = os.path.join(self.output_folder, f"{potential_output_base}.mp4")
                    else:
                        candidate = os.path.join(self.output_folder, f"{potential_output_base}_{counter:02d}.mp4")

                    if os.path.exists(candidate):
                        if file_path in self.input_files:
                            self.input_files.remove(file_path)
                        if file_path in self.tree_items:
                            self.file_tree.delete(self.tree_items[file_path])
                            del self.tree_items[file_path]
                        skipped_files.append(os.path.basename(file_path))
                        self.add_log_entry("INFO", f"Kihagyva (már feldolgozva): {os.path.basename(file_path)}")
                        found_existing = True
                        break
                    counter += 1

            if skipped_files:
                self.status.set(f"{len(skipped_files)} fájl kihagyva (már feldolgozva).")
            total_files = len(self.input_files)

            if total_files == 0:
                self.status.set("Nincs új fájl feldolgozásra. Minden kihagyva.")
                self.start_button.config(text="Feldolgozás indítása")
                self.stop_button.config(state="disabled")
                self.overall_progress['value'] = 100
                self.overall_progress_label_text.set("Összes fájl feldolgozása: 100%")

                try:
                    self.generate_log_files()
                except Exception as e:
                    self.add_log_entry("ERROR", f"Hiba a logfájlok generálása közben: {e}")

                self.show_summary_popup(
                    f"✅ Nincs új fájl feldolgozásra!\n\n"
                    f"Kihagyott fájlok száma: {len(skipped_files)}\n"
                    f"Log mappa: {self.log_dir}",
                    self.log_dir
                )
                self.root.update_idletasks()
                return

        # --- 2) Per-run process log inicializálása ---
        timestamp = self.start_time.strftime("%Y-%m-%d_%H-%M-%S")
        log_dir = getattr(self, "log_dir", None)
        if not log_dir:
            self.setup_log_directory()
            log_dir = self.log_dir

        process_log_path = os.path.join(log_dir, f"LOG_Ch1_{timestamp}.txt")
        with open(process_log_path, "w", encoding="utf-8") as lf:
            lf.write(f"Process log - Kezdés: {timestamp}\n")
            lf.write(f"Bemeneti könyvtár: {self.input_directory}\n")
            lf.write(f"Kimeneti könyvtár: {self.output_folder}\n")
            lf.write(f"Összes fájl: {total_files}\n")

        self.add_log_entry("INFO", f"Process log létrehozva: {process_log_path}")

        # --- 3) Fő feldolgozó ciklus (itt kezdődik majd a gyorsított mag) ---
        start_index = self.current_file_index + 1
        for i in range(start_index, total_files):
            # Stop kérés?
            if self.stop_processing_flag:
                self.add_log_entry("WARNING", "Feldolgozás megszakítva a felhasználó által.")
                self.status.set("Feldolgozás megszakítva.")
                break

            # Paused?
            while self.is_paused and not self.stop_processing_flag:
                self.root.update()

            if self.stop_processing_flag:
                break

                # ✅ GUI folyamatos frissítése minden fájl előtt
                self.root.after(0, self.update_stats)
                self.root.update_idletasks()

            # Index és aktuális fájl
            self.current_file_index = i
            file_path = self.input_files[i]
            input_file_basename = os.path.basename(file_path)

            # Állapot mentés (folytatás támogatás)
            self.save_processing_state(i, "processing", file_path)

            # TreeView sor frissítés "feldolgozás alatt" állapotra
            item_id = self.tree_items[file_path]
            self.file_tree.item(
                item_id,
                values=(
                    f"{i + 1:04d}",
                    input_file_basename,
                    self.file_tree.item(item_id, 'values')[2],
                    self.file_tree.item(item_id, 'values')[3],
                    "N/A", "N/A", "N/A", "N/A", "N/A", "N/A",
                    "Feldolgozás...", "N/A", "N/A",
                    "Feldolgozás alatt",
                    "FFmpeg",
                    self.preset_var.get(),
                    file_path,
                    "N/A",
                    self.script_name,
                    self.script_dir
                )
            )
            self.root.update_idletasks()

            self.add_log_entry("INFO", f"Feldolgozás alatt: {input_file_basename}")
            self.status.set(f"Feldolgozás alatt: {input_file_basename}")

            # GUI progress (összes)
            overall_pct = ((i + 1) / total_files) * 100
            self.overall_progress_label_text.set(f"Összes fájl feldolgozása: {int(overall_pct)}%")
            self.overall_progress['value'] = overall_pct

            # GUI progress (aktuális fájl)
            self.file_progress_label_text.set("Aktuális fájl feldolgozása: 0%")
            self.file_progress['value'] = 0
            self.root.update_idletasks()

            # Egyedi fájl statok induló értékei
            file_duration_sec = self.get_video_duration(file_path)
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
            processing_start_time = datetime.now()

            # =============================================================
            #  === LÉNYEGI FELDOLGOZÁS (ÚJ: FFmpeg gyors motor integrálva) ===
            # =============================================================

            try:
                file_stats = self.process_single_file_fast({"path": file_path})
                # a gyors motor visszaadja a motion_duration_sec és motion_pct értékeket
                if not file_stats:
                    file_stats = {"status": "Hiba", "output_file": None,
                                  "motion_duration": 0, "motion_percent": 0.0}
            except Exception as e:
                self.add_log_entry("ERROR", f"Hiba a gyors feldolgozás közben: {e}")
                file_stats = {"status": "Hiba", "output_file": None,
                              "motion_duration": 0, "motion_percent": 0.0}

            # =============================================================
            #  === A feldolgozás után minden marad a régi módon (statisztika) ===
            # =============================================================

            processing_end_time = datetime.now()
            processing_time = (processing_end_time - processing_start_time).total_seconds()
            minutes, seconds = divmod(int(processing_time), 60)
            formatted_processing_time = f"{minutes:02d}:{seconds:02d}"

            # Kimeneti méret / hossz
            if file_stats.get('output_file') and os.path.exists(file_stats['output_file']):
                output_file_path = file_stats['output_file']
                output_size_mb = os.path.getsize(output_file_path) / (1024 * 1024)
                output_duration_sec = self.get_video_duration(output_file_path)
            else:
                output_file_path = os.path.join(self.output_folder, f"failed_{input_file_basename}")
                output_size_mb = 0.0
                output_duration_sec = 0.0
                if not file_stats.get('status'):
                    file_stats['status'] = "Hiba"
                file_stats['output_file'] = output_file_path

            # --- Fájlméret-változás statisztika gyűjtése ---
            diff_mb = output_size_mb - file_size_mb
            if diff_mb < 0:
                self.smaller_files_count += 1
                self.smaller_total_diff_mb += abs(diff_mb)
            elif diff_mb > 0:
                self.larger_files_count += 1
                self.larger_total_diff_mb += diff_mb

            # Tömörítési arány
            if file_size_mb > 0 and output_size_mb > 0:
                compression_percent = 100 - (output_size_mb / file_size_mb) * 100
            elif file_size_mb > 0 and output_size_mb == 0:
                compression_percent = 100
            else:
                compression_percent = 0

            # LOG1 sor felvétele (változatlanul, de az új statokat használjuk, ha elérhetők)
            self.log1_data.append({
                "01. Index": f"{i + 1:04d}",
                "02. Be név": input_file_basename,
                "03. Be MB": f"{file_size_mb:.2f} MB",
                "04. Be Idő": self.format_time(file_duration_sec),
                "05. Ki név": os.path.basename(output_file_path) if output_file_path else "N/A",
                "06. Ki MB": f"{output_size_mb:.2f} MB" if output_size_mb > 0 else "N/A",
                "07. Ki Idő": self.format_time(output_duration_sec) if output_duration_sec > 0 else "N/A",
                "08. Tömörítés (%)": f"{compression_percent:.2f} %" if compression_percent > 0 else "N/A",
                "09. Mozgás Idő (s)": f"{file_stats.get('motion_time_sec', 0.0):.2f}" if file_stats.get('motion_time_sec', 0.0) > 0 else "N/A",
                "10. Mozgás (%)": f"{file_stats.get('motion_pct', 0.0):.2f}" if file_stats.get('motion_pct', 0.0) > 0 else "N/A",
                "10/b. Mozgás összesen": (
                    f"{sec_to_hms(file_stats.get('motion_time_sec', 0.0))} "
                    f"({int(file_stats.get('motion_segments_count', 0))} rész)"
                    if file_stats.get("motion_time_sec", 0.0) > 0 else "N/A"
                ),

                "11. Feld. Kezdés": processing_start_time.strftime("%H:%M:%S"),
                "12. Feld. Végzés": processing_end_time.strftime("%H:%M:%S"),
                "13. Futásidő": formatted_processing_time,
                "14. Státusz": file_stats.get('status', 'N/A'),
                "15. Eljárás": "FFmpeg (gyors)",
                "16. Profil": ", ".join([
                    f"Preset={self.preset_var.get()}",
                    f"CRF={self.crf_scale.get()}",
                    f"MinMotion={self.min_motion_duration_scale.get():.1f}s",
                    f"Idle={self.idle_duration_scale.get():.1f}s",
                    f"PreBuffer={self.pre_motion_buffer_scale.get():.1f}s",
                    f"PostBuffer={self.motion_end_buffer_scale.get():.1f}s",
                    f"Pixel={int(float(self.pixel_threshold_scale.get()))}",
                    f"Áttűnés={self.crossfade_duration_scale.get():.1f}s"
                    f"FFmpegThr={getattr(self, 'current_threshold_value', 0.3):.2f}"

                ]),
                "17. Bemeneti útvonal": file_path,
                "18. Kimeneti útvonal": output_file_path if output_file_path else "N/A",
                "19. Script neve": self.script_name,
                "20. Script könyvtára": self.script_dir
            })

            # TreeView frissítése "Kész" állapotra (változatlan)
            self.file_tree.item(
                item_id,
                values=(
                    f"{i + 1:04d}",
                    input_file_basename,
                    f"{file_size_mb:.2f} MB",
                    self.format_time(file_duration_sec),
                    os.path.basename(output_file_path) if output_file_path else "N/A",
                    f"{output_size_mb:.2f} MB" if output_size_mb > 0 else "N/A",
                    self.format_time(output_duration_sec) if output_duration_sec > 0 else "N/A",
                    f"{compression_percent:.2f} %" if compression_percent > 0 else "N/A",
                    f"{file_stats.get('motion_time_sec', 0.0):.2f}" if file_stats.get('motion_time_sec', 0.0) > 0 else "N/A",
                    f"{file_stats.get('motion_pct', 0.0):.2f}" if file_stats.get('motion_pct', 0.0) > 0 else "N/A",
                    processing_start_time.strftime("%H:%M:%S"),
                    processing_end_time.strftime("%H:%M:%S"),
                    formatted_processing_time,
                    file_stats.get('status', 'N/A'),
                    "FFmpeg (gyors)",
                    self.preset_var.get(),
                    file_path,
                    output_file_path if output_file_path else "N/A",
                    self.script_name,
                    self.script_dir
                )
            )
            self.root.update_idletasks()
            # Összesített számlálók frissítése
            self.processed_files_count += 1
            self.processed_input_size_mb += file_size_mb
            self.processed_output_size_mb += output_size_mb
            self.processed_input_duration_sec += file_duration_sec
            self.processed_output_duration_sec += output_duration_sec
            self.end_time = datetime.now()

            # feldolgozott fájl nevét eltesszük a Log2-hez
            self.processed_files_list.append(input_file_basename)

            # mentjük az állapotot "completed"-re
            self.save_processing_state(i, "completed", file_path)

            # stat frissítés GUI-ra
            self.update_stats()

            # --- részösszeg popup logika (új) ---
            try:
                popup_interval_raw = self.popup_interval_var.get() if hasattr(self, "popup_interval_var") else "0"
                popup_interval = int(popup_interval_raw) if popup_interval_raw.strip() != "" else 0
            except Exception:
                popup_interval = 0

            if popup_interval > 0 and (self.processed_files_count % popup_interval == 0):
                total_so_far = (datetime.now() - self.start_time).total_seconds()
                try:
                    self.show_completion_popup(
                        processed_count=self.processed_files_count,
                        total_duration=total_so_far,
                        errors=self.errors
                    )
                except Exception as e:
                    self.add_log_entry("ERROR", f"Részösszeg popup hiba: {e}")

            # kis csipogás siker esetén (nálad van play_success_sound() is)
            self.play_success_sound()

        # --- 4) ciklus vége / normál lezárás ---
        self.end_time = datetime.now()
        total_run_time_seconds = (self.end_time - self.start_time).total_seconds()
        total_run_time_formatted = self.format_time(total_run_time_seconds)

        # GUI végső állapot
        self.overall_progress_label_text.set("Összes fájl feldolgozása: 100%")
        self.overall_progress['value'] = 100
        self.file_progress_label_text.set("Aktuális fájl feldolgozása: Kész")
        self.status.set(f"✅ Kész. Összesen {total_files} fájl feldolgozva.")
        self.update_stats()

        # gombok visszaállítása
        self.start_button.config(text="Feldolgozás indítása")
        self.stop_button.config(state="disabled")
        self.resume_button.config(state="normal" if os.path.exists(self.processing_state_file) else "disabled")

        # --- 5) Logok generálása (Log1, Log2, Log3 TXT / JSON / XLSX / PDF) ---
        try:
            self.generate_log_files()
            self.add_log_entry("INFO", "Log fájlok elkészültek: TXT, JSON, XLSX, PDF formátumban (Log1, Log2, Log3)")
        except Exception as e:
            self.add_log_entry("ERROR", f"Hiba a logfájlok generálása közben: {e}")

        # --- 6) Végső popup (összefoglaló ablak) ---
        try:
            try:
                self.show_summary_window()
            except AttributeError:
                self.show_completion_popup(
                    processed_count=self.processed_files_count,
                    total_duration=total_run_time_seconds,
                    errors=self.errors
                )
        except Exception as e:
            self.add_log_entry("ERROR", f"Végső popup megjelenítésekor hiba: {e}")

        # --- 6/b) Szöveges összefoglaló logikus százalékkal ---
        try:
            total_in = self.processed_input_size_mb
            total_out = self.processed_output_size_mb

            if total_in > 0:
                diff_mb = total_out - total_in
                change_percent = abs((diff_mb / total_in) * 100)

                if total_out < total_in:
                    change_type = "Megtakarítás"
                    sign = "-"
                    emoji = "💾"
                elif total_out > total_in:
                    change_type = "Növekedés"
                    sign = "+"
                    emoji = "📈"
                else:
                    change_type = "Nincs változás"
                    sign = "±"
                    emoji = "⚪"

                summary_text = (
                    f"{emoji} Feldolgozás befejezve!\n\n"
                    f"Feldolgozott fájlok száma: {total_files}\n"
                    f"Teljes bemeneti méret: {total_in:.2f} MB\n"
                    f"Teljes kimeneti méret: {total_out:.2f} MB\n"
                    f"{change_type}: {sign}{abs(diff_mb):.2f} MB ({sign}{change_percent:.2f} %)\n"
                    f"Kisebb fájlok: {self.smaller_files_count} db (-{self.smaller_total_diff_mb:.2f} MB)\n"
                    f"Nagyobb fájlok: {self.larger_files_count} db (+{self.larger_total_diff_mb:.2f} MB)\n"
                    f"Teljes futásidő: {total_run_time_formatted}\n"
                    f"Kezdés: {self.start_time.strftime('%Y-%m-%d %H:%M:%S')}\n"
                    f"Befejezés: {self.end_time.strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                    f"📂 Log mappa: {self.log_dir}"
                )
            else:
                summary_text = (
                    "⚠️ Nincs elérhető statisztika (nincs bemeneti adat vagy feldolgozás sikertelen)."
                )

            self.show_summary_popup(summary_text, self.log_dir)
        except Exception as e:
            self.add_log_entry("ERROR", f"Összegző popup / stat frissítés hiba: {e}")

        # --- 7) Végi stat mezők GUI frissítése ---
        try:
            if self.processed_input_size_mb > 0:
                total_in = self.processed_input_size_mb
                total_out = self.processed_output_size_mb
                diff_mb = total_out - total_in
                saving_percent = abs((diff_mb / total_in) * 100)

                if total_out < total_in:
                    label_text = f"Megtakarítás: {saving_percent:.2f} %"
                    color = "blue"
                elif total_out > total_in:
                    label_text = f"Növekedés: {saving_percent:.2f} %"
                    color = "red"
                else:
                    label_text = "Nincs változás"
                    color = "black"

                self.total_output_mb_var.set(f"Összes kimenő MB: {total_out:.2f} MB")
                self.saving_percent_var.set(label_text)
                # --- Színezés a megtakarítás / növekedés alapján ---
                try:
                    if hasattr(self, "summary_saving_label"):
                        self.summary_saving_label.config(foreground=color)
                    else:
                        self.add_log_entry("DEBUG", f"Megtakarítás szín: {color}")
                except Exception as e:
                    self.add_log_entry("ERROR", f"Nem sikerült a szín beállítása: {e}")

                self.summary_text_var.set(
                    f"Összesített: Be: {total_in:.2f} MB → Ki: {total_out:.2f} MB ({label_text})\n"
                    f"Kisebb fájlok: {self.smaller_files_count} db (-{self.smaller_total_diff_mb:.2f} MB) | "
                    f"Nagyobb fájlok: {self.larger_files_count} db (+{self.larger_total_diff_mb:.2f} MB)"
                )
            else:
                self.total_output_mb_var.set("Összes kimenő MB: N/A")
                self.saving_percent_var.set("Megtakarítás: N/A %")
                self.summary_text_var.set("Összesített: N/A")
        except Exception as e:
            self.add_log_entry("ERROR", f"Hiba a végi statisztika frissítésekor: {e}")


    def reset_for_new_run(self):
        if os.path.exists(self.processing_state_file):
            os.remove(self.processing_state_file)
        self.input_files = []
        self.file_tree.delete(*self.file_tree.get_children())
        self.tree_items = {}
        self.processed_files_count = 0
        self.processed_input_size_mb = 0
        self.processed_output_size_mb = 0
        self.processed_input_duration_sec = 0
        self.processed_output_duration_sec = 0
        self.current_file_index = -1
        self.update_button_states()
        self.update_stats()
        self.add_log_entry("INFO", "Állapot visszaállítva. Készen áll egy új feldolgozásra.")
        self.status.set("Készen áll egy új feldolgozásra. Válassz mappát, ha szükséges.")
        self.root.update_idletasks()

    def process_video(self, file_path, file_index, file_duration_sec, file_size_mb, debug_dir, process_log_path):
        processed_count = 0
        total_duration = 0
        self.errors = []  # Tiszta hibalista minden futás elején
        log_line_number = 1
        # Létrehozzuk a videó-specifikus logfájl elérési útvonalát
        # process_log_path = os.path.join(self.output_folder, f"LOG_{os.path.basename(file_path).split('.')[0]}.txt")
        process_log_path = os.path.join(self.log_dir, f"LOG_{os.path.basename(file_path).split('.')[0]}.txt")
        # A feldolgozás kezdetének naplózása
        self.add_log_entry("INFO", f"[{log_line_number}] Feldolgozás megkezdve: {os.path.basename(file_path)}",
                           process_log_path)
        log_line_number += 1
        start_time = datetime.now()
        # Naplóbejegyzés a szkript nevével és a kezdési időponttal
        script_name = os.path.basename(__file__)
        self.add_log_entry("INFO", f"[{log_line_number}] Futtató szkript: {script_name}, Kezdési időpont: {start_time}")
        log_line_number += 1
        with open(process_log_path, "a", encoding="utf-8") as log_file:
            log_file.write(
                f"[{log_line_number}] {start_time} - Futtató szkript: {script_name}, Kezdési időpont: {start_time}\n")
        log_line_number += 1
        # Bemeneti fájl méretének logolása (de ne adjuk hozzá a globális összesítőhöz itt)
        self.add_log_entry("INFO", f"[{log_line_number}] Bemeneti fájl mérete: {file_size_mb:.2f} MB")
        log_line_number += 1
        with open(process_log_path, "a", encoding="utf-8") as log_file:
            log_file.write(
                f"[{log_line_number}] {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - Bemeneti fájl mérete: {file_size_mb:.2f} MB\n")
        log_line_number += 1
        self.current_file_progress_input_size = 0
        self.current_file_progress_input_duration = 0
        pixel_threshold = self.pixel_threshold_scale.get()
        min_motion_duration = self.min_motion_duration_scale.get()
        motion_end_buffer_duration = self.motion_end_buffer_scale.get()
        idle_duration = self.idle_duration_scale.get()
        pre_motion_buffer_duration = self.pre_motion_buffer_scale.get()
        crossfade_duration = self.crossfade_duration_scale.get()
        file_stats = {
            'motion_duration': 0,
            'motion_percent': 0,
            'output_file': None,
            'status': "Feldolgozva",
            'input_size_mb': file_size_mb,
            'output_size_mb': 0,
            'input_duration_sec': file_duration_sec,
            'output_duration_sec': 0,
            'start_time': start_time.strftime("%Y-%m-%d %H:%M:%S"),
            'end_time': ""
        }
        try:
            self.add_log_entry("INFO",
                               f"[{log_line_number}] Mozgásérzékelés a {os.path.basename(file_path)} fájlban...")
            log_line_number += 1
            self.status.set(f"Mozgásérzékelés a {os.path.basename(file_path)} fájlban...")
            with open(process_log_path, "a", encoding="utf-8") as log_file:
                log_file.write(
                    f"[{log_line_number}] {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - Mozgásérzékelés a {os.path.basename(file_path)} fájlban...\n")
            log_line_number += 1
            cap = cv2.VideoCapture(file_path)
            if not cap.isOpened():
                self.add_log_entry("ERROR", f"[{log_line_number}] Hiba: A videó nem olvasható.")
                log_line_number += 1
                self.status.set("Hiba: A videó nem olvasható.")
                with open(process_log_path, "a", encoding="utf-8") as log_file:
                    log_file.write(
                        f"[{log_line_number}] {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - Hiba: A videó nem olvasható.\n")
                log_line_number += 1
                file_stats['status'] = "Hiba"
                self.errors.append("A videó nem olvasható.")
                end_time = datetime.now()
                total_duration = (end_time - start_time).total_seconds()
                file_stats['end_time'] = end_time.strftime("%Y-%m-%d %H:%M:%S")
                self.show_completion_popup(processed_count, total_duration, self.errors)
                return file_stats
            fps = cap.get(cv2.CAP_PROP_FPS)
            frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            buffer_frames = int(motion_end_buffer_duration * fps)
            pre_buffer_frames = int(pre_motion_buffer_duration * fps)
            # --- Záró időpont OCR az utolsó képkockából ---
            cap.set(cv2.CAP_PROP_POS_FRAMES, frame_count - 1)
            ret, last_frame = cap.read()
            end_time_str = None
            if ret:
                crop = last_frame[0:150, 0:1000]
                gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
                _, thresh = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY)
                timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                cv2.imwrite(os.path.join(debug_dir, f"debug_ocr_crop_{timestamp}.png"), crop)
                cv2.imwrite(os.path.join(debug_dir, f"debug_ocr_gray_{timestamp}.png"), gray)
                cv2.imwrite(os.path.join(debug_dir, f"debug_ocr_thresh_{timestamp}.png"), thresh)
                ocr_text = ""
                for psm in [6, 7, 8, 10, 11, 13]:
                    text = pytesseract.image_to_string(
                        thresh,
                        config=f'--psm {psm} --oem 3 -c tessedit_char_whitelist=0123456789: '
                    )
                    text = ' '.join(text.strip().split())
                    self.add_log_entry("DEBUG", f"OCR szöveg (PSM {psm}): '{text}'")
                    with open(process_log_path, "a", encoding="utf-8") as log_file:
                        log_file.write(
                            f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - OCR szöveg (PSM {psm}): '{text}'\n")
                    if text:
                        ocr_text = text
                        break
                # Csak HH:MM:SS mintát keresünk
                match = re.search(r'(\d{2}):(\d{2}):(\d{2})', ocr_text)
                if match:
                    hours, minutes, seconds = map(int, match.groups())
                    if 0 <= hours <= 23 and 0 <= minutes <= 59 and 0 <= seconds <= 59:
                        end_time_str = f"{hours:02d}-{minutes:02d}-{seconds:02d}"
                        self.add_log_entry("INFO", f"Záró időpont OCR-rel kiolvasva: {end_time_str}")
                        with open(process_log_path, "a", encoding="utf-8") as log_file:
                            log_file.write(
                                f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - Záró időpont OCR-rel kiolvasva: {end_time_str}\n")
                    else:
                        self.add_log_entry("WARNING",
                                           "OCR hiba a záró időpont kiolvasásánál (érvénytelen érték). Fallback a duration-re.")
                        with open(process_log_path, "a", encoding="utf-8") as log_file:
                            log_file.write(
                                f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - OCR hiba a záró időpont kiolvasásánál (érvénytelen érték). Fallback a duration-re.\n")
                else:
                    self.add_log_entry("WARNING", f"OCR nem talált időbélyeget: '{ocr_text}'. Fallback a duration-re.")
                    with open(process_log_path, "a", encoding="utf-8") as log_file:
                        log_file.write(
                            f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - OCR nem talált időbélyeget: '{ocr_text}'. Fallback a duration-re.\n")
            cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
            # Fallback időpont generálása, ha az OCR nem sikerült
            if not end_time_str:
                total_seconds = int(file_duration_sec)
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                seconds = total_seconds % 60
                end_time_str = f"{hours:02d}-{minutes:02d}-{seconds:02d}"
                self.add_log_entry("INFO", f"Fallback záró időpont generálva: {end_time_str}")
                with open(process_log_path, "a", encoding="utf-8") as log_file:
                    log_file.write(
                        f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - Fallback záró időpont generálva: {end_time_str}\n")
            # --- Mozgásérzékelés ---
            motion_periods = []
            motion_start_frame = None
            no_motion_frames = 0
            frame_index = 0
            ret, prev_frame = cap.read()
            if not ret:
                cap.release()
                self.add_log_entry("ERROR", f"[{log_line_number}] Hiba: A videó első képkockája nem olvasható.")
                log_line_number += 1
                self.status.set("Hiba: A videó első képkockája nem olvasható.")
                with open(process_log_path, "a", encoding="utf-8") as log_file:
                    log_file.write(
                        f"[{log_line_number}] {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - Hiba: A videó első képkockája nem olvasható.\n")
                log_line_number += 1
                file_stats['status'] = "Hiba"
                self.errors.append("A videó első képkockája nem olvasható.")
                end_time = datetime.now()
                total_duration = (end_time - start_time).total_seconds()
                file_stats['end_time'] = end_time.strftime("%Y-%m-%d %H:%M:%S")
                self.show_completion_popup(processed_count, total_duration, self.errors)
                return file_stats
            prev_gray = cv2.cvtColor(prev_frame, cv2.COLOR_BGR2GRAY)
            while True:
                if self.stop_processing_flag or self.is_paused:
                    break
                ret, frame = cap.read()
                if not ret:
                    if motion_start_frame is not None:
                        motion_periods.append((max(0, motion_start_frame - pre_buffer_frames), frame_index - 1))
                    break
                gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                diff = cv2.absdiff(prev_gray, gray)
                _, thresh = cv2.threshold(diff, 25, 255, cv2.THRESH_BINARY)
                motion_level = cv2.countNonZero(thresh)
                is_motion = motion_level > pixel_threshold
                if is_motion:
                    no_motion_frames = 0
                    if motion_start_frame is None:
                        motion_start_frame = frame_index
                else:
                    if motion_start_frame is not None:
                        no_motion_frames += 1
                        if no_motion_frames >= buffer_frames:
                            motion_end_frame = frame_index - no_motion_frames
                            if (motion_end_frame - motion_start_frame) / fps >= min_motion_duration:
                                motion_periods.append(
                                    (max(0, motion_start_frame - pre_buffer_frames), motion_end_frame + buffer_frames))
                            motion_start_frame = None
                            no_motion_frames = 0
                progress_percent = (frame_index / frame_count) * 100
                self.file_progress['value'] = progress_percent
                self.file_progress_label_text.set(f"Aktuális fájl feldolgozása: {progress_percent:.2f}%")
                self.current_file_progress_input_duration = (frame_index / fps) if fps > 0 else 0
                self.current_file_progress_input_size = (frame_index / frame_count) * file_size_mb
                self.update_stats()
                prev_gray = gray
                frame_index += 1
            cap.release()
            if self.stop_processing_flag or self.is_paused:
                self.add_log_entry("WARNING", f"[{log_line_number}] A feldolgozás leállítva/felfüggesztve.")
                log_line_number += 1
                self.status.set("A feldolgozás leállítva/felfüggesztve.")
                with open(process_log_path, "a", encoding="utf-8") as log_file:
                    log_file.write(
                        f"[{log_line_number}] {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - A feldolgozás leállítva/felfüggesztve.\n")
                log_line_number += 1
                file_stats['status'] = "Megszakítva"
                self.errors.append("A feldolgozás leállítva a felhasználó által.")
                end_time = datetime.now()
                total_duration = (end_time - start_time).total_seconds()
                file_stats['end_time'] = end_time.strftime("%Y-%m-%d %H:%M:%S")
                self.show_completion_popup(processed_count, total_duration, self.errors)
                return file_stats
            if motion_periods:
                motion_periods.sort()
                merged_periods = [motion_periods[0]]
                for start, end in motion_periods[1:]:
                    last_start, last_end = merged_periods[-1]
                    if start <= last_end + (idle_duration * fps):
                        merged_periods[-1] = (last_start, max(last_end, end))
                    else:
                        merged_periods.append((start, end))
                motion_periods = merged_periods
                total_motion_duration = sum((end - start) / fps for start, end in motion_periods)
                file_stats['motion_duration'] = total_motion_duration
                file_stats['motion_percent'] = (
                    (total_motion_duration / file_duration_sec) * 100 if file_duration_sec > 0 else 0)
                self.add_log_entry("INFO",
                                   f"[{log_line_number}] Összes mozgásos idő: {total_motion_duration:.2f} másodperc")
                log_line_number += 1
                self.status.set(f"Összes mozgásos idő: {total_motion_duration:.2f} másodperc")
                with open(process_log_path, "a", encoding="utf-8") as log_file:
                    log_file.write(
                        f"[{log_line_number}] {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - Összes mozgásos idő: {total_motion_duration:.2f} másodperc\n")
                log_line_number += 1
            else:
                self.add_log_entry("INFO", f"[{log_line_number}] Nem észleltünk mozgást.")
                log_line_number += 1
                self.status.set("Nem észleltünk mozgást.")
                with open(process_log_path, "a", encoding="utf-8") as log_file:
                    log_file.write(
                        f"[{log_line_number}] {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - Nem észleltünk mozgást.\n")
                log_line_number += 1
                file_stats['status'] = "Nincs mozgás"
                end_time = datetime.now()
                total_duration = (end_time - start_time).total_seconds()
                file_stats['end_time'] = end_time.strftime("%Y-%m-%d %H:%M:%S")
                self.show_completion_popup(processed_count, total_duration, self.errors)
                return file_stats
            # --- FFmpeg kivágás + összefűzés ---
            self.add_log_entry("INFO", f"[{log_line_number}] Mozgásos részek kivágása (gyors vágás)...")
            log_line_number += 1
            self.status.set("Mozgásos részek kivágása (gyors vágás)...")
            with open(process_log_path, "a", encoding="utf-8") as log_file:
                log_file.write(
                    f"[{log_line_number}] {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - Mozgásos részek kivágása (gyors vágás)...\n")
            log_line_number += 1
            temp_dir = os.path.join(self.output_folder, "_temp_motion")
            os.makedirs(temp_dir, exist_ok=True)
            clips_to_concat = []
            durations = []
            for idx, (start_frame, end_frame) in enumerate(motion_periods):
                start_sec = max(0, start_frame / fps)
                end_sec = min(file_duration_sec, end_frame / fps)
                duration = end_sec - start_sec
                durations.append(duration)
                out_clip = os.path.join(temp_dir, f"clip_{idx:03}.mp4")
                cmd = [
                    "ffmpeg",
                    "-y",
                    "-i", file_path,
                    "-ss", str(start_sec),
                    "-to", str(end_sec),
                    "-c", "copy",
                    out_clip
                ]
                subprocess.run(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                               creationflags=subprocess.CREATE_NO_WINDOW)
                clips_to_concat.append(out_clip)
                self.add_log_entry("INFO",
                                   f"[{log_line_number}] Kivágott klip (gyors vágás): Index: {idx}, Kezdés: {start_sec:.2f} mp, "
                                   f"Vég: {end_sec:.2f} mp, Hossz: {duration:.2f} mp")
                log_line_number += 1
            # Kimeneti fájlnév generálása
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            # Formátum: Ch2_2020-12-04-P_12-38-51__13-12-32_1920_01
            date_part = base_name[4:12]  # Pl. "20201204" -> "2020-12-04"
            time_part = base_name[12:18]  # Pl. "123851" -> "12-38-51"
            formatted_date = f"{date_part[:4]}-{date_part[4:6]}-{date_part[6:8]}"
            formatted_time = f"{time_part[:2]}-{time_part[2:4]}-{time_part[4:6]}"
            out_name_base = f"{base_name[:3]}_{formatted_date}-P_{formatted_time}__{end_time_str}_1920"
            out_path = self.get_unique_filename(self.output_folder, out_name_base, ".mp4")
            file_stats['output_file'] = out_path
            self.add_log_entry("INFO", f"[{log_line_number}] Generált kimeneti fájlnév: {os.path.basename(out_path)}")
            log_line_number += 1
            # --- Összefűzés + egyszeri kódolás ---
            concat_list = os.path.join(temp_dir, "concat_list.txt")
            with open(concat_list, "w", encoding="utf-8") as f:
                for clip in clips_to_concat:
                    f.write(f"file '{clip}'\n")
            cmd_concat = [
                "ffmpeg",
                "-y",
                "-f", "concat",
                "-safe", "0",
                "-i", concat_list,
                "-c:v", "libx264",
                "-preset", self.preset_var.get(),
                "-crf", self.crf_var.get(),
                "-c:a", "aac",
                "-b:a", "192k",
                out_path
            ]
            try:
                process = subprocess.Popen(cmd_concat, stderr=subprocess.PIPE, stdout=subprocess.DEVNULL,
                                           universal_newlines=True, creationflags=subprocess.CREATE_NO_WINDOW)
                stderr_output = []
                for line in process.stderr:
                    stderr_output.append(line)
                    if "time=" in line:
                        match = re.search(r"time=(\d+:\d+:\d+\.\d+)", line)
                        if match:
                            elapsed_str = match.group(1)
                            h, m, s = elapsed_str.split(":")
                            elapsed_seconds = int(h) * 3600 + int(m) * 60 + float(s)
                            self.current_file_progress_input_duration = elapsed_seconds
                            self.update_stats()
                process.wait()
                return_code = process.returncode
                stderr_output = "".join(stderr_output)
                if not hasattr(self, "processed_files_list"):
                    self.processed_files_list = []
                self.processed_files_list.append(os.path.basename(out_path))
                if return_code != 0:
                    self.add_log_entry("ERROR", f"[{log_line_number}] FFmpeg összefűzés hiba: {stderr_output}")
                    log_line_number += 1
                    with open(process_log_path, "a", encoding="utf-8") as log_file:
                        log_file.write(
                            f"[{log_line_number}] {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - FFmpeg összefűzés hiba: {stderr_output}\n")
                    log_line_number += 1
                    file_stats['status'] = "Hiba"
                    self.errors.append(f"FFmpeg összefűzés hiba: {stderr_output}")
                    end_time = datetime.now()
                    total_duration = (end_time - start_time).total_seconds()
                    file_stats['end_time'] = end_time.strftime("%Y-%m-%d %H:%M:%S")
                   # self.show_completion_popup(1, total_duration, self.errors)
                    return file_stats
                # Kimeneti fájl méretének és időtartamának lekérdezése
                if os.path.exists(out_path) and os.path.getsize(out_path) > 0:
                    output_size_mb = os.path.getsize(out_path) / (1024 * 1024)
                    file_stats['output_size_mb'] = output_size_mb
                    # Ne adjuk hozzá a globális összesítőhöz itt - ezt a process_all_files kezeli
                    file_stats['output_duration_sec'] = self.get_video_duration(out_path)
                    # Ne adjuk hozzá a globális összesítőhöz itt - ezt a process_all_files kezeli
                    self.add_log_entry("INFO", f"[{log_line_number}] Kimeneti fájl mérete: {output_size_mb:.2f} MB")
                    log_line_number += 1
                    with open(process_log_path, "a", encoding="utf-8") as log_file:
                        log_file.write(
                            f"[{log_line_number}] {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - Kimeneti fájl mérete: {output_size_mb:.2f} MB\n")
                    log_line_number += 1
                    self.add_log_entry("INFO",
                                       f"[{log_line_number}] Kimeneti videó hossza a klipek összefűzése után: {file_stats['output_duration_sec']:.2f} mp")
                    log_line_number += 1
                    self.add_log_entry("INFO",
                                       f"[{log_line_number}] ✅ Kész. Kimeneti fájl: {os.path.basename(out_path)}")
                    log_line_number += 1
                    self.status.set(f"✅ Kész. Kimeneti fájl: {os.path.basename(out_path)}")
                    with open(process_log_path, "a", encoding="utf-8") as log_file:
                        log_file.write(
                            f"[{log_line_number}] {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - ✅ Kész. Kimeneti fájl: {os.path.basename(out_path)}\n")
                    log_line_number += 1
                    file_stats['status'] = "Sikeres"
                    processed_count = 1
                else:
                    self.add_log_entry("ERROR",
                                       f"[{log_line_number}] Hiba: A kimeneti fájl üres vagy nem jött létre. {out_path}")
                    log_line_number += 1
                    self.status.set(f"Hiba: A kimeneti fájl üres vagy nem jött létre. {out_path}")
                    with open(process_log_path, "a", encoding="utf-8") as log_file:
                        log_file.write(
                            f"[{log_line_number}] {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - Hiba: A kimeneti fájl üres vagy nem jött létre. {out_path}\n")
                    log_line_number += 1
                    file_stats['status'] = "Hiba"
                    self.errors.append("A kimeneti fájl üres vagy nem jött létre.")
                    end_time = datetime.now()
                    total_duration = (end_time - start_time).total_seconds()
                    file_stats['end_time'] = end_time.strftime("%Y-%m-%d %H:%M:%S")
                    self.show_completion_popup(processed_count, total_duration, self.errors)
                    return file_stats
            except Exception as e:
                self.add_log_entry("ERROR", f"[{log_line_number}] FFmpeg futtatási hiba: {str(e)}")
                log_line_number += 1
                with open(process_log_path, "a", encoding="utf-8") as log_file:
                    log_file.write(
                        f"[{log_line_number}] {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - FFmpeg futtatási hiba: {str(e)}\n")
                log_line_number += 1
                file_stats['status'] = "Hiba"
                self.errors.append(f"FFmpeg futtatási hiba: {str(e)}")
                end_time = datetime.now()
                total_duration = (end_time - start_time).total_seconds()
                file_stats['end_time'] = end_time.strftime("%Y-%m-%d %H:%M:%S")
                self.show_completion_popup(processed_count, total_duration, self.errors)
                return file_stats
            # Ideiglenes fájlok törlése
            if self.delete_temp_var.get() == "1":
                self.add_log_entry("INFO", f"[{log_line_number}] Ideiglenes fájlok törlése...")
                log_line_number += 1
                self.status.set("Ideiglenes fájlok törlése...")
                with open(process_log_path, "a", encoding="utf-8") as log_file:
                    log_file.write(
                        f"[{log_line_number}] {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - Ideiglenes fájlok törlése...\n")
                log_line_number += 1
                shutil.rmtree(temp_dir, ignore_errors=True)
        except Exception as e:
            self.add_log_entry("ERROR",
                               f"[{log_line_number}] Hiba a(z) {os.path.basename(file_path)} feldolgozásánál: {str(e)}")
            log_line_number += 1
            self.status.set(f"Hiba a(z) {os.path.basename(file_path)} feldolgozásánál: {str(e)}")
            with open(process_log_path, "a", encoding="utf-8") as log_file:
                log_file.write(
                    f"[{log_line_number}] {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - Hiba a(z) {os.path.basename(file_path)} feldolgozásánál: {str(e)}\n")
            log_line_number += 1
            file_stats['status'] = "Hiba"
            self.errors.append(f"Hiba a fájl feldolgozásánál: {str(e)}")
            end_time = datetime.now()
            total_duration = (end_time - start_time).total_seconds()
            file_stats['end_time'] = end_time.strftime("%Y-%m-%d %H:%M:%S")
        file_stats['end_time'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return file_stats



    # ----------------------------------------------------------------
    def save_processing_state(self, index, status, file_path):
        state = {
            'current_index': index,
            'status': status,
            'interrupted_file': os.path.basename(file_path),
            'interruption_time': datetime.now().isoformat(),
            'input_directory': self.input_directory,
            'output_directory': self.output_folder,
            'processed_files_count': self.processed_files_count,
            'processed_input_size_mb': self.processed_input_size_mb,
            'processed_output_size_mb': self.processed_output_size_mb,
            'processed_input_duration_sec': self.processed_input_duration_sec,
            'processed_output_duration_sec': self.processed_output_duration_sec
        }
        with open(self.processing_state_file, 'w') as f:
            json.dump(state, f, indent=4)


    def check_for_interrupted_processing(self):
        if os.path.exists(self.processing_state_file):
            with open(self.processing_state_file, 'r') as f:
                state = json.load(f)
            if state.get('status') == 'processing':
                self.add_log_entry("INFO",
                                   f"Félbeszakadt feldolgozás észlelve: {state.get('interrupted_file', 'N/A')}")
                self.resume_button.config(state="normal")

    def save_settings(self):
        """Beállítások mentése a settings.ini fájlba (mindig biztonságosan létrehozza a [SETTINGS] szekciót)."""
        import configparser
        import os

        config = configparser.ConfigParser()

        # Ha létezik régi ini, betöltjük, hogy ne vesszen el semmi
        if os.path.exists(self.settings_file):
            config.read(self.settings_file, encoding="utf-8")

        # Ha nincs [SETTINGS] szekció, létrehozzuk
        if "SETTINGS" not in config:
            config["SETTINGS"] = {}

        # Most biztonságosan tölthetjük fel az értékeket
        config["SETTINGS"].update({
            "input_directory": self.input_directory,
            "output_folder": self.output_folder,
            "log_folder": self.log_folder,
            "crf": str(self.crf_var.get()),
            "preset": str(self.preset_var.get()),
            "pixel_threshold": str(self.pixel_threshold_scale.get()),
            "min_motion_duration": str(self.min_motion_duration_scale.get()),
            "motion_end_buffer": str(self.motion_end_buffer_scale.get()),
            "idle_duration": str(self.idle_duration_scale.get()),
            "pre_motion_buffer": str(self.pre_motion_buffer_scale.get()),
            "crossfade_duration": str(self.crossfade_duration_scale.get()),
            "custom_resolution": str(self.custom_resolution_var.get()),
            "output_width": str(self.output_width_var.get()),
            "output_height": str(self.output_height_var.get()),
            "skip_processed": str(self.skip_processed_var.get()),
            "delete_temp": str(self.delete_temp_var.get()),
            "popup_interval": str(self.popup_interval_var.get()),  # ← itt kellett a vessző!

            # 🔹 ÚJ sorok az FFmpeg küszöbhöz
            "auto_threshold": str(self.auto_threshold_var.get()),
            "ffmpeg_threshold": str(self.ffmpeg_threshold_var.get()),
        })

        # Fájl mentése
        with open(self.settings_file, "w", encoding="utf-8") as configfile:
            config.write(configfile)

        self.add_log_entry("INFO", "Beállítások mentve a settings.ini fájlba.")

    def load_settings(self):
        """Betölti a beállításokat a settings.ini fájlból (kódolás-ellenőrzéssel és hiányzó értékek pótlásával)."""
        import configparser
        import os

        config = configparser.ConfigParser()

        # --- 1) settings.ini ellenőrzés ---
        if not os.path.exists(self.settings_file):
            self.add_log_entry("WARNING", "A settings.ini fájl nem létezik. Alapértelmezett értékek használata.")
            return

        # --- 2) Fájl beolvasása, kódolás-ellenőrzéssel ---
        try:
            config.read(self.settings_file, encoding="utf-8")
        except UnicodeDecodeError:
            try:
                config.read(self.settings_file, encoding="cp1250")
                self.add_log_entry("WARNING", "A settings.ini nem UTF-8, CP1250 kódolással olvasva.")
            except Exception as e:
                self.add_log_entry("ERROR", f"Nem sikerült beolvasni a settings.ini fájlt: {e}")
                return

        # --- 3) Ha hiányzik a [SETTINGS] szekció, létrehozzuk üresen ---
        if "SETTINGS" not in config:
            config["SETTINGS"] = {}
            self.add_log_entry("WARNING", "A settings.ini nem tartalmaz [SETTINGS] szekciót. Üresen létrehozva.")

        s = config["SETTINGS"]

        # --- 4) Értékek betöltése GUI változókba ---
        self.input_directory = s.get("input_directory", "")
        self.output_folder = s.get("output_folder", "")
        self.log_folder = s.get("log_folder", "")

        # CRF, Preset stb.
        self.crf_scale.set(int(float(s.get("crf", 23))))
        self.preset_var.set(s.get("preset", "medium"))

        # Mozgásérzékelés / időzítések
        self.pixel_threshold_scale.set(float(s.get("pixel_threshold", 2400)))
        self.min_motion_duration_scale.set(float(s.get("min_motion_duration", 2.0)))
        self.motion_end_buffer_scale.set(float(s.get("motion_end_buffer", 1.0)))
        self.idle_duration_scale.set(float(s.get("idle_duration", 5.0)))
        self.pre_motion_buffer_scale.set(float(s.get("pre_motion_buffer", 1.0)))
        self.crossfade_duration_scale.set(float(s.get("crossfade_duration", 0.5)))

        # Egyedi kimeneti felbontás
        self.custom_resolution_var.set(s.get("custom_resolution", "0"))
        self.output_width_var.set(s.get("output_width", "1920"))
        self.output_height_var.set(s.get("output_height", "1080"))

        # Kapcsolók
        self.skip_processed_var.set(s.get("skip_processed", "0"))
        self.delete_temp_var.set(s.get("delete_temp", "1"))

        # Popup gyakoriság
        popup_raw = s.get("popup_interval", "10")
        try:
            self.popup_interval_var.set(str(int(popup_raw)))
        except ValueError:
            self.popup_interval_var.set("10")

        # --- 5) FFmpeg küszöbérték és automatikus mód betöltése ---
        try:
            auto_thr_val = s.get("auto_threshold", "False").lower() in ("true", "1", "yes")
            self.auto_threshold_var.set(auto_thr_val)

            thr_val = float(s.get("ffmpeg_threshold", "0.30"))
            self.ffmpeg_threshold_var.set(thr_val)

            # Spinbox engedélyezés / tiltás az automata mód szerint
            self.ffmpeg_threshold_spin.config(state=("disabled" if auto_thr_val else "normal"))

            # Felirat frissítése a GUI-n
            mode_txt = "auto" if auto_thr_val else "fix"
            self.auto_threshold_value_var.set(f"Küszöb: {thr_val:.2f} ({mode_txt})")

            self.add_log_entry("INFO", f"FFmpeg küszöb betöltve: {thr_val:.2f}, automata={auto_thr_val}")
        except Exception as e:
            self.add_log_entry("WARNING", f"FFmpeg küszöb betöltése sikertelen: {e}")
            self.ffmpeg_threshold_var.set(0.30)
            self.auto_threshold_var.set(False)

        # --- 6) Naplózás és státusz ---
        self.add_log_entry("INFO", "Beállítások sikeresen betöltve a settings.ini fájlból.")

    def get_unique_filename(self, directory, base_name, extension):
        counter = 1
        output_file = os.path.join(directory, f"{base_name}{extension}")
        while os.path.exists(output_file):
            output_file = os.path.join(directory, f"{base_name}_{counter}{extension}")
            counter += 1
        return output_file


    def adjust_column_widths(self):
        for col in self.file_tree["columns"]:
            max_width = 0
            for item in self.file_tree.get_children():
                value = self.file_tree.item(item, 'values')[self.file_tree["columns"].index(col)]
                width = len(str(value)) * 10
                if width > max_width:
                    max_width = width
            self.file_tree.column(col, width=max_width + 20)


    def generate_report(self):
        # Ide jöhet a jelentéskészítő logika
        # Például, a napló adatok alapján PDF vagy más formátumú jelentés generálása
        self.add_log_entry("INFO", "Jelentés generálása...")
        try:
            pdf = FPDF()
            pdf.add_page()

            # Betűkészlet beállítása
            pdf.add_font("DejaVuSans", "", font_path_regular)
            pdf.add_font("DejaVuSans", "B", font_path_bold)
            pdf.set_font("DejaVuSans", "B", 16)

            pdf.cell(w=0, h=10, txt="Video Feldolgozási Jelentés", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.set_font("DejaVuSans", "", 12)

            pdf.cell(w=0, h=10, txt=f"Jelentés dátuma: {datetime.now().strftime('%Y-%m-%d %H:%M')}", new_x=XPos.LMARGIN,
                     new_y=YPos.NEXT)
            pdf.cell(w=0, h=10, txt=f"Bemeneti mappa: {self.input_directory}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.cell(w=0, h=10, txt=f"Kimeneti mappa: {self.output_folder}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.cell(w=0, h=10, txt=f"Feldolgozott fájlok száma: {self.processed_files_count}", new_x=XPos.LMARGIN,
                     new_y=YPos.NEXT)
            pdf.cell(w=0, h=10, txt=f"Összes feldolgozott méret: {self.processed_size_mb:.2f} MB", new_x=XPos.LMARGIN,
                     new_y=YPos.NEXT)
            pdf.cell(w=0, h=10,
                     txt=f"Összes feldolgozott időtartam: {str(timedelta(seconds=self.processed_duration_sec)).split('.')[0]}",
                     new_x=XPos.LMARGIN, new_y=YPos.NEXT)

            pdf.ln(10)  # Sor kihagyása

            pdf.set_font("DejaVuSans", "B", 14)
            pdf.cell(w=0, h=10, txt="Feldolgozási napló", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.set_font("DejaVuSans", "", 10)

            # Napló adatok hozzáadása
            for item in self.log_text.get_children():
                time, level, message = self.log_text.item(item)['values']
                pdf.multi_cell(w=0, h=5, txt=f"[{time}] [{level}] {message}")

            report_filename = os.path.join(self.output_folder,
                                           f"Jelentés_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.pdf")
            pdf.output(report_filename)
            self.add_log_entry("INFO", f"Jelentés sikeresen mentve: {report_filename}")

        except Exception as e:
            self.add_log_entry("ERROR", f"Hiba a jelentés generálása közben: {e}")

    def show_summary_popup(self, summary_text=None, log_dir=None):
        """
        Feldolgozás összesített adatai – automatikus ikon, színezés és fájlstatisztika.
        summary_text: opcionális előre generált szöveg
        log_dir: log mappa elérési út (opcionális)
        """
        import os, subprocess
        from datetime import timedelta
        from tkinter import Toplevel, ttk

        popup = Toplevel(self.root)
        popup.title("Összesített feldolgozás")
        popup.geometry("560x480")
        popup.configure(padx=20, pady=20)
        popup.resizable(False, False)

        # -------------------------------------------------
        # 1️⃣ Adatok a self.* változókból
        # -------------------------------------------------
        input_mb = getattr(self, "processed_input_size_mb", 0.0) or 0.0
        output_mb = getattr(self, "processed_output_size_mb", 0.0) or 0.0
        total_files = getattr(self, "processed_files_count", 0)
        smaller_files = getattr(self, "smaller_files_count", 0)
        larger_files = getattr(self, "larger_files_count", 0)
        smaller_diff_mb = getattr(self, "smaller_total_diff_mb", 0.0)
        larger_diff_mb = getattr(self, "larger_total_diff_mb", 0.0)
        start_time = getattr(self, "start_time", None)
        end_time = getattr(self, "end_time", None)

        total_time = (end_time - start_time).total_seconds() if start_time and end_time else 0

        # -------------------------------------------------
        # 2️⃣ Megtakarítás / növekedés számítása
        # -------------------------------------------------
        diff_mb = output_mb - input_mb
        change_percent = 0.0
        if input_mb > 0:
            change_percent = abs((diff_mb / input_mb) * 100)

        if input_mb == 0 and output_mb == 0:
            emoji = "⚪"
            color = "black"
            change_text = "Nincs adat"
            popup.title("Összesített feldolgozás – ⚪ Nincs adat")
        elif output_mb < input_mb:
            emoji = "💾"
            color = "#0078D7"  # kék
            change_text = f"{emoji} Megtakarítás: -{abs(diff_mb):.2f} MB ({change_percent:.2f} %)"
            popup.title("Összesített feldolgozás – 💾 Megtakarítás")
        elif output_mb > input_mb:
            emoji = "📈"
            color = "#C00000"  # piros
            change_text = f"{emoji} Növekedés: +{abs(diff_mb):.2f} MB ({change_percent:.2f} %)"
            popup.title("Összesített feldolgozás – 📈 Növekedés")
        else:
            emoji = "⚪"
            color = "black"
            change_text = f"{emoji} Nincs változás (0.00 %)"
            popup.title("Összesített feldolgozás – ⚪ Nincs változás")

        # -------------------------------------------------
        # 3️⃣ Fejléc és fő adatok blokk
        # -------------------------------------------------
        ttk.Label(
            popup,
            text="📊 Feldolgozás befejezve – Összesített adatok",
            font=("Helvetica", 12, "bold")
        ).pack(pady=(5, 10))

        data_frame = ttk.Frame(popup)
        data_frame.pack(pady=5, fill="x")

        ttk.Label(data_frame, text=f"Bemeneti méret:  {input_mb:.2f} MB", font=("Consolas", 10)).pack(anchor="w")
        ttk.Label(data_frame, text=f"Kimeneti méret:  {output_mb:.2f} MB", font=("Consolas", 10)).pack(anchor="w")

        ttk.Label(
            data_frame,
            text=change_text,
            font=("Consolas", 10, "bold"),
            foreground=color
        ).pack(anchor="w", pady=(4, 4))

        # -------------------------------------------------
        # 4️⃣ Fájl-statisztikák
        # -------------------------------------------------
        file_stats_frame = ttk.LabelFrame(popup, text="📦 Fájlstatisztika", padding=(10, 6))
        file_stats_frame.pack(pady=(8, 5), fill="x")

        ttk.Label(file_stats_frame, text=f"Kisebb fájlok: {smaller_files} db  (−{smaller_diff_mb:.2f} MB)",
                  font=("Consolas", 9), foreground="#0078D7").pack(anchor="w")
        ttk.Label(file_stats_frame, text=f"Nagyobb fájlok: {larger_files} db  (+{larger_diff_mb:.2f} MB)",
                  font=("Consolas", 9), foreground="#C00000").pack(anchor="w")

        ttk.Label(file_stats_frame,
                  text=f"Összes fájl: {total_files} db",
                  font=("Consolas", 9, "bold")).pack(anchor="w", pady=(4, 0))

        # -------------------------------------------------
        # 5️⃣ Idők és futásadatok
        # -------------------------------------------------
        info_frame = ttk.LabelFrame(popup, text="🕒 Időadatok", padding=(10, 6))
        info_frame.pack(pady=(8, 5), fill="x")

        runtime = str(timedelta(seconds=int(total_time)))
        start_str = start_time.strftime("%Y-%m-%d %H:%M:%S") if start_time else "N/A"
        end_str = end_time.strftime("%Y-%m-%d %H:%M:%S") if end_time else "N/A"

        ttk.Label(info_frame, text=f"Teljes futásidő: {runtime}", font=("Consolas", 9)).pack(anchor="w")
        ttk.Label(info_frame, text=f"Kezdés:         {start_str}", font=("Consolas", 9)).pack(anchor="w")
        ttk.Label(info_frame, text=f"Befejezés:      {end_str}", font=("Consolas", 9)).pack(anchor="w")

        # -------------------------------------------------
        # 6️⃣ Gombok
        # -------------------------------------------------
        def open_log_dir():
            try:
                if log_dir and os.path.exists(log_dir):
                    if os.name == "nt":
                        os.startfile(log_dir)
                    elif os.name == "posix":
                        subprocess.Popen(["xdg-open", log_dir])
            except Exception as e:
                self.add_log_entry("ERROR", f"Nem sikerült megnyitni a log mappát: {e}")

        btn_frame = ttk.Frame(popup)
        btn_frame.pack(pady=15)

        if log_dir:
            ttk.Button(btn_frame, text="📂 Log mappa megnyitása", command=open_log_dir).pack(side="left", padx=10)
        ttk.Button(btn_frame, text="✅ OK (Új futásra kész)",
                   command=lambda: (popup.destroy(), self.reset_for_new_run())).pack(side="left", padx=10)
        ttk.Button(btn_frame, text="❌ Kilépés (adatok megmaradnak)", command=popup.destroy).pack(side="left", padx=10)


if __name__ == "__main__":
    locale.setlocale(locale.LC_ALL, 'hu_HU.UTF-8')
    root = Tk()
    app = MotionExtractorApp(root)
    root.mainloop()
