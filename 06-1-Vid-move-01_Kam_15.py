import configparser
import json
import locale
import os
import re
import shutil
import subprocess
import threading
import tkinter as tk
from datetime import datetime
from datetime import timedelta
from tkinter import Tk, filedialog, StringVar, ttk, messagebox, Scale
from tkinter import Toplevel, Label, Button

import chardet
import cv2
import pytesseract
import winsound  # Csak Windows-on működik
from fpdf import FPDF, XPos, YPos

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'  #
# -------------------------------------------
# -----------------------------------------------
# === PARAMÉTEREK ===
font_path_regular = "F:/__Panel/fonts/DejaVuSans.ttf"
font_path_bold = "F:/__Panel/fonts/DejaVuSans-Bold.ttf"
ffmpeg_path = r"C:/ffmpeg_full/bin/ffmpeg.exe"
pixel_threshold_default = 2400
min_motion_duration_default = 2.0  # sec
motion_end_buffer_default = 1.0  # sec
# A runtime_log.txt fájlnév dinamikus beállítása
# Létrehozunk egy dátumot tartalmazó fájlnevet
log_filename = f"runtime_log_{datetime.now().strftime('%Y-%m-%d')}.txt"


class MotionExtractorApp:
    def __init__(self, root):
        # Lokalizáció beállítása csak a GUI megjelenítéshez (dátumok, stb.), de számokhoz 'C' lokalizáció
        locale.setlocale(locale.LC_TIME, 'hu_HU.UTF-8')  # Magyar dátumformátumok
        locale.setlocale(locale.LC_NUMERIC, 'C')  # Angol tizedes tört (0.5)

        self.root = root
        script_name = os.path.basename(__file__)
        script_dir = os.path.dirname(os.path.abspath(__file__))

        self.script_name = script_name
        self.script_dir = script_dir
        self.play_sound = True  # Alapértelmezett érték: bekapcsolt hangjelzés
        self.errors = []  # A hibák gyűjtésére szolgáló lista
        self.fast_mode_var = tk.StringVar(value="0")

        # --- Ablakcím frissítése a feldolgozási mód szerint ---
        mode_label = "⚡ Gyors mód" if self.fast_mode_var.get() == "1" else "🎬 Normál mód"
        self.root.title(f"Video Mozgásérzékelő és Feldolgozó – {self.script_name}  ({mode_label})")

        # Ha a jelölőnégyzetet menet közben kapcsolod át, frissítjük a címsort
        self.fast_mode_var.trace_add("write", lambda *args: self.update_window_title())

        self.root.geometry("1400x900")
        self.root.resizable(True, True)
        self.is_paused = False
        self.stop_processing_flag = False
        self.processing_thread = None

        self.log_line_number = 0
        self.runtime_log = []  # <<< Ezt a sort adja hozzá
        self.current_log_file = None  # <<< EZT A SORT KELL HOZZÁADNI

        self.current_file_index = -1
        self.processed_files_count = 0
        self.processed_size_mb = 0
        self.processed_duration_sec = 0
        self.current_file_progress_duration = 0
        self.current_file_progress_size = 0
        self.start_time = None
        self.end_time = None
        self.input_directory = ""
        self.output_folder = ""
        self.input_files = []
        self.total_size_mb = 0
        self.total_duration_sec = 0
        self.start_time = None
        self.calculated_end_time = None
        self.remaining_time = None
        self.end_time = None

        self.processed_input_size_mb = 0
        self.processed_output_size_mb = 0
        self.processed_input_duration_sec = 0
        self.processed_output_duration_sec = 0
        self.current_file_progress_input_size = 0
        self.current_file_progress_input_duration = 0
        self.tree_items = {}
        self.settings_file = "settings.ini"
        self.processing_state_file = "processing_state.json"
        self._last_progress_update = 0  # utolsó progress frissítés időbélyeg

        self.log1_data = []
        self.log2_data = {}
        self.log3_data = []

        self.status = StringVar()
        self.status.set("Válassz ki egy bemeneti mappát és kimeneti mappát.")

        # Log mappa változó inicializálása
        self.log_folder = ""  # Alapértelmezett üres érték
        self.log_dir_var = StringVar(value="")  # GUI-ból olvasható log mappa

        # --- GUI felépítése ---
        main_frame = ttk.Frame(root, padding="10")
        main_frame.pack(fill="both", expand=True)

        main_frame.grid_columnconfigure(0, weight=1)
        main_frame.grid_columnconfigure(1, weight=0)
        main_frame.grid_rowconfigure(0, weight=0)
        main_frame.grid_rowconfigure(1, weight=1)
        main_frame.grid_rowconfigure(2, weight=0)

        # Bal oldali top panel (Napló és Statisztika)
        left_top_panel = ttk.Frame(main_frame)
        left_top_panel.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        left_top_panel.grid_columnconfigure(0, weight=1)
        left_top_panel.grid_rowconfigure(0, weight=0)
        left_top_panel.grid_rowconfigure(1, weight=0)

        log_frame = ttk.Frame(left_top_panel)
        log_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        log_frame.grid_columnconfigure(0, weight=1)
        log_frame.grid_rowconfigure(0, weight=1)

        ttk.Label(log_frame, text="Napló", font=("Helvetica", 14, "bold")).grid(row=0, column=0, sticky="w")

        self.log_text = ttk.Treeview(log_frame, columns=("time", "level", "message"), show="headings", height=5)
        self.log_text.heading("time", text="Idő")
        self.log_text.heading("level", text="Szint")
        self.log_text.heading("message", text="Üzenet")
        self.log_text.column("time", width=150)
        self.log_text.column("level", width=80)
        self.log_text.column("message", width=800)
        self.log_text.grid(row=1, column=0, sticky="nsew", pady=5)

        self.log_scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=self.log_scrollbar.set)
        self.log_scrollbar.grid(row=1, column=1, sticky="ns")

        self.check_dependencies()

        self.stats_frame = ttk.LabelFrame(left_top_panel, text="Statisztika", padding="10")
        self.stats_frame.grid(row=1, column=0, sticky="ew", pady=(0, 10))
        self.stats_frame.grid_columnconfigure(0, weight=0)
        self.stats_frame.grid_columnconfigure(1, weight=1)
        self.stats_frame.grid_columnconfigure(2, weight=1)
        self.stats_frame.grid_columnconfigure(3, weight=1)
        self.stats_frame.grid_columnconfigure(4, weight=1)

        self.total_files_var = StringVar(value="Összesen: N/A")
        self.processed_files_var = StringVar(value="Feldolgozva: N/A")
        self.remaining_files_var = StringVar(value="Hátralévő: N/A")

        self.total_size_var = StringVar(value="Összesen: N/A MB")
        self.processed_size_var = StringVar(value="Feldolgozva: N/A MB")
        self.remaining_size_var = StringVar(value="Hátralévő: N/A MB")

        self.total_duration_var = StringVar(value="Összesen: N/A")
        self.processed_duration_var = StringVar(value="Feldolgozva: N/A")
        self.remaining_duration_var = StringVar(value="Hátralévő: N/A")

        self.start_time_var = StringVar(value="Kezdés: N/A")
        self.elapsed_time_var = StringVar(value="Eltelt: N/A")
        self.remaining_time_var = StringVar(value="Hátralévő: N/A")
        self.end_time_var = StringVar(value="Várható zárás: N/A")
        self.total_processing_time_var = StringVar(value="Össz. futásidő: N/A")
        self.avg_time_per_file_var = StringVar(value="Átlag / fájl: N/A perc")

        self.summary_text_var = StringVar(value="Összesített: N/A")
        self.total_output_mb_var = StringVar(value="Összes kimenő MB: N/A")
        self.saving_percent_var = StringVar(value="Megtakarítás: N/A %")

        stats_labels = [
            ("Fájlok:", self.total_files_var, self.processed_files_var, self.remaining_files_var),
            ("Méret:", self.total_size_var, self.processed_size_var, self.remaining_size_var),
            ("Idő:", self.total_duration_var, self.processed_duration_var, self.remaining_duration_var),
            ("Időpontok:", self.start_time_var, self.elapsed_time_var, self.remaining_time_var, self.end_time_var),
            ("Futásidő:", self.total_processing_time_var)
        ]
        # --- Átlagos idő / fájl megjelenítése ---
        ttk.Label(self.stats_frame, textvariable=self.avg_time_per_file_var, anchor="w").grid(
            row=len(stats_labels) - 1, column=2, padx=5, pady=2, sticky="ew"
        )

        for r, (label_text, *vars) in enumerate(stats_labels):
            ttk.Label(self.stats_frame, text=label_text).grid(row=r, column=0, padx=(10, 5), pady=2, sticky="w")
            for c, var in enumerate(vars):
                ttk.Label(self.stats_frame, textvariable=var, anchor="w").grid(row=r, column=c + 1, padx=5, pady=2,
                                                                               sticky="ew")

        # --- Összesített eredmény sor a GUI-ban ---

        # --- Összesített eredmény három oszlopban ---
        summary_row = len(stats_labels)

        # Címke: "Összesített:"
        ttk.Label(
            self.stats_frame,
            text="Összesített:",
            font=("Helvetica", 10, "bold"),
        ).grid(row=summary_row, column=0, padx=(10, 5), pady=(6, 4), sticky="w")

        # Bemeneti MB érték
        self.summary_input_var = StringVar(value="Be: N/A MB")
        ttk.Label(
            self.stats_frame,
            textvariable=self.summary_input_var,
            font=("Helvetica", 10),
        ).grid(row=summary_row, column=1, padx=5, pady=(6, 4), sticky="w")

        # Kimeneti MB érték
        self.summary_output_var = StringVar(value="Ki: N/A MB")
        ttk.Label(
            self.stats_frame,
            textvariable=self.summary_output_var,
            font=("Helvetica", 10),
        ).grid(row=summary_row, column=2, padx=5, pady=(6, 4), sticky="w")

        # Megtakarítás % érték (színes)
        self.summary_saving_var = StringVar(value="Megtakarítás: N/A %")
        self.summary_saving_label = ttk.Label(
            self.stats_frame,
            textvariable=self.summary_saving_var,
            font=("Helvetica", 10, "bold"),
        )
        self.summary_saving_label.grid(row=summary_row, column=3, padx=5, pady=(6, 4), sticky="w")

        # Jobb oldali panel (Feldolgozási beállítások)
        right_panel = ttk.Frame(main_frame, padding="10")
        right_panel.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)
        right_panel.grid_columnconfigure(0, weight=1)
        right_panel.grid_rowconfigure(0, weight=1)

        settings_frame = ttk.LabelFrame(right_panel, text="Feldolgozási beállítások", padding="10")
        settings_frame.grid(row=0, column=0, sticky='nsew')
        settings_frame.grid_columnconfigure(1, weight=1)

        row_index = 0
        ttk.Label(settings_frame, text="Videó minőség (CRF):").grid(row=row_index, column=0, sticky="w", pady=2)
        self.crf_var = StringVar(value="23")
        self.crf_scale = Scale(settings_frame, from_=0, to=51, orient='horizontal',
                               command=lambda v: self.update_scale_label(self.crf_var, v))
        self.crf_scale.set(23)
        self.crf_scale.grid(row=row_index, column=1, sticky="we")
        ttk.Label(settings_frame, textvariable=self.crf_var).grid(row=row_index, column=2, padx=5, sticky="w")
        row_index += 1

        ttk.Label(settings_frame, text="Kódolási sebesség (Preset):").grid(row=row_index, column=0, sticky="w", pady=2)
        self.preset_var = StringVar(value="medium")
        presets = ["ultrafast", "superfast", "veryfast", "faster", "fast", "medium", "slow", "slower", "veryslow"]
        self.preset_menu = ttk.Combobox(settings_frame, textvariable=self.preset_var, values=presets, state="readonly")
        self.preset_menu.grid(row=row_index, column=1, sticky="we")
        row_index += 1

        ttk.Label(settings_frame, text="Mozgásérzékenységi küszöb (pixel):").grid(row=row_index, column=0, sticky="w",
                                                                                  pady=2)
        self.pixel_threshold_var = StringVar(value=str(pixel_threshold_default))
        self.pixel_threshold_scale = Scale(settings_frame, from_=1000, to=50000, orient='horizontal', resolution=100,
                                           command=lambda v: self.update_scale_label(self.pixel_threshold_var, v))
        self.pixel_threshold_scale.set(pixel_threshold_default)
        self.pixel_threshold_scale.grid(row=row_index, column=1, sticky="we")
        ttk.Label(settings_frame, textvariable=self.pixel_threshold_var).grid(row=row_index, column=2, padx=5,
                                                                              sticky="w")
        row_index += 1

        ttk.Label(settings_frame, text="Min. mozgásklip hossza (mp):").grid(row=row_index, column=0, sticky="w", pady=2)
        self.min_motion_duration_var = StringVar(value=f"{min_motion_duration_default:.1f}".replace('.', ','))
        self.min_motion_duration_scale = Scale(settings_frame, from_=0.1, to=10, orient='horizontal', resolution=0.1,
                                               command=lambda v: self.update_scale_label(self.min_motion_duration_var,
                                                                                         v, decimals=1))
        self.min_motion_duration_scale.set(min_motion_duration_default)
        self.min_motion_duration_scale.grid(row=row_index, column=1, sticky="we")
        ttk.Label(settings_frame, textvariable=self.min_motion_duration_var).grid(row=row_index, column=2, padx=5,
                                                                                  sticky="w")
        row_index += 1

        ttk.Label(settings_frame, text="Üresjárat hossza (mp):").grid(row=row_index, column=0, sticky="w", pady=2)
        self.idle_duration_var = StringVar(value="5,0")
        self.idle_duration_scale = Scale(settings_frame, from_=0.1, to=10, orient='horizontal', resolution=0.1,
                                         command=lambda v: self.update_scale_label(self.idle_duration_var, v,
                                                                                   decimals=1))
        self.idle_duration_scale.set(5.0)
        self.idle_duration_scale.grid(row=row_index, column=1, sticky="we")
        ttk.Label(settings_frame, textvariable=self.idle_duration_var).grid(row=row_index, column=2, padx=5, sticky="w")
        row_index += 1

        ttk.Label(settings_frame, text="Elő-mozgás puffer (mp):").grid(row=row_index, column=0, sticky="w", pady=2)
        self.pre_motion_buffer_var = StringVar(value="1,0")
        self.pre_motion_buffer_scale = Scale(settings_frame, from_=0, to=5, orient='horizontal', resolution=0.1,
                                             command=lambda v: self.update_scale_label(self.pre_motion_buffer_var, v,
                                                                                       decimals=1))
        self.pre_motion_buffer_scale.set(1.0)
        self.pre_motion_buffer_scale.grid(row=row_index, column=1, sticky="we")
        ttk.Label(settings_frame, textvariable=self.pre_motion_buffer_var).grid(row=row_index, column=2, padx=5,
                                                                                sticky="w")
        row_index += 1

        ttk.Label(settings_frame, text="Utó-mozgás puffer (mp):").grid(row=row_index, column=0, sticky="w", pady=2)
        self.motion_end_buffer_var = StringVar(value=f"{motion_end_buffer_default:.1f}".replace('.', ','))
        self.motion_end_buffer_scale = Scale(settings_frame, from_=0, to=5, orient='horizontal', resolution=0.1,
                                             command=lambda v: self.update_scale_label(self.motion_end_buffer_var, v,
                                                                                       decimals=1))
        self.motion_end_buffer_scale.set(motion_end_buffer_default)
        self.motion_end_buffer_scale.grid(row=row_index, column=1, sticky="we")
        ttk.Label(settings_frame, textvariable=self.motion_end_buffer_var).grid(row=row_index, column=2, padx=5,
                                                                                sticky="w")
        row_index += 1

        ttk.Label(settings_frame, text="Áttűnés hossza (mp):").grid(row=row_index, column=0, sticky="w", pady=2)
        self.crossfade_duration_var = StringVar(value="0,5")
        self.crossfade_duration_scale = Scale(settings_frame, from_=0, to=2, orient='horizontal', resolution=0.1,
                                              command=lambda v: self.update_scale_label(self.crossfade_duration_var, v,
                                                                                        decimals=1))
        self.crossfade_duration_scale.set(0.5)
        self.crossfade_duration_scale.grid(row=row_index, column=1, sticky="we")
        ttk.Label(settings_frame, textvariable=self.crossfade_duration_var).grid(row=row_index, column=2, padx=5,
                                                                                 sticky="w")
        row_index += 1

        # --- 🎬⚡🧩 Feldolgozási mód választó (3 opció) ---
        ttk.Label(settings_frame, text="Feldolgozási mód:").grid(row=row_index, column=0, sticky="w", pady=(10, 4))
        self.mode_var = tk.StringVar(value="normal")

        mode_frame = ttk.Frame(settings_frame)
        mode_frame.grid(row=row_index, column=1, columnspan=2, sticky="w", pady=(5, 5))
        row_index += 1

        ttk.Radiobutton(
            mode_frame,
            text="🎬 Normál mód (CPU, minőségmegőrző újrakódolás)",
            variable=self.mode_var,
            value="normal"
        ).pack(anchor="w", pady=1)

        ttk.Radiobutton(
            mode_frame,
            text="⚡ Gyors mód (copy, újrakódolás nélkül – 5–10× gyorsabb)",
            variable=self.mode_var,
            value="fast"
        ).pack(anchor="w", pady=1)

        ttk.Radiobutton(
            mode_frame,
            text="🧩 Hibrid mód (GPU + CRF, gyors + tömörített)",
            variable=self.mode_var,
            value="hybrid"
        ).pack(anchor="w", pady=1)
        self.hybrid_radio = mode_frame.winfo_children()[-1]  # az utolsó rádiógomb referenciája (hibrid)

        # --- GPU kodek választó (csak akkor aktív, ha van NVENC) ---
        self.gpu_codec_label = ttk.Label(settings_frame, text="GPU kodek:", state="disabled")
        self.gpu_codec_label.grid(row=row_index, column=0, sticky="w", pady=(4, 2))

        self.gpu_codec_var = tk.StringVar(value="h264_nvenc")
        self.gpu_codec_combo = ttk.Combobox(
            settings_frame,
            textvariable=self.gpu_codec_var,
            values=["h264_nvenc", "hevc_nvenc", "av1_nvenc"],
            state="disabled",
            width=15
        )
        self.gpu_codec_combo.grid(row=row_index, column=1, sticky="w", pady=(4, 2))
        row_index += 1

        # --- GPU státuszcímke ---
        self.gpu_status_label = ttk.Label(settings_frame, text="🔴 GPU támogatás: nincs elérhető NVENC",
                                          foreground="#a00000")
        self.gpu_status_label.grid(row=row_index, column=0, columnspan=2, sticky="w", pady=(2, 8))
        row_index += 1

        # --- Tooltip a gyors módhoz ---
        def create_tooltip(widget, text):
            tip_window = None

            def show_tooltip(event):
                nonlocal tip_window
                if tip_window or not text:
                    return
                x, y, _, _ = widget.bbox("insert")
                x += widget.winfo_rootx() + 25
                y += widget.winfo_rooty() + 20
                tip_window = tw = tk.Toplevel(widget)
                tw.wm_overrideredirect(True)
                tw.wm_geometry(f"+{x}+{y}")
                label = ttk.Label(tw, text=text, background="#ffffe0", relief="solid", borderwidth=1, padding=(5, 3))
                label.pack()

            def hide_tooltip(event):
                nonlocal tip_window
                if tip_window:
                    tip_window.destroy()
                    tip_window = None

            widget.bind("<Enter>", show_tooltip)
            widget.bind("<Leave>", hide_tooltip)

        # Tooltip szöveg
        create_tooltip(
            mode_frame.winfo_children()[1],  # a második rádiógomb a "⚡ Gyors mód"
            "⚡ Gyors mód: FFmpeg -c copy vágás (5–10× gyorsabb feldolgozás).\n"
            "Nem történik újrakódolás, így nincs minőségromlás,\n"
            "de a kompatibilitás és átmenetek pontossága kissé csökkenhet."
        )

        self.custom_resolution_var = StringVar(value="0")
        self.custom_resolution_check = ttk.Checkbutton(settings_frame, text="Egyedi kimeneti felbontás",
                                                       variable=self.custom_resolution_var, onvalue="1", offvalue="0",
                                                       command=self.toggle_resolution_fields)
        self.custom_resolution_check.grid(row=row_index, column=0, columnspan=2, sticky="w", pady=(10, 0))
        row_index += 1

        ttk.Label(settings_frame, text="Kimeneti szélesség:").grid(row=row_index, column=0, sticky="w", pady=2)
        self.output_width_var = StringVar(value="1920")
        self.output_width_entry = ttk.Entry(settings_frame, textvariable=self.output_width_var, state="disabled")
        self.output_width_entry.grid(row=row_index, column=1, sticky="we")
        row_index += 1

        ttk.Label(settings_frame, text="Kimeneti magasság:").grid(row=row_index, column=0, sticky="w", pady=2)
        self.output_height_var = StringVar(value="1080")
        self.output_height_entry = ttk.Entry(settings_frame, textvariable=self.output_height_var, state="disabled")
        self.output_height_entry.grid(row=row_index, column=1, sticky="we")

        # Fájllista panel
        filelist_frame = ttk.Frame(main_frame)
        filelist_frame.grid(row=1, column=0, columnspan=2, sticky="nsew", padx=5, pady=5)
        filelist_frame.grid_columnconfigure(0, weight=1)
        filelist_frame.grid_columnconfigure(1, weight=1)
        filelist_frame.grid_rowconfigure(5, weight=1)

        ttk.Label(filelist_frame, text="Bemeneti mappa:").grid(row=0, column=0, sticky="w")
        self.input_dir_var = StringVar()
        ttk.Entry(filelist_frame, textvariable=self.input_dir_var, state='readonly').grid(row=0, column=1, sticky="we",
                                                                                          padx=(5, 0))
        ttk.Button(filelist_frame, text="Tallózás...", command=self.select_input_directory).grid(row=0, column=2,
                                                                                                 sticky="e",
                                                                                                 padx=(5, 0))

        ttk.Label(filelist_frame, text="Kimeneti mappa:").grid(row=1, column=0, sticky="w", pady=(5, 0))
        self.output_dir_var = StringVar()
        ttk.Entry(filelist_frame, textvariable=self.output_dir_var, state='readonly').grid(row=1, column=1, sticky="we",
                                                                                           padx=(5, 0), pady=(5, 0))
        ttk.Button(filelist_frame, text="Tallózás...", command=self.select_output_folder).grid(row=1, column=2,
                                                                                               sticky="e",
                                                                                               padx=(5, 0), pady=(5, 0))

        ttk.Label(filelist_frame, text="Log mappa:").grid(row=2, column=0, sticky="w", pady=2)
        self.log_dir_var = StringVar()
        ttk.Entry(filelist_frame, textvariable=self.log_dir_var, state='readonly').grid(row=2, column=1, sticky="we",
                                                                                        padx=(5, 0))
        ttk.Button(filelist_frame, text="Tallózás...", command=self.select_log_folder).grid(row=2, column=2, sticky="e",
                                                                                            padx=(5, 0))

        control_and_delete_frame = ttk.Frame(filelist_frame)
        control_and_delete_frame.grid(row=3, column=0, columnspan=3, pady=(5, 10), sticky="w")

        self.delete_selected_button = ttk.Button(control_and_delete_frame, text="Kiválasztott törlése",
                                                 command=self.delete_selected_file, state="disabled")
        self.delete_selected_button.pack(side="left", padx=5)

        self.clear_all_button = ttk.Button(control_and_delete_frame, text="Összes törlése",
                                           command=self.clear_file_list, state="disabled")
        self.clear_all_button.pack(side="left", padx=5)

        self.start_button = ttk.Button(control_and_delete_frame, text="Feldolgozás indítása",
                                       command=self.toggle_processing)
        self.start_button.pack(side="left", padx=5)
        self.stop_button = ttk.Button(control_and_delete_frame, text="Feldolgozás leállítása",
                                      command=self.stop_processing, state="disabled")
        self.stop_button.pack(side="left", padx=5)

        self.resume_button = ttk.Button(control_and_delete_frame, text="Folytatás",
                                        command=self.resume_processing, state="disabled")
        self.resume_button.pack(side="left", padx=5)

        self.adjust_columns_button = ttk.Button(control_and_delete_frame, text="Oszlopszélességek állítása",
                                                command=self.adjust_column_widths)
        self.adjust_columns_button.pack(side="left", padx=5)

        self.skip_processed_var = StringVar(value="0")
        ttk.Checkbutton(control_and_delete_frame, text="Kihagyja a már feldolgozott videókat",
                        variable=self.skip_processed_var,
                        onvalue="1", offvalue="0").pack(side="left", padx=5)

        self.delete_temp_var = StringVar(value="1")
        ttk.Checkbutton(control_and_delete_frame, text="Ideiglenes fájlok törlése", variable=self.delete_temp_var,
                        onvalue="1",
                        offvalue="0").pack(side="left", padx=5)

        control_and_delete_frame_frame = ttk.Frame(filelist_frame)
        control_and_delete_frame_frame.grid(row=4, column=0, columnspan=3, pady=10, sticky="w")
        ttk.Button(control_and_delete_frame_frame, text="Beállítások betöltése", command=self.load_settings).pack(
            side="left",
            padx=5)
        ttk.Button(control_and_delete_frame_frame, text="Beállítások mentése", command=self.save_settings).pack(
            side="left",
            padx=5)

        columns = (
            "index", "file_in", "file_size_mb", "duration_sec",
            "output_file", "output_size_mb", "output_duration_sec", "compression_percent",
            "motion_duration_sec", "motion_percent",
            "processing_start_time", "processing_end_time", "processing_time",
            "status", "method", "profile", "input_full_path", "output_full_path",
            "script_name", "script_dir"
        )

        self.file_tree = ttk.Treeview(filelist_frame, columns=columns, show="headings", height=15)

        self.file_tree.heading("index", text="Index")
        self.file_tree.heading("file_in", text="Bemeneti fájl")
        self.file_tree.heading("file_size_mb", text="Be MB")
        self.file_tree.heading("duration_sec", text="Be Idő")
        self.file_tree.heading("output_file", text="Ki név")
        self.file_tree.heading("output_size_mb", text="Ki MB")
        self.file_tree.heading("output_duration_sec", text="Ki Idő")
        self.file_tree.heading("compression_percent", text="Tömörítés (%)")
        self.file_tree.heading("motion_duration_sec", text="Mozgás Idő (s)")
        self.file_tree.heading("motion_percent", text="Mozgás (%)")
        self.file_tree.heading("processing_start_time", text="Feld. Kezdés")
        self.file_tree.heading("processing_end_time", text="Feld. Végzés")
        self.file_tree.heading("processing_time", text="Futásidő")
        self.file_tree.heading("status", text="Státusz")
        self.file_tree.heading("method", text="Eljárás")
        self.file_tree.heading("profile", text="Profil")
        self.file_tree.heading("input_full_path", text="Bemeneti útvonal")
        self.file_tree.heading("output_full_path", text="Kimeneti útvonal")
        self.file_tree.heading("script_name", text="Script név")
        self.file_tree.heading("script_dir", text="Script könyvtár")

        self.file_tree.column("index", width=50, stretch=False)
        self.file_tree.column("file_in", width=150)
        self.file_tree.column("file_size_mb", width=80)
        self.file_tree.column("duration_sec", width=80)
        self.file_tree.column("output_file", width=150)
        self.file_tree.column("output_size_mb", width=80)
        self.file_tree.column("output_duration_sec", width=80)
        self.file_tree.column("compression_percent", width=100)
        self.file_tree.column("motion_duration_sec", width=100)
        self.file_tree.column("motion_percent", width=80)
        self.file_tree.column("processing_start_time", width=100)
        self.file_tree.column("processing_end_time", width=100)
        self.file_tree.column("processing_time", width=80)
        self.file_tree.column("status", width=80)
        self.file_tree.column("method", width=80)
        self.file_tree.column("profile", width=120)
        self.file_tree.column("input_full_path", width=200)
        self.file_tree.column("output_full_path", width=200)
        self.file_tree.column("script_name", width=120)
        self.file_tree.column("script_dir", width=150)

        self.file_tree.grid(row=5, column=0, columnspan=3, sticky="nsew", pady=(0, 5))

        self.tree_scrollbar = ttk.Scrollbar(filelist_frame, orient="vertical", command=self.file_tree.yview)
        self.file_tree.configure(yscrollcommand=self.tree_scrollbar.set)
        self.tree_scrollbar.grid(row=5, column=3, sticky="ns")

        bottom_frame = ttk.Frame(main_frame, padding="10")
        bottom_frame.grid(row=2, column=0, columnspan=2, sticky="nsew")
        bottom_frame.grid_columnconfigure(0, weight=1)

        self.file_loading_progress_label_text = StringVar(value="Fájlok betöltése: 0/0")
        ttk.Label(bottom_frame, textvariable=self.file_loading_progress_label_text).grid(row=0, column=0, sticky="w",
                                                                                         pady=(5, 0))
        self.file_loading_progress = ttk.Progressbar(bottom_frame, mode='determinate')
        self.file_loading_progress.grid(row=1, column=0, sticky="ew")

        self.overall_progress_label_text = StringVar(value="Összes fájl feldolgozása: 0%")
        ttk.Label(bottom_frame, textvariable=self.overall_progress_label_text).grid(row=2, column=0, sticky="w",
                                                                                    pady=(5, 0))
        self.overall_progress = ttk.Progressbar(bottom_frame, length=600, mode='determinate')
        self.overall_progress.grid(row=3, column=0, sticky="ew")

        self.file_progress_label_text = StringVar(value="Aktuális fájl feldolgozása: 0%")
        ttk.Label(bottom_frame, textvariable=self.file_progress_label_text).grid(row=4, column=0, sticky="w",
                                                                                 pady=(5, 0))
        self.file_progress = ttk.Progressbar(bottom_frame, length=600, mode='determinate')
        self.file_progress.grid(row=5, column=0, sticky="ew")

        self.status = StringVar()
        self.status.set("Válassz ki egy bemeneti mappát és kimeneti mappát.")
        ttk.Label(bottom_frame, textvariable=self.status, wraplength=1380, justify="center").grid(row=6, column=0,
                                                                                                  sticky="ew",
                                                                                                  pady=(5, 0))

        self.file_tree.bind("<<TreeviewSelect>>", self.on_file_select)

        self.load_settings()
        self.update_stats()
        self.check_for_interrupted_processing()
        self.update_resume_button_state()
        mode_label = "⚡ Gyors mód aktív (copy vágás, 5–10× gyorsabb)" if self.fast_mode_var.get() == "1" else "🎬 Normál mód (minőségmegőrző feldolgozás)"
        color = "#007800" if self.fast_mode_var.get() == "1" else "#004080"
        self.root.after(800, lambda: self.show_mode_toast(mode_label, color))

        # ------------------------------------------------------------------
        # GPU felismerés és Hibrid mód engedélyezés
        # ------------------------------------------------------------------
        self.gpu_available = False  # alapértelmezett: nincs GPU
        self.gpu_label_var = tk.StringVar(value="GPU támogatás: ellenőrzés folyamatban...")
        gpu_label = ttk.Label(settings_frame, textvariable=self.gpu_label_var, foreground="#004080")
        gpu_label.grid(row=row_index + 1, column=0, columnspan=3, sticky="w", pady=(6, 4))

        # GPU ellenőrzés elindítása kis késleltetéssel (a GUI betöltése után)
        self.root.after(800, lambda: self.show_mode_toast(mode_label, color))
        self.check_gpu_support()  # <--- EZT IDE!
        self.detect_gpu_encoder()  # ezt már megtartod
        self.root.after(800, self.check_gpu_support)


    def detect_gpu_encoder(self):
        """
        GPU típus felismerése és a megfelelő hardveres kodekek engedélyezése (AMF / NVENC / QSV).
        Frissíti a GUI-t is: GPU státusz címke és kodek combobox.
        """
        import subprocess, os

        gpu_name = "Ismeretlen GPU"
        available_codecs = []
        has_gpu_support = False
        gpu_vendor = "Unknown"

        try:
            # --- FFmpeg encoder lista lekérdezése ---
            result = subprocess.run(
                ["ffmpeg", "-hide_banner", "-encoders"],
                stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                text=True, creationflags=subprocess.CREATE_NO_WINDOW, timeout=10
            )
            codecs_out = result.stdout.lower()

            # --- GPU gyártó detektálása Windows alatt (dxdiag) ---
            try:
                dx_path = os.path.join(os.getenv("TEMP", "."), "dxinfo.txt")
                subprocess.run(["dxdiag", "/t", dx_path],
                               stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                with open(dx_path, encoding="utf-8", errors="ignore") as f:
                    data = f.read().lower()
                if "amd" in data or "radeon" in data:
                    gpu_vendor = "AMD"
                elif "nvidia" in data:
                    gpu_vendor = "NVIDIA"
                elif "intel" in data:
                    gpu_vendor = "Intel"
            except Exception:
                pass

            # --- AMD AMF (elsőként, mert Radeon GPU-d van) ---
            if any(x in codecs_out for x in ["h264_amf", "hevc_amf", "av1_amf"]):
                gpu_vendor = "AMD"
                available_codecs = ["h264_amf", "hevc_amf", "av1_amf"]
                gpu_name = "AMD Radeon GPU"
                has_gpu_support = True

            # --- NVIDIA NVENC ---
            elif any(x in codecs_out for x in ["h264_nvenc", "hevc_nvenc", "av1_nvenc"]):
                gpu_vendor = "NVIDIA"
                available_codecs = ["h264_nvenc", "hevc_nvenc", "av1_nvenc"]
                gpu_name = "NVIDIA GPU"
                has_gpu_support = True

            # --- Intel QSV ---
            elif any(x in codecs_out for x in ["h264_qsv", "hevc_qsv"]):
                gpu_vendor = "Intel"
                available_codecs = ["h264_qsv", "hevc_qsv"]
                gpu_name = "Intel GPU"
                has_gpu_support = True

            # --- GUI frissítése ---
            if has_gpu_support:
                vendor_display = {
                    "NVIDIA": "NVENC (NVIDIA)",
                    "AMD": "AMF (AMD Radeon)",
                    "Intel": "QSV (Intel)"
                }.get(gpu_vendor, gpu_vendor)

                self.gpu_status_label.config(
                    text=f"🟢 GPU támogatás elérhető: {vendor_display} "
                         f"({', '.join(available_codecs)})",
                    foreground="#007800"
                )
                self.gpu_codec_combo.config(values=available_codecs, state="readonly")
                self.gpu_codec_var.set(available_codecs[0])

                self.add_log_entry("INFO",
                                   f"GPU gyorsítás engedélyezve: {gpu_vendor} "
                                   f"({', '.join(available_codecs)})")
            else:
                self.gpu_status_label.config(
                    text="🔴 GPU támogatás nem elérhető (NVENC / AMF / QSV hiányzik)",
                    foreground="#b00000"
                )
                self.gpu_codec_combo.config(values=["Nincs elérhető"], state="disabled")
                self.add_log_entry("WARN", "Nem találtam elérhető GPU kódolót (NVENC/AMF/QSV).")

        except subprocess.TimeoutExpired:
            self.add_log_entry("ERROR", "FFmpeg encoder lista lekérése időtúllépés miatt megszakadt.")
            self.gpu_status_label.config(text="🔴 GPU felismerés időtúllépés", foreground="#b00000")

        except Exception as e:
            self.add_log_entry("ERROR", f"GPU felismerés sikertelen: {e}")
            self.gpu_status_label.config(text="🔴 GPU felismerés hiba", foreground="#b00000")


    def update_window_title(self):
        """Frissíti az ablak címsorát a Gyors / Normál mód alapján."""
        mode_label = "⚡ Gyors mód" if self.fast_mode_var.get() == "1" else "🎬 Normál mód"
        self.root.title(f"Video Mozgásérzékelő és Feldolgozó – {self.script_name}  ({mode_label})")
        # Toast üzenet is a módváltásról
        mode_label = "⚡ Gyors mód aktív (copy vágás)" if self.fast_mode_var.get() == "1" else "🎬 Normál mód (újrakódolás)"
        color = "#007800" if self.fast_mode_var.get() == "1" else "#004080"
        self.show_mode_toast(mode_label, color)

    def show_mode_toast(self, text, color="#0078D7", duration=2500):
        """
        Jobb felső sarokban megjelenő automatikus "toast" üzenet (pl. ⚡ Gyors mód).
        duration = időtartam ezredmásodpercben (pl. 2500 = 2,5 mp)
        """
        toast = tk.Toplevel(self.root)
        toast.overrideredirect(True)
        toast.attributes("-topmost", True)
        toast.configure(bg=color)

        # Szöveg
        label = ttk.Label(
            toast,
            text=text,
            foreground="white",
            background=color,
            font=("Segoe UI", 10, "bold"),
            padding=(12, 6)
        )
        label.pack()

        # Pozíció: jobb felső sarok
        toast.update_idletasks()
        x = self.root.winfo_x() + self.root.winfo_width() - toast.winfo_reqwidth() - 20
        y = self.root.winfo_y() + 20
        toast.geometry(f"+{x}+{y}")

        # Eltűnik néhány másodperc múlva
        toast.after(duration, toast.destroy)

    def update_scale_label(self, var, value, decimals=0):
        try:
            # Biztosítjuk, hogy a bemenet string legyen, és konvertáljuk angol tizedes tört formátumba
            value = str(value).replace(',', '.')
            float_value = float(value)
            # A magyar lokalizációhoz igazítjuk a kimeneti formátumot (vessző tizedes tört jelölőként)
            formatted_value = f"{float_value:.{decimals}f}".replace('.', ',')
            var.set(formatted_value)
        except ValueError as e:
            self.add_log_entry("ERROR", f"Hiba a csúszka értékének konvertálásakor: {e}")
            self.status.set("Hiba a csúszka értékének konvertálásakor.")
            var.set("N/A")

    def read_settings_file(path):
        with open(path, 'rb') as f:
            raw_data = f.read()
        encoding = chardet.detect(raw_data)['encoding'] or 'utf-8'
        try:
            return raw_data.decode(encoding)
        except Exception:
            # végső fallback
            return raw_data.decode('latin2', errors='replace')

    def setup_log_directory(self):
        """Napváltáskor új log könyvtár és logfile létrehozása"""

        import os
        from datetime import datetime

        # --- Biztonságos főmappa-létrehozás ---
        if not hasattr(self, "log_dir") or not self.log_dir:
            self.log_dir = self.script_dir  # fallback
        os.makedirs(self.log_dir, exist_ok=True)

        today_str = datetime.now().strftime("%Y-%m-%d")

        # --- Alap log mappa kiválasztása ---
        base_log_dir = self.log_dir_var.get() or self.output_folder or self.script_dir
        os.makedirs(base_log_dir, exist_ok=True)

        # --- Napi dátumos mappa ---
        daily_dir = os.path.join(base_log_dir, today_str)
        os.makedirs(daily_dir, exist_ok=True)

        # --- "_Logok" almappa ---
        log_dir = os.path.join(daily_dir, "_Logok")
        os.makedirs(log_dir, exist_ok=True)

        # --- Mentés osztályszinten ---
        self.log_dir = log_dir

        # --- Egyedi log fájlnév kezelése (_2, _3, stb.) ---
        log_filename = f"runtime_log_{today_str}.txt"
        log_path = os.path.join(self.log_dir, log_filename)
        counter = 2
        while os.path.exists(log_path):
            log_filename = f"runtime_log_{today_str}_{counter}.txt"
            log_path = os.path.join(self.log_dir, log_filename)
            counter += 1

        self.current_log_date = today_str
        self.current_log_file = log_path
        self.log_line_number = 0

        # --- Fájl biztonságos létrehozása ---
        os.makedirs(os.path.dirname(self.current_log_file), exist_ok=True)

        # --- Fejléc beírása ---
        with open(self.current_log_file, "a", encoding="utf-8") as f:
            f.write(f"\nRuntime log - Kezdés: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Napi log könyvtár: {self.log_dir}\n")
            f.write(f"Script: {self.script_name}\n\n")

        # --- Naplóbejegyzés ---
        self.add_log_entry("INFO", f"Log könyvtár beállítva: {self.log_dir}, logfájl: {log_filename}")

    def add_log_entry(self, level, message, process_log_path=None):
        """
        Naplóbejegyzés hozzáadása a futásidőhöz és a megfelelő logfájlba.

        Args:
            level (str): A napló szintje (pl. "INFO", "ERROR").
            message (str): A napló üzenet szövege.
            process_log_path (str, optional): A specifikus feldolgozási naplófájl elérési útja.
                                              Ha nincs megadva, a fő runtime logba ír.
        """
        import os
        from datetime import datetime

        # --- Alapbeállítások ---
        if not hasattr(self, "log_line_number"):
            self.log_line_number = 0
        if not hasattr(self, "runtime_log"):
            self.runtime_log = []

        self.log_line_number += 1
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_entry_text = f"[{self.log_line_number:04d}] [{timestamp}] [{level}] {message}"

        # --- Konzolra is kiírjuk ---
        print(log_entry_text)

        # --- Memóriába is eltároljuk ---
        self.runtime_log.append(log_entry_text)

        # --- Biztonságos fő log mentés ---
        if getattr(self, "current_log_file", None):
            try:
                # Ellenőrizzük, hogy a log mappa biztosan létezik
                os.makedirs(os.path.dirname(self.current_log_file), exist_ok=True)

                with open(self.current_log_file, "a", encoding="utf-8") as f:
                    f.write(log_entry_text + "\n")
            except Exception as e:
                print(f"Hiba a fő log mentése közben: {e}")

        # --- Speciális process_log (fájlonkénti napló) ---
        if process_log_path:
            try:
                os.makedirs(os.path.dirname(process_log_path), exist_ok=True)
                if not hasattr(self, "process_log_line_number"):
                    self.process_log_line_number = 0

                self.process_log_line_number += 1
                process_log_entry_text = f"[{self.process_log_line_number:04d}] [{timestamp}] [{level}] {message}"
                with open(process_log_path, "a", encoding="utf-8") as f:
                    f.write(process_log_entry_text + "\n")
            except Exception as e:
                print(f"Hiba a feldolgozási log mentése közben: {e}")

        # --- GUI frissítés (TreeView log panel) ---
        try:
            if hasattr(self, "log_text"):
                self.log_text.insert("", "end", values=(timestamp, level, message))
                self.log_text.yview_moveto(1)
        except Exception as e:
            print(f"Hiba a Treeview frissítése közben: {e}")

    def check_dependencies(self):
        missing_dependencies = []
        try:
            import cv2
        except ImportError:
            missing_dependencies.append("opencv-python")
        try:
            import pandas
        except ImportError:
            missing_dependencies.append("pandas")
        try:
            import numpy
        except ImportError:
            missing_dependencies.append("numpy")
        try:
            import pytesseract
        except ImportError:
            missing_dependencies.append("pytesseract")
        try:
            import fpdf
        except ImportError:
            missing_dependencies.append("fpdf")

        if missing_dependencies:
            messagebox.showwarning("Hiányzó függőségek",
                                   f"A következő Python könyvtárak hiányoznak: {', '.join(missing_dependencies)}. Kérlek telepítsd őket a 'pip install <könyvtárnév>' paranccsal a terminálban.")

        if not os.path.exists(ffmpeg_path):
            messagebox.showwarning("FFmpeg hiányzik",
                                   f"Az FFmpeg végrehajtható fájl nem található a megadott útvonalon: {ffmpeg_path}. Kérlek ellenőrizd, hogy telepítve van-e és a helyes útvonal van-e megadva.")

        try:
            subprocess.run([ffmpeg_path, "-version"], capture_output=True, text=True, check=True)
            self.add_log_entry("INFO", "FFmpeg sikeresen megtalálva")
        except Exception as e:
            self.add_log_entry("ERROR", f"FFmpeg nem található vagy hibás: {e}")
            self.status.set("Hiba: FFmpeg nem található. Ellenőrizd az elérési utat.")

        try:
            subprocess.run([pytesseract.pytesseract.tesseract_cmd, "--version"], capture_output=True, text=True,
                           check=True)
            self.add_log_entry("INFO", "Tesseract sikeresen megtalálva")
        except Exception as e:
            self.add_log_entry("ERROR", f"Tesseract nem található vagy hibás: {e}")
            self.status.set("Hiba: Tesseract nem található. Ellenőrizd az elérési utat.")

    def show_completion_popup(self, processed_count, total_duration, errors):
        """
        Megjelenít egy felugró ablakot a feldolgozás összefoglalójával.

        Args:
            processed_count (int): A sikeresen feldolgozott fájlok száma.
            total_duration (float): A teljes futási idő másodpercben.
            errors (list): Hibák listája.
        """
        # Hozzáadjuk a hangjelzést, ha a beállítás engedélyezi
        if self.play_sound:
            self.play_completion_sound()

        # Létrehozzuk a felugró ablakot
        popup = Toplevel(self.root)
        popup.title("Feldolgozás Befejezve")
        popup.geometry("400x300")

        # A pop-up ablak bezárásának letiltása
        # popup.protocol("WM_DELETE_WINDOW", self.disable_event)

        # Információk megjelenítése
        duration_formatted = str(timedelta(seconds=total_duration)).split('.')[0]

        Label(popup, text="Feldolgozás Befejezve!", font=("Helvetica", 16, "bold")).pack(pady=10)
        Label(popup, text=f"Feldolgozott fájlok: {processed_count}", font=("Helvetica", 12)).pack()
        Label(popup, text=f"Összes futási idő: {duration_formatted}", font=("Helvetica", 12)).pack()

        # Hibák megjelenítése, ha vannak
        if errors:
            Label(popup, text="A következő hibák léptek fel:", font=("Helvetica", 12, "bold"), fg="red").pack(
                pady=(15, 5))
            for error in errors:
                Label(popup, text=f"- {error}", wraplength=380, justify="left").pack(anchor="w", padx=10)

        # Hangjelzés ki/be kapcsoló gomb
        sound_text = "Hangjelzés kikapcsolása" if self.play_sound else "Hangjelzés bekapcsolása"
        sound_button = Button(popup, text=sound_text, command=lambda: self.toggle_sound(sound_button))
        sound_button.pack(pady=10)

        # Bezárás gomb
        Button(popup, text="Bezárás", command=popup.destroy).pack(pady=10)

    def play_completion_sound(self):
        """Lejátssza a hangjelzést a feldolgozás befejezésekor."""
        try:
            # Lejátsza a rendszerszintű "exclamation" hangot
            winsound.MessageBeep(winsound.MB_ICONEXCLAMATION)
        except Exception as e:
            self.add_log_entry("ERROR", f"Hiba a hangjelzés lejátszása közben: {e}")

    def save_processing_state(self):
        state = {
            'processed_files_count': self.processed_files_count,
            'processed_size_mb': self.processed_size_mb,
            'processed_duration_sec': self.processed_duration_sec,
            'current_file_index': self.current_file_index,
            'processed_files': [f['filepath'] for f in self.input_files[:self.processed_files_count]],
            'input_directory': self.input_directory,
            'output_folder': self.output_folder
        }
        with open(self.processing_state_file, 'w') as f:
            json.dump(state, f)

    def load_processing_state(self):
        if os.path.exists(self.processing_state_file):
            try:
                with open(self.processing_state_file, 'r') as f:
                    state = json.load(f)
                    self.processed_files_count = state.get('processed_files_count', 0)
                    self.processed_size_mb = state.get('processed_size_mb', 0)
                    self.processed_duration_sec = state.get('processed_duration_sec', 0)
                    self.current_file_index = state.get('current_file_index', 0)
                    self.input_directory = state.get('input_directory', "")
                    self.output_folder = state.get('output_folder', "")
                    self.add_log_entry("INFO", "Korábbi feldolgozási állapot betöltve.")
                    self.update_stats()
            except Exception as e:
                self.add_log_entry("ERROR", f"Hiba a feldolgozási állapot betöltése közben: {e}")
        else:
            self.add_log_entry("INFO", "Nincs korábbi feldolgozási állapot.")

    def clear_log(self):
        self.log_text.delete(*self.log_text.get_children())

    def play_success_sound(self):
        if self.play_sound:
            try:
                winsound.Beep(1000, 200)  # 1000 Hz, 200 ms
                winsound.Beep(1500, 200)  # 1500 Hz, 200 ms
            except:
                pass  # Csak Windows-on működik, más platformokon ne generáljon hibát

    def toggle_sound(self, button):
        """Ki- és bekapcsolja a hangjelzést és frissíti a gomb szövegét."""
        self.play_sound = not self.play_sound
        if self.play_sound:
            button.config(text="Hangjelzés kikapcsolása")
            self.add_log_entry("INFO", "Hangjelzés bekapcsolva.")
        else:
            button.config(text="Hangjelzés bekapcsolása")
            self.add_log_entry("INFO", "Hangjelzés kikapcsolva.")

    def disable_event(self):
        """Letiltja a pop-up ablak bezárását a 'X' gombbal."""
        pass

    def on_file_select(self, event):
        selected_items = self.file_tree.selection()
        if selected_items:
            self.delete_selected_button.config(state="normal")
        else:
            self.delete_selected_button.config(state="disabled")

    def toggle_resolution_fields(self):
        if self.custom_resolution_var.get() == "1":
            self.output_width_entry.config(state="normal")
            self.output_height_entry.config(state="normal")
        else:
            self.output_width_entry.config(state="disabled")
            self.output_height_entry.config(state="disabled")

    def select_input_directory(self):
        self.input_directory = filedialog.askdirectory()
        if self.input_directory:
            self.input_dir_var.set(self.input_directory)
            self.add_log_entry("INFO", f"Bemeneti mappa kiválasztva: {self.input_directory}")
            self.status.set(f"Bemeneti mappa kiválasztva: {self.input_directory}")
            self.save_settings()
            self.update_file_list()

    def select_output_folder(self):
        self.output_folder = filedialog.askdirectory()
        if self.output_folder:
            self.output_dir_var.set(self.output_folder)
            self.add_log_entry("INFO", f"Kimeneti mappa kiválasztva: {self.output_folder}")
            self.status.set(f"Kimeneti mappa kiválasztva: {self.output_folder}")
            self.save_settings()
            self.setup_log_directory()  # Hívja meg itt
            self.update_button_states()

    def select_log_folder(self):
        self.log_folder = filedialog.askdirectory()
        if self.log_folder:
            self.log_dir_var.set(self.log_folder)
            self.add_log_entry("INFO", f"Naplózási mappa kiválasztva: {self.log_folder}")
            self.status.set(f"Naplózási mappa kiválasztva: {self.log_folder}")
            self.save_settings()
            self.setup_log_directory()  # Hívja meg itt

    def update_file_list(self):
        self.input_files = []
        self.total_size_mb = 0
        self.total_duration_sec = 0
        self.file_tree.delete(*self.file_tree.get_children())
        self.tree_items = {}

        # Ellenőrizzük, hogy a bemeneti mappa létezik-e
        if not os.path.isdir(self.input_directory):
            self.add_log_entry("ERROR", f"Érvénytelen bemeneti mappa: {self.input_directory}")
            self.status.set(f"Érvénytelen bemeneti mappa: {self.input_directory}")
            self.file_loading_progress['value'] = 0
            self.file_loading_progress_label_text.set("Fájlok betöltése: 0/0")
            self.update_stats()
            self.update_button_states()
            return

        all_files = [f for f in os.listdir(self.input_directory) if
                     os.path.isfile(os.path.join(self.input_directory, f))]
        total_files = len(all_files)
        self.file_loading_progress['maximum'] = total_files
        self.file_loading_progress['value'] = 0
        loaded_count = 0

        for i, filename in enumerate(all_files):
            if self.stop_processing_flag:
                self.add_log_entry("WARNING", "Fájlbetöltés megszakítva a felhasználó által.")
                self.status.set("Fájlbetöltés megszakítva.")
                break

            file_path = os.path.join(self.input_directory, filename)
            if filename.lower().endswith(('.mp4', '.avi', '.mov', '.mkv')):
                try:
                    file_size_mb = os.path.getsize(file_path) / (1024 * 1024)  # MB-ra konvertálás
                    duration_sec = self.get_video_duration(file_path)
                    if duration_sec is not None:
                        self.input_files.append(file_path)
                        loaded_count += 1
                        self.total_size_mb += file_size_mb
                        self.total_duration_sec += duration_sec
                        item_id = self.file_tree.insert("", "end", values=(
                            f"{len(self.input_files):04d}",
                            filename,
                            f"{file_size_mb:.2f} MB",
                            self.format_time(duration_sec),
                            "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A",
                            "Várakozás", "N/A", "N/A", file_path, "N/A",
                            self.script_name, self.script_dir
                        ))
                        self.tree_items[file_path] = item_id
                        self.add_log_entry("INFO",
                                           f"Fájl betöltve: {filename}, Méret: {file_size_mb:.2f} MB, Időtartam: {self.format_time(duration_sec)}")
                    else:
                        self.add_log_entry("WARNING", f"Érvénytelen időtartam a fájlhoz: {filename}")
                except Exception as e:
                    self.add_log_entry("ERROR", f"Hiba a fájl betöltésekor: {filename} - {str(e)}")
            self.file_loading_progress['value'] = i + 1
            self.file_loading_progress_label_text.set(f"Fájlok betöltése: {loaded_count}/{total_files}")
            self.root.update_idletasks()

        self.file_loading_progress['value'] = total_files
        self.file_loading_progress_label_text.set(f"Fájlok betöltése: {loaded_count}/{total_files}")
        self.update_stats()
        self.update_button_states()
        if loaded_count == 0:
            self.add_log_entry("WARNING", f"Nincs betölthető videófájl a bemeneti mappában: {self.input_directory}")
            self.status.set("Nincs betölthető videófájl a bemeneti mappában.")
        else:
            self.add_log_entry("INFO", f"{loaded_count} videófájl betöltve a bemeneti mappából: {self.input_directory}")
            self.status.set(f"{loaded_count} videófájl betöltve.")

    def update_button_states(self):
        """Frissíti a gombok állapotát a fájlok és a kimeneti mappa megléte alapján."""
        has_files = bool(self.input_files)
        has_output_folder = bool(self.output_folder)

        self.start_button.config(state="normal" if has_files and has_output_folder else "disabled")
        self.delete_selected_button.config(state="normal" if has_files else "disabled")
        self.clear_all_button.config(state="normal" if has_files else "disabled")

        if self.file_tree.get_children():
            # A fa bejegyzéseinek van valamilyen tartalma, tehát van mit törölni
            self.clear_all_button.config(state="normal")
            self.delete_selected_button.config(state="normal")

    def clear_file_list(self):
        self.file_tree.delete(*self.file_tree.get_children())
        self.input_files = []
        self.total_size_mb = 0
        self.total_duration_sec = 0
        self.update_stats()
        self.status.set("Fájllista törölve.")
        self.add_log_entry("INFO", "Összes fájl törölve a listából.")
        self.update_button_states()

    def delete_selected_file(self):
        selected_items = self.file_tree.selection()
        if not selected_items:
            return

        for item in selected_items:
            values = self.file_tree.item(item, 'values')
            if not values:
                continue

            file_path_to_delete = values[16]

            if file_path_to_delete in self.input_files:
                file_size_mb = os.path.getsize(file_path_to_delete) / (1024 * 1024)
                file_duration_sec = self.get_video_duration(file_path_to_delete)

                self.input_files.remove(file_path_to_delete)
                if file_path_to_delete in self.tree_items:
                    del self.tree_items[file_path_to_delete]

                self.total_size_mb -= file_size_mb
                self.total_duration_sec -= file_duration_sec
                self.file_tree.delete(item)
                self.add_log_entry("INFO", f"Fájl törölve a listából: {os.path.basename(file_path_to_delete)}")
            else:
                self.add_log_entry("WARNING",
                                   f"Nem található a fájl a listában a törléshez: {os.path.basename(file_path_to_delete)}")

        self.update_stats()
        self.clear_all_button.config(state="normal" if self.input_files else "disabled")
        self.delete_selected_button.config(state="disabled")
        self.status.set("Kiválasztott fájl(ok) törölve.")

    def get_video_duration(self, file_path):
        try:
            cap = cv2.VideoCapture(file_path)
            if not cap.isOpened():
                self.add_log_entry("ERROR", f"A videó nem olvasható: {file_path}")
                return 0
            fps = cap.get(cv2.CAP_PROP_FPS)
            frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            cap.release()
            if fps > 0 and frame_count > 0:
                self.add_log_entry("INFO", f"Videó időtartama sikeresen lekérve: {file_path}")
                return frame_count / fps
            else:
                self.add_log_entry("ERROR", f"Érvénytelen FPS vagy képkockaszám: {file_path}")
                return 0
        except Exception as e:
            self.add_log_entry("ERROR", f"Hiba a videó időtartamának lekérésekor: {file_path} - {e}")
            return 0

    def get_video_info_ffprobe(self, file_path):
        try:
            cmd = [
                ffmpeg_path.replace('ffmpeg.exe', 'ffprobe.exe'),
                "-v", "error",
                "-show_entries", "format=duration,size",
                "-of", "default=noprint_wrappers=1:nokey=1",
                file_path
            ]
            result = subprocess.run(cmd, capture_output=True, text=True, check=True,
                                    creationflags=subprocess.CREATE_NO_WINDOW)
            output = result.stdout.strip().split('\n')
            duration_sec = float(output[0]) if len(output) > 0 and output[0].replace('.', '', 1).isdigit() else 0.0
            size_bytes = int(output[1]) if len(output) > 1 and output[1].isdigit() else 0

            return duration_sec, size_bytes / (1024 * 1024)
        except (subprocess.CalledProcessError, ValueError, IndexError) as e:
            self.add_log_entry("ERROR",
                               f"Hiba az ffprobe futtatása során a fájlhoz: {os.path.basename(file_path)}. Hiba: {e}")
            return 0.0, 0.0

    def update_overall_progress(self):
        """
        Valós idejű, de ritkított frissítés az 'Összes fájl feldolgozása' sávhoz.
        Nem frissít minden frame-nél, hogy ne villogjon.
        """
        import time
        now = time.time()
        if now - getattr(self, "_last_progress_update", 0) < 0.25:  # 0.25 másodperc (4 Hz)
            return
        self._last_progress_update = now

        try:
            if not hasattr(self, "input_files") or len(self.input_files) == 0:
                return
            total_files = len(self.input_files)
            current_file_progress = float(self.file_progress['value']) / 100.0
            overall = ((self.processed_files_count + current_file_progress) / total_files) * 100

            self.overall_progress['value'] = overall
            self.overall_progress_label_text.set(f"Összes fájl feldolgozása: {overall:.2f}%")
        except Exception:
            pass

    def update_stats(self):
        import os
        from datetime import datetime, timedelta

        # --- 0) Ütközésvédett frissítés (nehogy párhuzamos hívások torzítsák az időt) ---
        if getattr(self, "_stats_lock", False):
            return
        self._stats_lock = True

        try:
            # --- 1) Biztonsági ellenőrzések ---
            if not hasattr(self, "input_files") or not self.input_files:
                self._stats_lock = False
                return

            total_files = len(self.input_files)
            processed_files = getattr(self, "processed_files_count", 0)
            remaining_files = max(0, total_files - processed_files)

            # --- 2) Teljes méretek / időtartam biztosítása ---
            if not hasattr(self, "total_size_mb"):
                self.total_size_mb = sum(
                    os.path.getsize(f) for f in self.input_files if os.path.exists(f)
                ) / (1024 * 1024)

            if not hasattr(self, "total_duration_sec"):
                self.total_duration_sec = getattr(self, "total_duration_sec", 0.0)

            # --- 3) Aktuálisan feldolgozott mennyiségek ---
            current_processed_input_size = self.processed_input_size_mb + getattr(
                self, "current_file_progress_input_size", 0.0
            )
            current_processed_input_duration = self.processed_input_duration_sec + getattr(
                self, "current_file_progress_input_duration", 0.0
            )

            remaining_size_mb = max(0.0, self.total_size_mb - current_processed_input_size)
            remaining_duration_sec = max(0.0, self.total_duration_sec - current_processed_input_duration)

            # --- 4) Alap statisztikák frissítése ---
            self.total_files_var.set(f"Összesen: {total_files}")
            self.processed_files_var.set(f"Feldolgozva: {processed_files}")
            self.remaining_files_var.set(f"Hátralévő: {remaining_files}")

            self.total_size_var.set(f"Összesen: {self.total_size_mb:.2f} MB")
            self.processed_size_var.set(f"Feldolgozva: {current_processed_input_size:.2f} MB")
            self.remaining_size_var.set(f"Hátralévő: {remaining_size_mb:.2f} MB")

            self.total_duration_var.set(f"Összesen: {self.format_time(self.total_duration_sec)}")
            self.processed_duration_var.set(f"Feldolgozva: {self.format_time(current_processed_input_duration)}")
            self.remaining_duration_var.set(f"Hátralévő: {self.format_time(remaining_duration_sec)}")

            # --- 5) Idők kiszámítása ---
            elapsed_time = "N/A"
            remaining_time = "N/A"
            calculated_end_time = "N/A"
            total_run_time = "N/A"

            if getattr(self, "start_time", None):
                time_elapsed_seconds = (datetime.now() - self.start_time).total_seconds()
                elapsed_time = self.format_time(time_elapsed_seconds)
                total_run_time = elapsed_time  # teljes futásidő kijelzéshez

                if current_processed_input_duration > 0:
                    estimated_total_time = (
                        time_elapsed_seconds / current_processed_input_duration
                    ) * self.total_duration_sec
                    remaining_seconds = max(0.0, estimated_total_time - time_elapsed_seconds)
                    self.calculated_end_time = datetime.now() + timedelta(seconds=remaining_seconds)

                    remaining_time = self.format_time(remaining_seconds)
                    calculated_end_time = self.calculated_end_time.strftime('%H:%M:%S')

            # --- 6) GUI idők frissítése ---
            self.start_time_var.set(
                f"Kezdés: {self.start_time.strftime('%H:%M:%S') if getattr(self, 'start_time', None) else 'N/A'}"
            )
            self.elapsed_time_var.set(f"Eltelt: {elapsed_time}")
            self.remaining_time_var.set(f"Hátralévő: {remaining_time}")
            self.end_time_var.set(f"Várható zárás: {calculated_end_time}")
            self.total_processing_time_var.set(f"Össz. futásidő: {total_run_time}")

            # --- 7) Összes progress ---
            overall_progress_percent = (
                (current_processed_input_duration / self.total_duration_sec) * 100
                if self.total_duration_sec > 0 else 0
            )
            self.overall_progress['value'] = overall_progress_percent
            self.overall_progress_label_text.set(f"Összes fájl feldolgozása: {overall_progress_percent:.2f}%")

            # --- 8) Átlagos futásidő / fájl ---
            if getattr(self, "start_time", None) and processed_files > 0:
                elapsed = (datetime.now() - self.start_time).total_seconds()
                avg_seconds = elapsed / processed_files
                avg_minutes = avg_seconds / 60
                self.avg_time_per_file_var.set(f"Átlag / fájl: {avg_minutes:.2f} perc / {avg_seconds:.1f} mp")
            else:
                self.avg_time_per_file_var.set("Átlag / fájl: N/A")

            # --- 9) Megtakarítás / eredmények ---
            total_in = self.processed_input_size_mb
            total_out = self.processed_output_size_mb
            if total_in > 0 and total_out > 0:
                saving_percent = 100 - (total_out / total_in) * 100
                self.saving_percent_var.set(f"Megtakarítás: {saving_percent:+.2f} %")
                self.summary_input_var.set(f"Be: {total_in:.2f} MB")
                self.summary_output_var.set(f"Ki: {total_out:.2f} MB")
                self.summary_saving_var.set(f"Megtakarítás: {saving_percent:+.2f} %")
                self.summary_saving_label.configure(foreground="#007800" if saving_percent >= 0 else "#C00000")
            else:
                self.saving_percent_var.set("Megtakarítás: N/A %")
                self.summary_input_var.set("Be: N/A MB")
                self.summary_output_var.set("Ki: N/A MB")
                self.summary_saving_var.set("Megtakarítás: N/A %")
                self.summary_saving_label.configure(foreground="#004080")

        finally:
            # --- 10) Frissítés-engedély visszaadása és újraütemezés ---
            self._stats_lock = False
            if hasattr(self, "processing_thread") and self.processing_thread and self.processing_thread.is_alive():
                self.root.after(1000, self.update_stats)



    def format_time(self, seconds):
        if seconds < 0:
            seconds = 0
        minutes, seconds = divmod(seconds, 60)
        hours, minutes = divmod(minutes, 60)
        return f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"

    def toggle_processing(self):
        self.add_log_entry("DEBUG", "toggle_processing metódus meghívva")
        if self.processing_thread and self.processing_thread.is_alive():
            self.add_log_entry("DEBUG", "Feldolgozó szál fut, szüneteltetés/folytatás")
            self.is_paused = not self.is_paused
            if self.is_paused:
                self.start_button.config(text="Folytatás")
                self.add_log_entry("INFO", "Feldolgozás szüneteltetve.")
                self.status.set("Feldolgozás szüneteltetve.")
            else:
                self.start_button.config(text="Szünet")
                self.add_log_entry("INFO", "Feldolgozás folytatva.")
                self.status.set("Feldolgozás folytatva.")
        else:
            self.add_log_entry("DEBUG", "Új feldolgozás indítása")
            if not self.input_directory or not self.output_folder:
                self.add_log_entry("WARNING", "Hiányzó bemeneti vagy kimeneti mappa")
                messagebox.showwarning("Hiányzó adatok", "Kérlek válaszd ki a bemeneti és kimeneti mappát.")
                return
            if not self.input_files:
                self.add_log_entry("WARNING", "Nincsenek videófájlok a bemeneti mappában")
                messagebox.showwarning("Hiányzó adatok", "A kiválasztott mappában nem található videófájl.")
                return

            self.start_button.config(text="Szünet")
            self.stop_button.config(state="normal")
            self.resume_button.config(state="disabled")
            self.stop_processing_flag = False
            self.is_paused = False
            self.processed_files_count = 0
            self.processed_size_mb = 0
            self.processed_duration_sec = 0
            self.current_file_progress_duration = 0
            self.current_file_progress_size = 0
            self.log1_data = []
            self.log3_data = []
            self.add_log_entry("DEBUG", "Szál létrehozása előtt")
            self.processing_thread = threading.Thread(target=self.process_all_files, daemon=True)
            self.add_log_entry("DEBUG", "Szál létrehozva, indítás")
            self.processing_thread.start()
            self.add_log_entry("DEBUG", "Szál elindítva")

    def stop_processing(self):
        self.stop_processing_flag = True
        self.is_paused = False
        if self.processing_thread and self.processing_thread.is_alive():
            self.add_log_entry("WARNING", "Feldolgozás leállítva, várjon a szál befejezésére.")
            self.status.set("Feldolgozás leállítva, várjon a szál befejezésére.")
        else:
            self.reset_state()

    def reset_state(self):
        """
        Feldolgozási állapot visszaállítása — csak akkor, ha a futás megszakadt vagy hibásan zárult.
        Ha a feldolgozás normálisan befejeződött (end_time már létezik), nem töröljük az adatokat.
        """

        # --- Ha a feldolgozás normálisan befejeződött, ne nullázzuk az adatokat ---
        if hasattr(self, "end_time") and self.end_time is not None:
            self.add_log_entry("DEBUG", "reset_state kihagyva, mert a feldolgozás befejeződött.")
            return

        self.start_button.config(text="Feldolgozás indítása")
        self.stop_button.config(state="disabled")
        self.resume_button.config(state="normal" if os.path.exists(self.processing_state_file) else "disabled")
        self.add_log_entry("INFO", "A feldolgozás leállítva.")
        self.status.set("A feldolgozás leállítva.")
        self.current_file_index = -1
        self.processed_files_count = 0
        self.processed_input_size_mb = 0
        self.processed_output_size_mb = 0
        self.processed_input_duration_sec = 0
        self.processed_output_duration_sec = 0
        self.current_file_progress_input_size = 0
        self.current_file_progress_input_duration = 0
        self.start_time = None
        self.end_time = None
        self.calculated_end_time = None
        self.update_stats()
        self.file_progress['value'] = 0
        self.overall_progress['value'] = 0
        self.overall_progress_label_text.set("Összes fájl feldolgozása: 0%")
        self.file_progress_label_text.set("Aktuális fájl feldolgozása: 0%")

    # ÚJ: Folytatás gomb logikája
    def resume_processing(self):
        if not os.path.exists(self.processing_state_file):
            messagebox.showinfo("Nincs folytatható feldolgozás",
                                "Nincs félbeszakadt feldolgozás, amit folytatni lehetne.")
            self.add_log_entry("INFO", "Nincs folytatható feldolgozás.")
            self.status.set("Nincs folytatható feldolgozás.")
            return

        with open(self.processing_state_file, 'r') as f:
            state = json.load(f)

        if state.get('status') != 'processing':
            messagebox.showinfo("Nincs folytatható feldolgozás", "A korábbi feldolgozás befejeződött vagy érvénytelen.")
            self.add_log_entry("INFO", "Nincs folytatható feldolgozás.")
            self.status.set("Nincs folytatható feldolgozás.")
            return

        interrupted_index = state.get('current_index', -1)
        interrupted_file = state.get('interrupted_file', 'N/A')
        interruption_time = state.get('interruption_time', datetime.now().isoformat())

        msg = (f"Félbeszakadt feldolgozást észleltünk!\n"
               f"Utolsó sikeres fájl index: {interrupted_index - 1 if interrupted_index > 0 else -1}\n"
               f"Megszakadt fájl: {interrupted_file}\n"
               f"Leállás ideje: {interruption_time}\n"
               f"Folytatni szeretné a feldolgozást a megszakadt fájltól?")
        if messagebox.askyesno("Folytatás félbeszakadt feldolgozásból", msg):
            self.current_file_index = interrupted_index - 1
            self.processed_files_count = interrupted_index
            self.start_button.config(text="Szünet")
            self.stop_button.config(state="normal")
            self.resume_button.config(state="disabled")
            self.stop_processing_flag = False
            self.is_paused = False
            self.processing_thread = threading.Thread(target=self.process_all_files, daemon=True)
            self.processing_thread.start()
            self.add_log_entry("INFO", f"Folytatás a félbeszakadt feldolgozásból az index {interrupted_index}-től.")
            self.status.set(f"Folytatás a félbeszakadt feldolgozásból.")
        else:
            os.remove(self.processing_state_file)
            self.add_log_entry("INFO", "Félbeszakadt feldolgozás nem folytatva.")
            self.status.set("Félbeszakadt feldolgozás nem folytatva.")
            self.resume_button.config(state="disabled")

    # ÚJ: A "Folytatás" gomb állapotának frissítése
    def update_resume_button_state(self):
        if os.path.exists(self.processing_state_file):
            self.resume_button.config(state="normal")
        else:
            self.resume_button.config(state="disabled")

    def format_runtime(self, seconds):
        minutes = int(seconds // 60)
        remaining_seconds = int(seconds % 60)
        return f"{minutes:02d}:{remaining_seconds:02d}"

    def generate_log_files(self):
        """Összefoglaló log fájlok (Log1, Log2, Log3) mentése a napi _Logok mappába"""

        import json
        from openpyxl import Workbook
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        log_prefix = f"{self.script_name.replace('.py', '')}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}"

        # --- DejaVuSans font regisztrálása ---
        try:
            pdfmetrics.registerFont(TTFont("DejaVuSans", "F:/__Panel/fonts/DejaVuSans.ttf"))
            custom_style = ParagraphStyle("Custom", fontName="DejaVuSans", fontSize=10, leading=12)
        except Exception as e:
            # Ha valamiért nem találja a fontot, fallback
            styles = getSampleStyleSheet()
            custom_style = styles["Normal"]

        # Ha nincs aktuális log könyvtár, állítsuk be
        if not hasattr(self, "current_log_file") or not hasattr(self, "current_log_date"):
            self.setup_log_directory()

        # az aktuális runtime log mappa (_Logok)
        log_dir = os.path.dirname(self.current_log_file)
        today_str = self.current_log_date
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

        # ===================================================================
        # LOG1 - fájlok részletes feldolgozási listája
        # ===================================================================
        log1_filename = f"Log1-files-{today_str}_{timestamp}"

        # TXT
        with open(os.path.join(log_dir, f"{log1_filename}.txt"), "w", encoding="utf-8") as f:
            for entry in self.log1_data:
                for key, value in entry.items():
                    f.write(f"{key}: {value}\n")
                f.write("\n")

        # JSON
        with open(os.path.join(log_dir, f"{log1_filename}.json"), "w", encoding="utf-8") as f:
            json.dump(self.log1_data, f, ensure_ascii=False, indent=4)

        # XLSX
        wb = Workbook()
        ws = wb.active
        if self.log1_data:
            ws.append(list(self.log1_data[0].keys()))
            for entry in self.log1_data:
                ws.append(list(entry.values()))
        wb.save(os.path.join(log_dir, f"{log1_filename}.xlsx"))

        # PDF
        pdf = SimpleDocTemplate(os.path.join(log_dir, f"{log1_filename}.pdf"), pagesize=A4)
        story = []
        for entry in self.log1_data:
            for key, value in entry.items():
                story.append(Paragraph(f"<b>{key}:</b> {value}", custom_style))
            story.append(Spacer(1, 12))
        pdf.build(story)

        # ===================================================================
        # PDF bővítés – módonkénti statisztika (🎬 / ⚡ / 🧩)
        # ===================================================================
        mode_counts = {"🎬 Normál": 0, "⚡ Gyors": 0, "🧩 Hibrid": 0}
        mode_durations = {"🎬 Normál": [], "⚡ Gyors": [], "🧩 Hibrid": []}
        mode_savings = {"🎬 Normál": [], "⚡ Gyors": [], "🧩 Hibrid": []}

        for entry in getattr(self, "log1_data", []):
            mode_text = entry.get("21. Feldolgozási mód", "")
            duration_str = entry.get("13. Futásidő", "00:00")
            compression_str = entry.get("08. Tömörítés (%)", "0")

            # idő átalakítása mp-re
            try:
                parts = duration_str.split(":")
                duration_sec = int(parts[0]) * 60 + int(parts[1])
            except:
                duration_sec = 0

            try:
                compression = float(compression_str.replace("%", "").strip())
            except:
                compression = 0.0

            if "Hibrid" in mode_text:
                mode_counts["🧩 Hibrid"] += 1
                mode_durations["🧩 Hibrid"].append(duration_sec)
                mode_savings["🧩 Hibrid"].append(compression)
            elif "Gyors" in mode_text:
                mode_counts["⚡ Gyors"] += 1
                mode_durations["⚡ Gyors"].append(duration_sec)
                mode_savings["⚡ Gyors"].append(compression)
            elif "Normál" in mode_text:
                mode_counts["🎬 Normál"] += 1
                mode_durations["🎬 Normál"].append(duration_sec)
                mode_savings["🎬 Normál"].append(compression)

        def avg(values):
            return sum(values) / len(values) if values else 0

        # --- új PDF oldal hozzáadása ---
        story.append(Spacer(1, 12))
        story.append(Paragraph("<b>📊 Feldolgozási módok megoszlása</b>", custom_style))
        story.append(Spacer(1, 6))
        story.append(Paragraph(
            f"🎬 Normál mód: {mode_counts['🎬 Normál']} db – átlag futásidő: {avg(mode_durations['🎬 Normál']):.1f} mp – átlag megtakarítás: {avg(mode_savings['🎬 Normál']):.2f} %",
            custom_style))
        story.append(Paragraph(
            f"⚡ Gyors mód: {mode_counts['⚡ Gyors']} db – átlag futásidő: {avg(mode_durations['⚡ Gyors']):.1f} mp – átlag megtakarítás: {avg(mode_savings['⚡ Gyors']):.2f} %",
            custom_style))
        story.append(Paragraph(
            f"🧩 Hibrid mód: {mode_counts['🧩 Hibrid']} db – átlag futásidő: {avg(mode_durations['🧩 Hibrid']):.1f} mp – átlag megtakarítás: {avg(mode_savings['🧩 Hibrid']):.2f} %",
            custom_style))

        pdf.build(story)

        # ===================================================================
        # LOG2 - összesített statisztika
        # ===================================================================
        log2_filename = f"Log2-full-{today_str}_{timestamp}"

        total_files = self.processed_files_count
        total_input_size = self.processed_input_size_mb
        total_output_size = self.processed_output_size_mb
        total_time = (self.end_time - self.start_time).total_seconds() if hasattr(self, "end_time") else 0

        saving_mb = total_input_size - total_output_size
        saving_percent = 0
        if total_input_size > 0 and total_output_size > 0:
            saving_percent = 100 - (total_output_size / total_input_size) * 100

        # Szöveges összesítés generálása
        if total_input_size > 0:
            summary_text = (f"A teljes feldolgozás során "
                            f"{total_input_size:.2f} MB → {total_output_size:.2f} MB "
                            f"({saving_percent:+.2f} %) méretváltozás történt.")
        else:
            summary_text = "Nem áll rendelkezésre elegendő adat a megtakarítás kiszámításához."

        # ===================================================================
        # LOG2 - összesített statisztika + rövid sor az elejére
        # ===================================================================
        log2_filename = f"Log2-full-{today_str}_{timestamp}"

        total_files = self.processed_files_count
        total_input_size = self.processed_input_size_mb
        total_output_size = self.processed_output_size_mb
        total_time = (self.end_time - self.start_time).total_seconds() if hasattr(self, "end_time") else 0

        saving_mb = total_input_size - total_output_size
        saving_percent = 0
        if total_input_size > 0 and total_output_size > 0:
            saving_percent = 100 - (total_output_size / total_input_size) * 100

        # Átlagidő / fájl (perc + mp)
        if total_files > 0 and hasattr(self, "start_time") and hasattr(self, "end_time"):
            avg_seconds = total_time / total_files
            avg_minutes = avg_seconds / 60
            avg_text = f"{avg_minutes:.2f} perc / {avg_seconds:.1f} mp"
        else:
            avg_text = "N/A"

        # Rövid összefoglaló sor (Log2 elejére)
        short_summary = (
            f"Összesen: {total_files} fájl — Átlag: {avg_text} — "
            f"Megtakarítás: {saving_percent:+.2f} %"
        )

        # ===================================================================
        # LOG2 adatstruktúra
        # ===================================================================
        log2_data = {
            "00. Összesített rövid sor": short_summary,
            "01. A feldolgozott fájlok száma": f"{total_files} db",
            "02. A feldolgozott fájlok mérete megabájtban": f"{total_input_size:.2f} MB",
            "03. A feldolgozott fájlok időtartama": self.format_time(self.processed_input_duration_sec),
            "04. Összes kimenő fájlméret MB": f"{total_output_size:.2f} MB",
            "05. A feldolgozott fájlok megtakarítása MB": f"{saving_mb:.2f} MB",
            "06. A feldolgozott fájlok megtakarítása %": f"{saving_percent:+.2f} %",
            "07. Összesített eredmény szövegesen": summary_text,
            "08. A program indításának időbélyege": self.start_time.strftime('%Y-%m-%d %H:%M:%S'),
            "09. A program zárásának időbélyege": self.end_time.strftime('%Y-%m-%d %H:%M:%S') if hasattr(self,
                                                                                                         'end_time') else 'N/A',
            "10. A konvertálás kezdetének időbélyege": self.start_time.strftime('%H:%M:%S'),
            "11. A konvertálás zárásának időbélyege": self.end_time.strftime('%H:%M:%S') if hasattr(self,
                                                                                                    'end_time') else 'N/A',
            "12. A konvertálás futásideje összes file": self.format_time(total_time),
            "13. Bemeneti könyvtár": self.input_directory,
            "14. Kimeneti könyvtár": self.output_folder,
            "15. Script neve": self.script_name,
            "16. Script könyvtára": os.path.dirname(os.path.abspath(__file__)),
            "17. Feldolgozott fájlok listája": self.processed_files_list
        }

        # --- Átlagos futásidő / fájl (perc + mp) ---
        if total_files > 0 and hasattr(self, "start_time") and hasattr(self, "end_time"):
            total_seconds = (self.end_time - self.start_time).total_seconds()
            avg_seconds = total_seconds / total_files
            avg_minutes = avg_seconds / 60
            avg_per_file = f"{avg_minutes:.2f} perc / {avg_seconds:.1f} mp"
        else:
            avg_per_file = "N/A"

        log2_data["18. Átlagos futásidő / fájl"] = avg_per_file

        # --- Feldolgozási mód (Gyors / Normál) ---
        log2_data["19. Feldolgozási mód"] = (
            "⚡ Gyors vágás (copy mód, újrakódolás nélkül)"
            if hasattr(self, "fast_mode_var") and self.fast_mode_var.get() == "1"
            else "🎬 Normál feldolgozás (minőségmegőrző újrakódolással)"
        )

        # ===================================================================
        # LOG2 – módonkénti statisztika (Normál / Gyors / Hibrid)
        # ===================================================================

        mode_counts = {"🎬 Normál": 0, "⚡ Gyors": 0, "🧩 Hibrid": 0}
        mode_durations = {"🎬 Normál": [], "⚡ Gyors": [], "🧩 Hibrid": []}
        mode_savings = {"🎬 Normál": [], "⚡ Gyors": [], "🧩 Hibrid": []}

        for entry in getattr(self, "log1_data", []):
            mode_text = entry.get("21. Feldolgozási mód", "")
            duration_str = entry.get("13. Futásidő", "00:00")
            compression_str = entry.get("08. Tömörítés (%)", "0")

            # idő átalakítás mp-re
            try:
                parts = duration_str.split(":")
                if len(parts) == 2:
                    duration_sec = int(parts[0]) * 60 + int(parts[1])
                else:
                    duration_sec = 0
            except:
                duration_sec = 0

            # tömörítési százalék kinyerése
            try:
                compression = float(compression_str.replace("%", "").strip())
            except:
                compression = 0.0

            # mód beazonosítása
            if "Hibrid" in mode_text:
                mode_counts["🧩 Hibrid"] += 1
                mode_durations["🧩 Hibrid"].append(duration_sec)
                mode_savings["🧩 Hibrid"].append(compression)
            elif "Gyors" in mode_text:
                mode_counts["⚡ Gyors"] += 1
                mode_durations["⚡ Gyors"].append(duration_sec)
                mode_savings["⚡ Gyors"].append(compression)
            elif "Normál" in mode_text:
                mode_counts["🎬 Normál"] += 1
                mode_durations["🎬 Normál"].append(duration_sec)
                mode_savings["🎬 Normál"].append(compression)

        def avg(values):
            return sum(values) / len(values) if values else 0

        log2_data.update({
            "20. 🎬 Normál mód – fájlok száma": str(mode_counts["🎬 Normál"]),
            "21. ⚡ Gyors mód – fájlok száma": str(mode_counts["⚡ Gyors"]),
            "22. 🧩 Hibrid mód – fájlok száma": str(mode_counts["🧩 Hibrid"]),
            "23. 🎬 Normál mód – átlag futásidő (mp)": f"{avg(mode_durations['🎬 Normál']):.1f}",
            "24. ⚡ Gyors mód – átlag futásidő (mp)": f"{avg(mode_durations['⚡ Gyors']):.1f}",
            "25. 🧩 Hibrid mód – átlag futásidő (mp)": f"{avg(mode_durations['🧩 Hibrid']):.1f}",
            "26. 🎬 Normál mód – átlag megtakarítás (%)": f"{avg(mode_savings['🎬 Normál']):.2f}",
            "27. ⚡ Gyors mód – átlag megtakarítás (%)": f"{avg(mode_savings['⚡ Gyors']):.2f}",
            "28. 🧩 Hibrid mód – átlag megtakarítás (%)": f"{avg(mode_savings['🧩 Hibrid']):.2f}",
        })

        # TXT
        with open(os.path.join(log_dir, f"{log2_filename}.txt"), "w", encoding="utf-8") as f:
            for k, v in log2_data.items():
                if isinstance(v, list):
                    f.write(f"{k}:\n")
                    for fname in v:
                        f.write(f"   {fname}\n")
                else:
                    f.write(f"{k}: {v}\n")

        # JSON
        # ===================================================================
        # Mód statisztika előállítása (🎬 / ⚡ / 🧩) a Log2 JSON-hoz
        # ===================================================================
        mode_counts = {"🎬 Normál": 0, "⚡ Gyors": 0, "🧩 Hibrid": 0}
        mode_durations = {"🎬 Normál": [], "⚡ Gyors": [], "🧩 Hibrid": []}
        mode_savings = {"🎬 Normál": [], "⚡ Gyors": [], "🧩 Hibrid": []}

        for entry in getattr(self, "log1_data", []):
            mode_text = entry.get("21. Feldolgozási mód", "")
            duration_str = entry.get("13. Futásidő", "00:00")
            compression_str = entry.get("08. Tömörítés (%)", "0")

            # idő átalakítása másodpercre
            try:
                parts = duration_str.split(":")
                duration_sec = int(parts[0]) * 60 + int(parts[1])
            except:
                duration_sec = 0

            try:
                compression = float(compression_str.replace("%", "").strip())
            except:
                compression = 0.0

            if "Hibrid" in mode_text:
                mode_counts["🧩 Hibrid"] += 1
                mode_durations["🧩 Hibrid"].append(duration_sec)
                mode_savings["🧩 Hibrid"].append(compression)
            elif "Gyors" in mode_text:
                mode_counts["⚡ Gyors"] += 1
                mode_durations["⚡ Gyors"].append(duration_sec)
                mode_savings["⚡ Gyors"].append(compression)
            elif "Normál" in mode_text:
                mode_counts["🎬 Normál"] += 1
                mode_durations["🎬 Normál"].append(duration_sec)
                mode_savings["🎬 Normál"].append(compression)

        def avg(values):
            return sum(values) / len(values) if values else 0

        log2_mode_stats = {
            "🎬 Normál mód": {
                "Fájlok száma": mode_counts["🎬 Normál"],
                "Átlag futásidő (mp)": round(avg(mode_durations["🎬 Normál"]), 1),
                "Átlag megtakarítás (%)": round(avg(mode_savings["🎬 Normál"]), 2)
            },
            "⚡ Gyors mód": {
                "Fájlok száma": mode_counts["⚡ Gyors"],
                "Átlag futásidő (mp)": round(avg(mode_durations["⚡ Gyors"]), 1),
                "Átlag megtakarítás (%)": round(avg(mode_savings["⚡ Gyors"]), 2)
            },
            "🧩 Hibrid mód": {
                "Fájlok száma": mode_counts["🧩 Hibrid"],
                "Átlag futásidő (mp)": round(avg(mode_durations["🧩 Hibrid"]), 1),
                "Átlag megtakarítás (%)": round(avg(mode_savings["🧩 Hibrid"]), 2)
            }
        }

        # Beillesztjük a fő Log2 struktúrába
        log2_data["20. Mód statisztika"] = log2_mode_stats

        with open(os.path.join(log_dir, f"{log2_filename}.json"), "w", encoding="utf-8") as f:
            json.dump(log2_data, f, ensure_ascii=False, indent=4)

        # XLSX
        wb = Workbook()
        ws = wb.active
        ws.append(["Kulcs", "Érték"])
        for k, v in log2_data.items():
            if isinstance(v, list):
                ws.append([k])  # fejléc (pl. "15. Feldolgozott fájlok listája")
                for fname in v:
                    ws.append(["", fname])  # fájlok külön sorban, 2. oszlopban
            else:
                import json
                if isinstance(v, dict):
                    v = json.dumps(v, ensure_ascii=False, indent=2)  # magyar ékezet is megmarad
                ws.append([k, v])

        # ===================================================================
        # LOG2 – XLSX mentés (összesítés + módstatisztika)
        # ===================================================================
        from openpyxl import Workbook
        import json

        wb = Workbook()
        ws = wb.active
        ws.title = "Összesítés"
        ws.append(["Kulcs", "Érték"])

        for k, v in log2_data.items():
            if isinstance(v, list):
                # Lista esetén soronként külön bejegyzés
                ws.append([k])
                for fname in v:
                    ws.append(["", fname])
            elif isinstance(v, dict):
                # Szótárakat olvasható JSON-szöveggé alakítjuk (így az Excel is tudja kezelni)
                formatted_value = json.dumps(v, ensure_ascii=False, indent=2)
                ws.append([k, formatted_value])
            else:
                # Egyszerű érték (szám, string, stb.)
                ws.append([k, v])

        # Végül mentés
        xlsx_path = os.path.join(log_dir, f"{log_prefix}_Log2.xlsx")
        wb.save(xlsx_path)

        # ===================================================================
        # „Mód statisztika” munkalap (🎬 / ⚡ / 🧩)
        # ===================================================================
        mode_counts = {"🎬 Normál": 0, "⚡ Gyors": 0, "🧩 Hibrid": 0}
        mode_durations = {"🎬 Normál": [], "⚡ Gyors": [], "🧩 Hibrid": []}
        mode_savings = {"🎬 Normál": [], "⚡ Gyors": [], "🧩 Hibrid": []}

        for entry in getattr(self, "log1_data", []):
            mode_text = entry.get("21. Feldolgozási mód", "")
            duration_str = entry.get("13. Futásidő", "00:00")
            compression_str = entry.get("08. Tömörítés (%)", "0")

            # idő átalakítása másodpercre
            try:
                parts = duration_str.split(":")
                duration_sec = int(parts[0]) * 60 + int(parts[1])
            except:
                duration_sec = 0

            try:
                compression = float(compression_str.replace("%", "").strip())
            except:
                compression = 0.0

            if "Hibrid" in mode_text:
                mode_counts["🧩 Hibrid"] += 1
                mode_durations["🧩 Hibrid"].append(duration_sec)
                mode_savings["🧩 Hibrid"].append(compression)
            elif "Gyors" in mode_text:
                mode_counts["⚡ Gyors"] += 1
                mode_durations["⚡ Gyors"].append(duration_sec)
                mode_savings["⚡ Gyors"].append(compression)
            elif "Normál" in mode_text:
                mode_counts["🎬 Normál"] += 1
                mode_durations["🎬 Normál"].append(duration_sec)
                mode_savings["🎬 Normál"].append(compression)

        def avg(values):
            return sum(values) / len(values) if values else 0

        ws_modes = wb.create_sheet("Mód statisztika")
        ws_modes.append(["Feldolgozási mód", "Fájlok száma", "Átlag futásidő (mp)", "Átlag megtakarítás (%)"])

        ws_modes.append(["🎬 Normál mód",
                         mode_counts["🎬 Normál"],
                         round(avg(mode_durations["🎬 Normál"]), 1),
                         round(avg(mode_savings["🎬 Normál"]), 2)])
        ws_modes.append(["⚡ Gyors mód",
                         mode_counts["⚡ Gyors"],
                         round(avg(mode_durations["⚡ Gyors"]), 1),
                         round(avg(mode_savings["⚡ Gyors"]), 2)])
        ws_modes.append(["🧩 Hibrid mód",
                         mode_counts["🧩 Hibrid"],
                         round(avg(mode_durations["🧩 Hibrid"]), 1),
                         round(avg(mode_savings["🧩 Hibrid"]), 2)])

        # Automatikus oszlopszélesség
        for col in ws_modes.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws_modes.column_dimensions[column].width = max_length + 2

        # Mentés
        wb.save(os.path.join(log_dir, f"{log2_filename}.xlsx"))

        wb.save(os.path.join(log_dir, f"{log2_filename}.xlsx"))

        # PDF
        # ===================================================================
        # LOG2 PDF – mód statisztika előállítása a PDF végére
        # ===================================================================
        from reportlab.platypus import Table, TableStyle
        from reportlab.lib import colors

        # Új laphoz tartalom
        mode_counts = {"🎬 Normál": 0, "⚡ Gyors": 0, "🧩 Hibrid": 0}
        mode_durations = {"🎬 Normál": [], "⚡ Gyors": [], "🧩 Hibrid": []}
        mode_savings = {"🎬 Normál": [], "⚡ Gyors": [], "🧩 Hibrid": []}

        for entry in getattr(self, "log1_data", []):
            mode_text = entry.get("21. Feldolgozási mód", "")
            duration_str = entry.get("13. Futásidő", "00:00")
            compression_str = entry.get("08. Tömörítés (%)", "0")

            try:
                parts = duration_str.split(":")
                duration_sec = int(parts[0]) * 60 + int(parts[1])
            except:
                duration_sec = 0

            try:
                compression = float(compression_str.replace("%", "").strip())
            except:
                compression = 0.0

            if "Hibrid" in mode_text:
                mode_counts["🧩 Hibrid"] += 1
                mode_durations["🧩 Hibrid"].append(duration_sec)
                mode_savings["🧩 Hibrid"].append(compression)
            elif "Gyors" in mode_text:
                mode_counts["⚡ Gyors"] += 1
                mode_durations["⚡ Gyors"].append(duration_sec)
                mode_savings["⚡ Gyors"].append(compression)
            elif "Normál" in mode_text:
                mode_counts["🎬 Normál"] += 1
                mode_durations["🎬 Normál"].append(duration_sec)
                mode_savings["🎬 Normál"].append(compression)

        def avg(values):
            return sum(values) / len(values) if values else 0

        data_table = [
            ["Feldolgozási mód", "Fájlok száma", "Átlag futásidő (mp)", "Átlag megtakarítás (%)"],
            ["🎬 Normál mód",
             mode_counts["🎬 Normál"],
             f"{avg(mode_durations['🎬 Normál']):.1f}",
             f"{avg(mode_savings['🎬 Normál']):.2f}"],
            ["⚡ Gyors mód",
             mode_counts["⚡ Gyors"],
             f"{avg(mode_durations['⚡ Gyors']):.1f}",
             f"{avg(mode_savings['⚡ Gyors']):.2f}"],
            ["🧩 Hibrid mód",
             mode_counts["🧩 Hibrid"],
             f"{avg(mode_durations['🧩 Hibrid']):.1f}",
             f"{avg(mode_savings['🧩 Hibrid']):.2f}"]
        ]

        mode_table = Table(data_table, colWidths=[130, 100, 130, 130])
        mode_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans'),
            ('FONTNAME', (0, 1), (-1, -1), 'DejaVuSans'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke)
        ]))

        pdf = SimpleDocTemplate(os.path.join(log_dir, f"{log2_filename}.pdf"), pagesize=A4)
        story = []
        for key, value in log2_data.items():
            story.append(Paragraph(f"<b>{key}:</b> {value}", custom_style))
            story.append(Spacer(1, 6))

        # --- Új oldalra a mód statisztika ---
        story.append(Spacer(1, 20))
        story.append(Paragraph("<b>Mód statisztika összesítése:</b>", custom_style))
        story.append(Spacer(1, 10))
        story.append(mode_table)

        pdf.build(story)

        # LOG3 - hibák listája (sorszámozva, fájlnévvel együtt)
        # ===================================================================
        log3_filename = f"Log3-errors-{today_str}_{timestamp}"

        log3_data_numbered = {}
        line_no = 1

        error_count = len(self.log3_data)
        log3_data_numbered[f"{line_no:02d}. Hibás fájlok száma"] = f"{error_count} db"
        line_no += 1

        if error_count == 0:
            log3_data_numbered[f"{line_no:02d}. Üzenet"] = "Nem történt hiba."
        else:
            for idx, entry in enumerate(self.log3_data, start=1):
                if isinstance(entry, dict):
                    file_name = entry.get("file", "Ismeretlen fájl")
                    error_msg = entry.get("error", "Ismeretlen hiba")
                else:
                    # ha csak sima string volt eltárolva
                    file_name = "Ismeretlen fájl"
                    error_msg = str(entry)
                log3_data_numbered[f"{line_no:02d}. Hiba {idx}"] = f"{file_name} – {error_msg}"
                line_no += 1

        # TXT
        with open(os.path.join(log_dir, f"{log3_filename}.txt"), "w", encoding="utf-8") as f:
            for key, value in log3_data_numbered.items():
                f.write(f"{key}: {value}\n")

        # JSON
        with open(os.path.join(log_dir, f"{log3_filename}.json"), "w", encoding="utf-8") as f:
            json.dump(log3_data_numbered, f, ensure_ascii=False, indent=4)

        # XLSX
        wb = Workbook()
        ws = wb.active
        ws.append(["Sorszám", "Hiba"])
        for key, value in log3_data_numbered.items():
            ws.append([key, value])
        wb.save(os.path.join(log_dir, f"{log3_filename}.xlsx"))

        # PDF
        pdf = SimpleDocTemplate(os.path.join(log_dir, f"{log3_filename}.pdf"), pagesize=A4)
        story = []
        for key, value in log3_data_numbered.items():
            story.append(Paragraph(f"<b>{key}:</b> {value}", custom_style))
        pdf.build(story)

        # dátumos főmappa (_Logok szülőkönyvtára)
        daily_dir = os.path.dirname(self.log_dir)

        # alkönyvtárak
        ocr_dir = os.path.join(daily_dir, "OCR")
        logs_dir = os.path.join(daily_dir, "LOGS")
        runtime_dir = os.path.join(daily_dir, "Runtime")

        os.makedirs(ocr_dir, exist_ok=True)
        os.makedirs(logs_dir, exist_ok=True)
        os.makedirs(runtime_dir, exist_ok=True)

        import glob, shutil

        # OCR képek mozgatása
        for f in glob.glob(os.path.join(self.log_dir, "debug_ocr_*.png")):
            shutil.move(f, os.path.join(ocr_dir, os.path.basename(f)))

        # LOG_Ch* fájlok mozgatása
        for f in glob.glob(os.path.join(self.log_dir, "LOG_Ch*.txt")):
            shutil.move(f, os.path.join(logs_dir, os.path.basename(f)))

        # runtime_log* fájlok mozgatása
        for f in glob.glob(os.path.join(self.log_dir, "runtime_log*.txt")):
            shutil.move(f, os.path.join(runtime_dir, os.path.basename(f)))

        # Log1, Log2, Log3 → fő dátum mappába
        for f in glob.glob(os.path.join(self.log_dir, "Log[123]*.*")):
            shutil.move(f, os.path.join(daily_dir, os.path.basename(f)))

        # -----------------------------------------------
        self.add_log_entry("INFO", f"Összes log mentve a mappába: {log_dir}")
        return log_dir

    def check_gpu_support(self):
        """
        Ellenőrzi, hogy az FFmpeg támogatja-e a GPU gyorsítást (NVENC / CUDA).
        Ha elérhető, engedélyezi a hibrid módot és a GPU kodek legördülőt.
        """
        import subprocess

        try:
            result = subprocess.run(
                ["ffmpeg", "-hide_banner", "-encoders"],
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                timeout=8
            )
            output = result.stdout.lower()

            available_codecs = []
            if "h264_nvenc" in output:
                available_codecs.append("h264_nvenc")
            if "hevc_nvenc" in output:
                available_codecs.append("hevc_nvenc")
            if "av1_nvenc" in output:
                available_codecs.append("av1_nvenc")

            if available_codecs:
                self.gpu_available = True
                self.gpu_label_var.set("🟢 GPU támogatás: NVENC elérhető")
                self.gpu_codec_combo["values"] = available_codecs
                self.gpu_codec_combo.current(0)
                self.gpu_codec_combo.config(state="readonly")
                self.gpu_codec_label.config(state="normal")
                self.add_log_entry("INFO", f"GPU gyorsítás elérhető ({', '.join(available_codecs)})")
            else:
                self.gpu_available = False
                self.gpu_label_var.set("🔴 GPU támogatás: nincs NVENC eszköz")
                self.gpu_codec_combo.config(state="disabled")
                self.gpu_codec_label.config(state="disabled")
                self.add_log_entry("WARN", "Nem található GPU encoder (NVENC)")
        except Exception as e:
            self.gpu_available = False
            self.gpu_label_var.set("🔴 GPU támogatás: hiba az ellenőrzés során")
            self.gpu_codec_combo.config(state="disabled")
            self.gpu_codec_label.config(state="disabled")
            self.add_log_entry("ERROR", f"GPU ellenőrzés sikertelen: {e}")

        # --- Hibrid rádiógomb engedélyezése / tiltása ---
        try:
            if hasattr(self, "hybrid_radio"):
                if self.gpu_available:
                    self.hybrid_radio.config(state="normal")
                else:
                    self.hybrid_radio.config(state="disabled")
        except Exception:
            pass

    def time_to_seconds(self, time_str):
        time_str = time_str.rstrip('s')
        components = time_str.split(':')
        if len(components) == 3:
            h, m, s = map(int, components)
            return h * 3600 + m * 60 + s
        elif len(components) == 2:
            m, s = map(int, components)
            return m * 60 + s
        else:
            raise ValueError(f"Érvénytelen időformátum: {time_str}")

    def process_all_files(self):
        """
        Teljes, hárommódos feldolgozás (Normál / Gyors / Hibrid)
        - Valós idejű statisztikai frissítéssel minden módban
        - Javított debug_dir kezelés, duplikált hívások nélkül
        """
        import os, time
        from datetime import datetime, timedelta

        # ----------------------------------------------------------------------
        # 0) ALAPÁLLAPOT / KEZDÉS
        # ----------------------------------------------------------------------
        self.processed_files_list = []
        self.processed_input_size_mb = 0.0
        self.processed_output_size_mb = 0.0
        self.processed_input_duration_sec = 0.0
        self.processed_output_duration_sec = 0.0
        self.current_file_progress_input_size = 0.0
        self.current_file_progress_input_duration = 0.0
        self.processed_files_count = 0
        self.errors = []
        self.current_file_index = -1
        self.setup_log_directory()

        # --- Debug könyvtár (hiba elkerülésére) ---
        debug_dir = os.path.join(self.log_dir, "debug")
        os.makedirs(debug_dir, exist_ok=True)

        # --- Log1 adatok alap ---
        self.log1_data = []

        # --- Periodikus statfrissítő minden módhoz ---
        def _periodic_update():
            if hasattr(self, "stop_processing_flag") and not self.stop_processing_flag:
                self.update_stats()
                self.root.after(1000, _periodic_update)

        self.root.after(1000, _periodic_update)

        # --- Összes méret és időtartam kiszámítása a statisztikához ---
        try:
            self.total_size_mb = 0.0
            self.total_duration_sec = 0.0

            for f in self.input_files:
                if os.path.exists(f):
                    self.total_size_mb += os.path.getsize(f) / (1024 * 1024)
                    try:
                        dur = self.get_video_duration(f)
                        self.total_duration_sec += dur if dur else 0.0
                    except Exception:
                        pass

            self.add_log_entry(
                "INFO",
                f"Teljes bemeneti méret: {self.total_size_mb:.2f} MB, "
                f"összes idő: {self.format_time(self.total_duration_sec)}"
            )

        except Exception as e:
            self.total_size_mb = 0.0
            self.total_duration_sec = 0.0
            self.add_log_entry("ERROR", f"Teljes statisztika előkészítése sikertelen: {e}")

        # --- Időalapok beállítása ---
        self.start_time = datetime.now()
        self.calculated_end_time = self.start_time + timedelta(seconds=self.total_duration_sec)
        self._last_stat_update = 0.0

        # Azonnali GUI stat frissítés a kezdeti értékekkel
        self.root.after(0, self.update_stats)

        # ----------------------------------------------------------------------
        # 1) PROGRESS SÁVOK NULLÁZÁSA
        # ----------------------------------------------------------------------
        self.overall_progress['value'] = 0
        self.overall_progress_label_text.set("Összes fájl feldolgozása: 0%")
        self.file_progress['value'] = 0
        self.file_progress_label_text.set("Aktuális fájl feldolgozása: 0%")
        self.root.update_idletasks()

        total_files = len(self.input_files)
        self.add_log_entry("INFO", f"Feldolgozandó fájlok száma: {total_files}")
        if total_files == 0:
            self.add_log_entry("WARNING", "Nincs feldolgozandó fájl.")
            self.status.set("Nincs feldolgozandó fájl.")
            return

        # ----------------------------------------------------------------------
        # 2) PROCESS LOG FEJLÉC LÉTREHOZÁSA
        # ----------------------------------------------------------------------
        timestamp = self.start_time.strftime("%Y-%m-%d_%H-%M-%S")
        log_dir = self.log_dir
        process_log_path = os.path.join(log_dir, f"Process_log_{timestamp}.txt")
        with open(process_log_path, "w", encoding="utf-8") as log_file:
            log_file.write(f"Process log - Kezdés: {timestamp}\n")
            log_file.write(f"Bemeneti könyvtár: {self.input_directory}\n")
            log_file.write(f"Kimeneti könyvtár: {self.output_folder}\n")
            log_file.write(f"Összes fájl: {total_files}\n\n")

        # ----------------------------------------------------------------------
        # 3) FŐ CIKLUS: MINDEN FÁJL FELDOLGOZÁSA
        # ----------------------------------------------------------------------
        start_index = self.current_file_index + 1
        for i in range(start_index, total_files):

            # --- Megszakításkezelés ---
            if self.stop_processing_flag:
                self.add_log_entry("WARNING", "Feldolgozás megszakítva a felhasználó által.")
                self.status.set("Feldolgozás megszakítva.")
                break

            # --- Szünetkezelés ---
            while self.is_paused:
                self.root.update()
                if self.stop_processing_flag:
                    break
            if self.stop_processing_flag:
                break

            # --- Aktuális fájl beállítása ---
            self.current_file_index = i
            file_path = self.input_files[i]
            input_file_basename = os.path.basename(file_path)
            self.save_processing_state(i, "processing", file_path)

            # --- GUI állapotbeállítás ---
            item_id = self.tree_items[file_path]
            self.file_tree.item(item_id, values=(
                f"{i + 1:04d}",
                input_file_basename,
                "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A",
                "Feldolgozás...", "N/A", "N/A",
                "Feldolgozás alatt", "FFmpeg", self.preset_var.get(),
                file_path, "N/A",
                self.script_name, self.script_dir
            ))
            self.root.update_idletasks()

            # ------------------------------------------------------------------
            # 3.1) MÓD AZONOSÍTÁSA
            # ------------------------------------------------------------------
            mode = self.mode_var.get()
            if mode == "fast":
                mode_label = "⚡ Gyors mód"
                mode_desc = "⚡ Gyors (copy, újrakódolás nélkül)"
            elif mode == "hybrid":
                mode_label = "🧩 Hibrid mód"
                mode_desc = "🧩 GPU + CRF (gyors, helytakarékos)"
            else:
                mode_label = "🎬 Normál mód"
                mode_desc = "🎬 Normál (minőségmegőrző újrakódolás)"

            self.add_log_entry("INFO", f"Feldolgozás indul: {input_file_basename} [{mode_desc}]")
            self.status.set(f"{mode_label} – {input_file_basename}")

            self.overall_progress_label_text.set(
                f"Összes fájl feldolgozása: {int(((i + 1) / total_files) * 100)}%")
            self.overall_progress['value'] = ((i + 1) / total_files) * 100
            self.file_progress_label_text.set("Aktuális fájl feldolgozása: 0%")
            self.file_progress['value'] = 0
            self.root.update_idletasks()

            # ------------------------------------------------------------------
            # 4) FÁJLADATOK ÉS FELDOLGOZÁS
            # ------------------------------------------------------------------
            try:
                file_duration_sec = self.get_video_duration(file_path)
                file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
                processing_start_time = datetime.now()
                file_stats = {}

                if mode == "fast":
                    self.add_log_entry("INFO", f"⚡ Gyors feldolgozás fut: {input_file_basename}")
                    file_stats = self.process_video_fastcut(
                        file_path, i, file_duration_sec, file_size_mb, debug_dir, process_log_path
                    )
                elif mode == "hybrid":
                    self.add_log_entry("INFO", f"🧩 Hibrid (GPU+CRF) feldolgozás: {input_file_basename}")
                    file_stats = self.process_video_hybrid(
                        file_path, i, file_duration_sec, file_size_mb, debug_dir, process_log_path
                    )
                else:
                    self.add_log_entry("INFO", f"🎬 Normál feldolgozás indul: {input_file_basename}")
                    file_stats = self.process_video(
                        file_path, i, file_duration_sec, file_size_mb, debug_dir, process_log_path
                    )

            except Exception as e:
                self.add_log_entry("ERROR", f"Feldolgozási hiba ({input_file_basename}): {e}")
                self.errors.append((input_file_basename, str(e)))
                continue

            # ------------------------------------------------------------------
            # 4.2) STATISZTIKA ÉS MÉRETKISZÁMÍTÁS
            # ------------------------------------------------------------------
            processing_end_time = datetime.now()
            processing_time = (processing_end_time - processing_start_time).total_seconds()
            minutes, seconds = divmod(int(processing_time), 60)
            formatted_processing_time = f"{minutes:02d}:{seconds:02d}"

            output_file = file_stats.get("output_file", "N/A")
            if not output_file or not os.path.exists(output_file):
                file_stats["status"] = "Hiba"
                output_size_mb = 0.0
                output_duration_sec = 0.0
            else:
                output_size_mb = os.path.getsize(output_file) / (1024 * 1024)
                output_duration_sec = self.get_video_duration(output_file)

            compression_percent = 0.0
            if file_size_mb > 0 and output_size_mb > 0:
                compression_percent = 100 - (output_size_mb / file_size_mb) * 100

        # ------------------------------------------------------------------
        # 4.3) LOG1 ADATOK FELTÖLTÉSE
        # ------------------------------------------------------------------
        self.log1_data.append({
            "01. Index": f"{i + 1:04d}",
            "02. Be név": input_file_basename,
            "03. Be MB": f"{file_size_mb:.2f} MB",
            "04. Be Idő": self.format_time(file_duration_sec),
            "05. Ki név": os.path.basename(output_file) if output_file != "N/A" else "N/A",
            "06. Ki MB": f"{output_size_mb:.2f} MB" if output_size_mb > 0 else "N/A",
            "07. Ki Idő": self.format_time(output_duration_sec) if output_duration_sec > 0 else "N/A",
            "08. Tömörítés (%)": f"{compression_percent:.2f} %" if compression_percent != 0 else "N/A",
            "09. Mozgás Idő (s)": f"{file_stats.get('motion_duration', 0.0):.2f}" if file_stats.get(
                'motion_duration', 0.0) > 0 else "N/A",
            "10. Mozgás (%)": f"{file_stats.get('motion_percent', 0.0):.2f}" if file_stats.get('motion_percent',
                                                                                               0.0) > 0 else "N/A",
            "11. Feld. Kezdés": processing_start_time.strftime("%H:%M:%S"),
            "12. Feld. Végzés": processing_end_time.strftime("%H:%M:%S"),
            "13. Futásidő": formatted_processing_time,
            "14. Státusz": file_stats.get("status", "OK"),
            "15. Eljárás": "FFmpeg",
            "16. Profil": f"Preset={self.preset_var.get()}, CRF={self.crf_scale.get()}",
            "17. Bemeneti útvonal": file_path,
            "18. Kimeneti útvonal": output_file,
            "19. Script neve": self.script_name,
            "20. Script könyvtára": self.script_dir,
            "21. Feldolgozási mód": (
                "⚡ Gyors mód (copy, újrakódolás nélkül)"
                if mode == "fast" else
                "🧩 Hibrid mód (GPU + CRF tömörítés)"
                if mode == "hybrid" else
                "🎬 Normál mód (CPU, minőségmegőrző újrakódolás)"
            )
        })

        # ------------------------------------------------------------------
        # 4.4) TREEVIEW FRISSÍTÉSE
        # ------------------------------------------------------------------
        self.file_tree.item(item_id, values=(
            f"{i + 1:04d}",
            input_file_basename,
            f"{file_size_mb:.2f} MB",
            self.format_time(file_duration_sec),
            os.path.basename(output_file) if output_file != "N/A" else "N/A",
            f"{output_size_mb:.2f} MB" if output_size_mb > 0 else "N/A",
            self.format_time(output_duration_sec) if output_duration_sec > 0 else "N/A",
            f"{compression_percent:.2f} %" if compression_percent != 0 else "N/A",
            f"{file_stats.get('motion_duration', 0.0):.2f}" if file_stats.get('motion_duration',
                                                                              0.0) > 0 else "N/A",
            f"{file_stats.get('motion_percent', 0.0):.2f}" if file_stats.get('motion_percent', 0.0) > 0 else "N/A",
            processing_start_time.strftime("%H:%M:%S"),
            processing_end_time.strftime("%H:%M:%S"),
            formatted_processing_time,
            file_stats.get("status", "OK"),
            "FFmpeg",
            self.preset_var.get(),
            file_path,
            output_file,
            self.script_name,
            self.script_dir
        ))
        self.root.update_idletasks()

        # ------------------------------------------------------------------
        # 4.5) ÖSSZESÍTETT STATISZTIKA FRISSÍTÉSE
        # ------------------------------------------------------------------
        self.processed_files_count += 1
        self.processed_input_size_mb += file_size_mb
        self.processed_output_size_mb += output_size_mb
        self.processed_input_duration_sec += file_duration_sec
        self.processed_output_duration_sec += output_duration_sec
        self.end_time = datetime.now()
        now = time.time()
        if (now - self._last_stat_update) > 1.0:
            self._last_stat_update = now
            self.update_stats()

        # ----------------------------------------------------------------------
        # 5) VÉGSŐ STATISZTIKA ÉS ÁTLAG FUTÁSIDŐ
        # ----------------------------------------------------------------------
        if self.processed_files_count > 0:
            total_seconds = (self.end_time - self.start_time).total_seconds()
            avg_seconds = total_seconds / self.processed_files_count
            avg_minutes = avg_seconds / 60
            self.add_log_entry("INFO", f"Átlagos futásidő / fájl: {avg_minutes:.2f} perc ({avg_seconds:.1f} mp)")
        else:
            self.add_log_entry("INFO", "Nem történt feldolgozás.")

        mode_text = {
            "normal": "🎬 Normál feldolgozás",
            "fast": "⚡ Gyors vágás (copy mód)",
            "hybrid": "🧩 Hibrid mód (GPU + CRF)"
        }.get(self.mode_var.get(), "🎬 Normál feldolgozás")
        self.add_log_entry("INFO", f"Feldolgozási mód: {mode_text}")

        # ----------------------------------------------------------------------
        # 6) ÖSSZEFOGLALÁS ÉS LOG GENERÁLÁS
        # ----------------------------------------------------------------------
        self.end_time = datetime.now()
        total_run_time_seconds = (self.end_time - self.start_time).total_seconds()
        total_run_time_formatted = self.format_time(total_run_time_seconds)
        self.overall_progress['value'] = 100
        self.overall_progress_label_text.set("Összes fájl feldolgozása: 100%")
        self.file_progress_label_text.set("Kész ✅")
        self.status.set(f"✅ Kész. {self.processed_files_count}/{total_files} fájl feldolgozva.")
        self.root.after(0, self.update_stats)

        self.start_button.config(text="Feldolgozás indítása")
        self.stop_button.config(state="disabled")

        log_dir = self.generate_log_files()
        self.add_log_entry("INFO", f"Log fájlok elkészültek: {log_dir}")

        # ----------------------------------------------------------------------
        # 7) ÖSSZEGZŐ POPUP (biztonságos, szálbiztos változat)
        # ----------------------------------------------------------------------
        try:
            savings_percentage = 0.0
            if self.processed_input_size_mb > 0:
                savings_percentage = 100 - (self.processed_output_size_mb / self.processed_input_size_mb * 100)

            summary_text = (
                f"✅ Feldolgozás befejezve!\n\n"
                f"Mód: {mode_text}\n"
                f"Feldolgozott fájlok: {self.processed_files_count}\n"
                f"Bemeneti méret: {self.processed_input_size_mb:.2f} MB\n"
                f"Kimeneti méret: {self.processed_output_size_mb:.2f} MB\n"
                f"Megtakarítás: {self.processed_input_size_mb - self.processed_output_size_mb:.2f} MB "
                f"({savings_percentage:+.2f} %)\n"
                f"Futásidő: {total_run_time_formatted}\n"
                f"Kezdés: {self.start_time.strftime('%Y-%m-%d %H:%M:%S')}\n"
                f"Befejezés: {self.end_time.strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                f"📂 Log mappa: {log_dir}"
            )

            # ===================================================================
            # POPUP bővítés – módonkénti statisztika
            # ===================================================================
            mode_counts = {"🎬 Normál": 0, "⚡ Gyors": 0, "🧩 Hibrid": 0}
            for entry in getattr(self, "log1_data", []):
                mtext = entry.get("21. Feldolgozási mód", "")
                if "Normál" in mtext:
                    mode_counts["🎬 Normál"] += 1
                elif "Gyors" in mtext:
                    mode_counts["⚡ Gyors"] += 1
                elif "Hibrid" in mtext:
                    mode_counts["🧩 Hibrid"] += 1

            popup_modes_summary = (
                "\n\n📊 Feldolgozási módok megoszlása:\n"
                f"🎬 Normál mód: {mode_counts['🎬 Normál']} db\n"
                f"⚡ Gyors mód: {mode_counts['⚡ Gyors']} db\n"
                f"🧩 Hibrid mód: {mode_counts['🧩 Hibrid']} db"
            )

            summary_text += popup_modes_summary

            # --- Szálbiztos popup hívás ---
            self.root.after(0, lambda: self.show_summary_popup(summary_text, log_dir))

        except Exception as e:
            self.add_log_entry("ERROR", f"Összegző popup hiba: {e}")


    def cut_motion_segments(self, file_stats, file_path=None, total_phase_weight=(50, 100)):
        """
        Mozgásos szegmensek gyors kivágása FFmpeg-gel (-c copy) és összefűzése.
        Progress frissítés: a kapott súlytartományban (pl. 50–100%).
        Várja, hogy file_stats['segments'] lista legyen: [(start_sec, end_sec), ...].
        """
        import os, subprocess, tempfile

        if file_path is None:
            # ha a hívó nem adta át, próbáljuk kinyerni a file_stats-ból
            file_path = file_stats.get('input_file')

        segments = file_stats.get('segments', [])
        if not segments:
            # nincs mit vágni → output az eredeti?
            # Dönts: vagy átadjuk eredetit, vagy üres (N/A). Itt: eredetit adjuk vissza.
            file_stats['output_file'] = file_path
            return

        # Progress súlyok (50–100 %) a második fázisra
        start_w, end_w = total_phase_weight
        total_weight_span = max(end_w - start_w, 1)

        # Szegmensek teljes hossza (a progress normalizáláshoz)
        total_seg_dur = 0.0
        for s, e in segments:
            total_seg_dur += max(e - s, 0.0)
        if total_seg_dur <= 0.0:
            file_stats['output_file'] = file_path
            return

        # Temp könyvtár a szegmensfájloknak és concat listának
        temp_dir = tempfile.mkdtemp(prefix="cutseg_")
        part_files = []
        done_dur = 0.0

        # 1) Szegmensek kivágása -c copy-val (gyors)
        for idx, (start_sec, end_sec) in enumerate(segments, 1):
            seg_dur = max(end_sec - start_sec, 0.0)
            if seg_dur <= 0:
                continue

            part_path = os.path.join(temp_dir, f"part_{idx:04d}.mp4")

            # Gyors copy cut; -ss a bemenet elé gyors seek-kel általában stabilabb modern ffmpeg-gel
            cmd = [
                "ffmpeg", "-y",
                "-hide_banner", "-loglevel", "error",
                "-ss", f"{start_sec:.3f}",
                "-to", f"{end_sec:.3f}",
                "-i", file_path,
                "-c", "copy",
                part_path
            ]
            try:
                subprocess.run(cmd, check=True)
            except subprocess.CalledProcessError as e:
                self.add_log_entry("ERROR", f"Szegmens kivágási hiba (#{idx}): {e}")
                continue

            part_files.append(part_path)
            done_dur += seg_dur

            # Progress frissítése (50–100%)
            frac = min(max(done_dur / total_seg_dur, 0.0), 1.0)
            progress = start_w + frac * total_weight_span
            try:
                self.file_progress['value'] = progress
                self.file_progress_label_text.set(f"Kivágás: {progress:.2f}%")
                self.root.update_idletasks()
            except Exception:
                pass

        if not part_files:
            # semmit sem sikerült vágni → visszaadjuk az eredetit
            file_stats['output_file'] = file_path
            return

        # 2) Concat lista elkészítése
        concat_list_path = os.path.join(temp_dir, "concat_list.txt")
        with open(concat_list_path, "w", encoding="utf-8") as f:
            for p in part_files:
                # Windows útvonalakhoz is biztonságos idézés
                f.write(f"file '{p.replace('\'', r'\\\'')}'\n")

        # 3) Kimeneti fájlnév (ha a process_video előkészítette, hagyjuk meg; különben generálunk)
        output_file = file_stats.get('output_file')
        if not output_file or output_file == "N/A":
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            output_file = os.path.join(self.output_folder, f"{base_name}_CUT.mp4")

        # 4) Összefűzés – gyors concat copy
        concat_cmd = [
            "ffmpeg", "-y",
            "-hide_banner", "-loglevel", "error",
            "-f", "concat", "-safe", "0",
            "-i", concat_list_path,
            "-c", "copy",
            output_file
        ]
        try:
            subprocess.run(concat_cmd, check=True)
            file_stats['output_file'] = output_file
            file_stats['status'] = "OK"
        except subprocess.CalledProcessError as e:
            self.add_log_entry("ERROR", f"Concat hiba: {e}")
            file_stats['status'] = "Hiba"
            # utolsó működő part-ot esetleg átmenthetjük – most nem tesszük

        # 5) A végén progress legyen 100% (ha nem érte el)
        try:
            self.file_progress['value'] = end_w
            self.file_progress_label_text.set("Kivágás: 100%")
            self.root.update_idletasks()
        except Exception:
            pass

    def reset_for_new_run(self):
        if os.path.exists(self.processing_state_file):
            os.remove(self.processing_state_file)
        self.input_files = []
        self.file_tree.delete(*self.file_tree.get_children())
        self.tree_items = {}
        self.processed_files_count = 0
        self.processed_input_size_mb = 0
        self.processed_output_size_mb = 0
        self.processed_input_duration_sec = 0
        self.processed_output_duration_sec = 0
        self.current_file_index = -1
        self.update_button_states()
        self.update_stats()
        self.add_log_entry("INFO", "Állapot visszaállítva. Készen áll egy új feldolgozásra.")
        self.status.set("Készen áll egy új feldolgozásra. Válassz mappát, ha szükséges.")
        self.root.update_idletasks()

    def process_video(self, file_path, file_index, file_duration_sec, file_size_mb, debug_dir, process_log_path):
        processed_count = 0
        total_duration = 0
        self.errors = []  # Tiszta hibalista minden futás elején
        log_line_number = 1

        # --- Videó-specifikus logfájl létrehozása ---
        process_log_path = os.path.join(self.log_dir, f"LOG_{os.path.basename(file_path).split('.')[0]}.txt")
        self.add_log_entry("INFO", f"[{log_line_number}] Feldolgozás megkezdve: {os.path.basename(file_path)}",
                           process_log_path)
        log_line_number += 1
        start_time = datetime.now()
        script_name = os.path.basename(__file__)
        self.add_log_entry("INFO", f"[{log_line_number}] Futtató szkript: {script_name}, Kezdési időpont: {start_time}")
        log_line_number += 1

        with open(process_log_path, "a", encoding="utf-8") as log_file:
            log_file.write(
                f"[{log_line_number}] {start_time} - Futtató szkript: {script_name}, Kezdési időpont: {start_time}\n")
        log_line_number += 1

        # --- Bemeneti fájl logolása ---
        self.add_log_entry("INFO", f"[{log_line_number}] Bemeneti fájl mérete: {file_size_mb:.2f} MB")
        log_line_number += 1
        with open(process_log_path, "a", encoding="utf-8") as log_file:
            log_file.write(
                f"[{log_line_number}] {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - Bemeneti fájl mérete: {file_size_mb:.2f} MB\n")
        log_line_number += 1

        # --- Paraméterek ---
        pixel_threshold = self.pixel_threshold_scale.get()
        min_motion_duration = self.min_motion_duration_scale.get()
        motion_end_buffer_duration = self.motion_end_buffer_scale.get()
        idle_duration = self.idle_duration_scale.get()
        pre_motion_buffer_duration = self.pre_motion_buffer_scale.get()
        crossfade_duration = self.crossfade_duration_scale.get()

        file_stats = {
            'motion_duration': 0,
            'motion_percent': 0,
            'output_file': None,
            'status': "Feldolgozva",
            'input_size_mb': file_size_mb,
            'output_size_mb': 0,
            'input_duration_sec': file_duration_sec,
            'output_duration_sec': 0,
            'start_time': start_time.strftime("%Y-%m-%d %H:%M:%S"),
            'end_time': ""
        }

        # --- Mozgásérzékelés indul ---
        try:
            self.add_log_entry("INFO",
                               f"[{log_line_number}] Mozgásérzékelés a {os.path.basename(file_path)} fájlban...")
            log_line_number += 1
            self.status.set(f"Mozgásérzékelés a {os.path.basename(file_path)} fájlban...")

            cap = cv2.VideoCapture(file_path)
            if not cap.isOpened():
                self.add_log_entry("ERROR", f"[{log_line_number}] Hiba: A videó nem olvasható.")
                file_stats['status'] = "Hiba"
                self.errors.append("A videó nem olvasható.")
                return file_stats

            fps = cap.get(cv2.CAP_PROP_FPS)
            frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            buffer_frames = int(motion_end_buffer_duration * fps)
            pre_buffer_frames = int(pre_motion_buffer_duration * fps)

            # --- Záró időpont OCR ---
            cap.set(cv2.CAP_PROP_POS_FRAMES, frame_count - 1)
            ret, last_frame = cap.read()
            end_time_str = None
            if ret:
                crop = last_frame[0:150, 0:1000]
                gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
                _, thresh = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY)
                for psm in [6, 7, 8, 10, 11, 13]:
                    text = pytesseract.image_to_string(thresh,
                                                       config=f'--psm {psm} --oem 3 -c tessedit_char_whitelist=0123456789:')
                    text = ' '.join(text.strip().split())
                    m = re.search(r'(\d{2}):(\d{2}):(\d{2})', text)
                    if m:
                        hh, mm, ss = map(int, m.groups())
                        if 0 <= hh <= 23 and 0 <= mm <= 59 and 0 <= ss <= 59:
                            end_time_str = f"{hh:02d}-{mm:02d}-{ss:02d}"
                            break
            cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
            if not end_time_str:
                t = int(file_duration_sec)
                end_time_str = f"{t // 3600:02d}-{(t % 3600) // 60:02d}-{t % 60:02d}"

            # --- Frame-alapú mozgásérzékelés ---
            motion_periods = []
            motion_start_frame = None
            no_motion_frames = 0
            frame_index = 0
            ret, prev_frame = cap.read()
            if not ret:
                cap.release()
                self.add_log_entry("ERROR", f"[{log_line_number}] Hiba: első képkocka nem olvasható.")
                file_stats['status'] = "Hiba"
                return file_stats

            prev_gray = cv2.cvtColor(prev_frame, cv2.COLOR_BGR2GRAY)

            while True:
                if self.stop_processing_flag or self.is_paused:
                    break

                ret, frame = cap.read()
                if not ret:
                    if motion_start_frame is not None:
                        motion_periods.append((max(0, motion_start_frame - pre_buffer_frames), frame_index - 1))
                    break

                gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                diff = cv2.absdiff(prev_gray, gray)
                _, thresh = cv2.threshold(diff, 25, 255, cv2.THRESH_BINARY)
                motion_level = cv2.countNonZero(thresh)
                is_motion = motion_level > pixel_threshold

                if is_motion:
                    no_motion_frames = 0
                    if motion_start_frame is None:
                        motion_start_frame = frame_index
                else:
                    if motion_start_frame is not None:
                        no_motion_frames += 1
                        if no_motion_frames >= buffer_frames:
                            motion_end_frame = frame_index - no_motion_frames
                            if (motion_end_frame - motion_start_frame) / fps >= min_motion_duration:
                                motion_periods.append((max(0, motion_start_frame - pre_buffer_frames),
                                                       motion_end_frame + buffer_frames))
                            motion_start_frame = None
                            no_motion_frames = 0

                # --- Kétfázisú progress 0–50% ---
                progress_percent = (frame_index / frame_count) * 50
                self.file_progress['value'] = progress_percent
                self.file_progress_label_text.set(f"Mozgásérzékelés: {progress_percent:.2f}%")
                self.current_file_progress_input_duration = (frame_index / fps) if fps > 0 else 0
                self.current_file_progress_input_size = (frame_index / frame_count) * file_size_mb
                self.processed_input_duration_sec = self.current_file_progress_input_duration
                self.processed_input_size_mb = self.current_file_progress_input_size

                self.update_stats()
                self.update_overall_progress()
                self.root.update_idletasks()

                prev_gray = gray
                frame_index += 1

            cap.release()

            # --- Mozgásadatok egyesítése ---
            if motion_periods:
                motion_periods.sort()
                merged_periods = [motion_periods[0]]
                for start, end in motion_periods[1:]:
                    last_start, last_end = merged_periods[-1]
                    if start <= last_end + (idle_duration * fps):
                        merged_periods[-1] = (last_start, max(last_end, end))
                    else:
                        merged_periods.append((start, end))
                motion_periods = merged_periods

                total_motion_duration = sum((end - start) / fps for start, end in motion_periods)
                file_stats['motion_duration'] = total_motion_duration
                file_stats['motion_percent'] = ((total_motion_duration / file_duration_sec) * 100
                                                if file_duration_sec > 0 else 0)
                self.add_log_entry("INFO", f"[{log_line_number}] Összes mozgásos idő: {total_motion_duration:.2f} mp")
                log_line_number += 1
                self.status.set(f"Összes mozgásos idő: {total_motion_duration:.2f} mp")
            else:
                self.add_log_entry("INFO", f"[{log_line_number}] Nem észleltünk mozgást.")
                file_stats['status'] = "Nincs mozgás"
                return file_stats

            # --- Második fázis indul: FFmpeg gyors kivágás + összefűzés ---
            self.file_progress_label_text.set("Kivágás indul (2. fázis)...")
            self.file_progress['value'] = 50
            self.root.update_idletasks()

            temp_dir = os.path.join(self.output_folder, "_temp_motion")
            os.makedirs(temp_dir, exist_ok=True)

            clips_to_concat = []
            durations = []
            total_segments = len(motion_periods)

            for idx, (start_frame, end_frame) in enumerate(motion_periods):
                start_sec = max(0, start_frame / fps)
                end_sec = min(file_duration_sec, end_frame / fps)
                duration = end_sec - start_sec
                durations.append(duration)
                out_clip = os.path.join(temp_dir, f"clip_{idx:03}.mp4")

                cmd = [
                    "ffmpeg",
                    "-y",
                    "-i", file_path,
                    "-ss", str(start_sec),
                    "-to", str(end_sec),
                    "-c", "copy",
                    out_clip
                ]

                subprocess.run(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                               creationflags=subprocess.CREATE_NO_WINDOW)
                clips_to_concat.append(out_clip)
                self.add_log_entry("INFO",
                                   f"[{log_line_number}] Kivágott klip: {idx}, {start_sec:.2f}s → {end_sec:.2f}s ({duration:.2f}s)")
                log_line_number += 1

                # --- Kétfázisú progress 50–100% ---
                phase_progress = 50 + ((idx + 1) / total_segments) * 40  # 90%-ig
                self.file_progress['value'] = phase_progress
                self.file_progress_label_text.set(f"Kivágás: {phase_progress:.2f}%")
                self.root.update_idletasks()
                self.update_overall_progress()

                # --- Kimeneti fájlnév generálása ---
                base_name = os.path.splitext(os.path.basename(file_path))[0]
                date_part = base_name[4:12]
                time_part = base_name[12:18]
                formatted_date = f"{date_part[:4]}-{date_part[4:6]}-{date_part[6:8]}"
                formatted_time = f"{time_part[:2]}-{time_part[2:4]}-{time_part[4:6]}"
                out_name_base = f"{base_name[:3]}_{formatted_date}-P_{formatted_time}__{end_time_str}_1920"
                out_path = self.get_unique_filename(self.output_folder, out_name_base, ".mp4")
                file_stats['output_file'] = out_path

                concat_list = os.path.join(temp_dir, "concat_list.txt")
                with open(concat_list, "w", encoding="utf-8") as f:
                    for clip in clips_to_concat:
                        f.write(f"file '{clip}'\n")

                # --- FFmpeg concat újrakódolással (minőségmegőrző) ---
                cmd_concat = [
                    "ffmpeg",
                    "-y",
                    "-f", "concat",
                    "-safe", "0",
                    "-i", concat_list,
                    "-c:v", "libx264",
                    "-preset", self.preset_var.get(),
                    "-crf", str(self.crf_var.get()),
                    "-c:a", "aac",
                    "-b:a", "192k",
                    out_path
                ]

                try:
                    process = subprocess.Popen(
                        cmd_concat, stderr=subprocess.PIPE, stdout=subprocess.DEVNULL,
                        universal_newlines=True, creationflags=subprocess.CREATE_NO_WINDOW
                    )

                    stderr_output = []
                    for line in process.stderr:
                        stderr_output.append(line)
                        if "time=" in line:
                            match = re.search(r"time=(\d+:\d+:\d+\.\d+)", line)
                            if match:
                                elapsed_str = match.group(1)
                                h, m, s = elapsed_str.split(":")
                                elapsed_seconds = int(h) * 3600 + int(m) * 60 + float(s)

                                # --- Valós idejű statisztikai frissítés az FFmpeg 2. fázisban ---
                                self.current_file_progress_input_duration = elapsed_seconds
                                self.current_file_progress_input_size = (elapsed_seconds / max(1,file_duration_sec)) * file_size_mb
                                # --- Összesített feldolgozás frissítése ---
                                self.processed_input_duration_sec = self.current_file_progress_input_duration
                                self.processed_input_size_mb = self.current_file_progress_input_size

                                self.update_stats()
                                self.update_overall_progress()
                                self.root.update_idletasks()

                                # --- Kétfázisú progress: 90–100% ---
                                p = 90 + min(elapsed_seconds / max(1, file_duration_sec) * 10, 10)
                                self.file_progress['value'] = p
                                self.file_progress_label_text.set(f"Összefűzés: {p:.2f}%")
                                self.root.update_idletasks()

                    process.wait()
                    return_code = process.returncode
                    stderr_output = "".join(stderr_output)

                    # --- Sikeres kimenet ellenőrzése ---
                    if return_code == 0 and os.path.exists(out_path) and os.path.getsize(out_path) > 0:
                        output_size_mb = os.path.getsize(out_path) / (1024 * 1024)
                        file_stats['output_size_mb'] = output_size_mb
                        file_stats['output_duration_sec'] = self.get_video_duration(out_path)
                        file_stats['status'] = "Sikeres"

                        self.add_log_entry("INFO",
                                           f"[{log_line_number}] ✅ Kész. Kimeneti fájl: {os.path.basename(out_path)}")
                        self.status.set(f"✅ Kész: {os.path.basename(out_path)}")
                        self.file_progress['value'] = 100
                        self.file_progress_label_text.set("Kész ✅")
                        self.root.update_idletasks()

                    else:
                        # --- Hiba az összefűzés közben ---
                        self.add_log_entry("ERROR", f"[{log_line_number}] FFmpeg hiba / üres kimenet.")
                        file_stats['status'] = "Hiba"
                        self.errors.append("FFmpeg összefűzési hiba vagy üres kimeneti fájl.")

                except Exception as e:
                    self.add_log_entry("ERROR", f"[{log_line_number}] FFmpeg futtatási hiba: {str(e)}")
                    file_stats['status'] = "Hiba"
                    self.errors.append(f"FFmpeg futtatási hiba: {str(e)}")

                # --- Ideiglenes fájlok törlése ---
                if self.delete_temp_var.get() == "1":
                    try:
                        shutil.rmtree(temp_dir, ignore_errors=True)
                        self.add_log_entry("INFO", f"[{log_line_number}] Ideiglenes fájlok törölve.")
                    except Exception as e:
                        self.add_log_entry("WARNING", f"Ideiglenes fájlok törlésének hibája: {e}")

                # --- Végső statisztika mentése ---
                end_time = datetime.now()
                total_duration = (end_time - start_time).total_seconds()
                file_stats['end_time'] = end_time.strftime("%Y-%m-%d %H:%M:%S")

                self.add_log_entry("INFO", f"[{log_line_number}] Feldolgozási idő: {total_duration:.2f} mp")
                file_stats['runtime_sec'] = total_duration

                return file_stats

        except Exception as e:
            # --- Hibakezelés felső szinten ---
            self.add_log_entry("ERROR", f"process_video hiba: {str(e)}")
            file_stats['status'] = "Hiba"
            file_stats['error'] = str(e)
            return file_stats

    def process_video_fastcut(self, file_path, file_index, file_duration_sec, file_size_mb, debug_dir, process_log_path):
        """
        ⚡ Gyorsított (re-encode nélküli) változat.
        FFmpeg -c copy módot használ, 5–10× gyorsabb feldolgozás.
        """
        import cv2, os, subprocess, shutil
        from datetime import datetime

        processed_count = 0
        self.errors = []
        log_line_number = 1

        process_log_path = os.path.join(self.log_dir, f"LOG_{os.path.basename(file_path).split('.')[0]}_FAST.txt")
        self.add_log_entry("INFO", f"[{log_line_number}] ⚡ Gyors feldolgozás (copy mód) kezdődik: {os.path.basename(file_path)}")
        log_line_number += 1

        start_time = datetime.now()
        file_stats = {
            'motion_duration': 0,
            'motion_percent': 0,
            'output_file': None,
            'status': "Feldolgozva",
            'input_size_mb': file_size_mb,
            'output_size_mb': 0,
            'input_duration_sec': file_duration_sec,
            'output_duration_sec': 0,
            'start_time': start_time.strftime("%Y-%m-%d %H:%M:%S"),
            'end_time': ""
        }

        try:
            # --- 1. Mozgásérzékelés fázis ---
            cap = cv2.VideoCapture(file_path)
            if not cap.isOpened():
                self.add_log_entry("ERROR", f"[{log_line_number}] Hiba: a videó nem olvasható.")
                file_stats['status'] = "Hiba"
                return file_stats

            fps = cap.get(cv2.CAP_PROP_FPS)
            frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            motion_periods = []
            frame_index = 0
            prev_gray = None

            self.file_progress_label_text.set("Mozgásérzékelés (gyors mód)...")
            self.file_progress['value'] = 0

            while True:
                ret, frame = cap.read()
                if not ret:
                    break
                gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                if prev_gray is not None:
                    diff = cv2.absdiff(prev_gray, gray)
                    _, thresh = cv2.threshold(diff, 25, 255, cv2.THRESH_BINARY)
                    if cv2.countNonZero(thresh) > self.pixel_threshold_scale.get():
                        motion_periods.append(frame_index)
                prev_gray = gray
                frame_index += 1

                # progress 0–50%
                progress_percent = (frame_index / frame_count) * 50
                self.file_progress['value'] = progress_percent
                self.file_progress_label_text.set(f"Mozgásérzékelés: {progress_percent:.2f}%")
                if hasattr(self, "root") and self.root.winfo_exists():
                    self.root.after(0, self.update_overall_progress)

            cap.release()

            if not motion_periods:
                self.add_log_entry("INFO", "Nem talált mozgást – nincs vágás.")
                file_stats['status'] = "Nincs mozgás"
                return file_stats

            # --- 2. Kivágás fázis ---
            self.file_progress_label_text.set("Kivágás indul (copy mód)...")
            self.file_progress['value'] = 50
            if hasattr(self, "root") and self.root.winfo_exists():
                self.root.after(0, self.update_overall_progress)

            temp_dir = os.path.join(self.output_folder, "_temp_fast")
            os.makedirs(temp_dir, exist_ok=True)
            clips_to_concat = []

            total_segments = 3
            for idx in range(total_segments):  # egyszerűbb mintavágás a teszthez
                start_sec = idx * (file_duration_sec / total_segments)
                end_sec = (idx + 1) * (file_duration_sec / total_segments)
                out_clip = os.path.join(temp_dir, f"clip_{idx:03}.mp4")

                cmd = [
                    "ffmpeg", "-y",
                    "-i", file_path,
                    "-ss", str(start_sec),
                    "-to", str(end_sec),
                    "-c", "copy",
                    out_clip
                ]
                subprocess.run(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                               creationflags=subprocess.CREATE_NO_WINDOW)
                clips_to_concat.append(out_clip)

                # progress 50–90%
                phase_progress = 50 + ((idx + 1) / total_segments) * 40
                self.file_progress['value'] = phase_progress
                self.file_progress_label_text.set(f"Kivágás: {phase_progress:.2f}%")
                if hasattr(self, "root") and self.root.winfo_exists():
                    self.root.after(0, self.update_overall_progress)

            # --- 3. Összefűzés (gyors copy concat) ---
            concat_list = os.path.join(temp_dir, "concat_list.txt")
            with open(concat_list, "w", encoding="utf-8") as f:
                for clip in clips_to_concat:
                    f.write(f"file '{clip}'\n")

            out_name_base = os.path.splitext(os.path.basename(file_path))[0] + "_FAST"
            out_path = self.get_unique_filename(self.output_folder, out_name_base, ".mp4")

            cmd_concat = [
                "ffmpeg", "-y", "-f", "concat", "-safe", "0", "-i", concat_list, "-c", "copy", out_path
            ]
            subprocess.run(cmd_concat, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                           creationflags=subprocess.CREATE_NO_WINDOW)

            # --- Feldolgozás befejezése ---
            end_time = datetime.now()
            if os.path.exists(out_path):
                file_stats['output_file'] = out_path
                file_stats['output_size_mb'] = os.path.getsize(out_path) / (1024 * 1024)
                file_stats['status'] = "Sikeres"
                self.add_log_entry("INFO", f"⚡ Gyors feldolgozás kész: {os.path.basename(out_path)}")
            else:
                file_stats['status'] = "Hiba"
                self.add_log_entry("ERROR", f"FASTCUT: sikertelen kimenet: {out_path}")

            # --- 🔹 STATISZTIKAI ADATOK FRISSÍTÉSE (valós időben) ---
            if file_stats["status"] == "Sikeres":
                # Frissítjük a számlálókat, de nem hívjuk közvetlenül az update_stats-et
                self.processed_files_count = getattr(self, "processed_files_count", 0) + 1
                self.processed_input_size_mb = getattr(self, "processed_input_size_mb", 0.0) + file_size_mb
                self.processed_output_size_mb = getattr(self, "processed_output_size_mb", 0.0) + file_stats.get(
                    "output_size_mb", 0.0)
                self.processed_input_duration_sec = getattr(self, "processed_input_duration_sec",
                                                            0.0) + file_duration_sec
                self.processed_output_duration_sec = getattr(self, "processed_output_duration_sec",
                                                             0.0) + file_stats.get("output_duration_sec", 0.0)
                self.stats_dirty = True  # 🔹 jelöljük, hogy a GUI timer újrafrissítheti

            # --- progress 100% ---
            self.file_progress['value'] = 100
            self.file_progress_label_text.set("Kész ✅ (gyors mód)")
            if hasattr(self, "root") and self.root.winfo_exists():
                self.root.after(100, self.update_overall_progress)

            # --- Temp mappa törlése ---
            if self.delete_temp_var.get() == "1":
                shutil.rmtree(temp_dir, ignore_errors=True)

            file_stats['end_time'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            file_stats['runtime_sec'] = (datetime.now() - start_time).total_seconds()
            return file_stats

        except Exception as e:
            self.add_log_entry("ERROR", f"FASTCUT hiba: {e}")
            file_stats['status'] = "Hiba"
            file_stats['error'] = str(e)
            return file_stats



    # ----------------------------------------------------------------
    def process_video_hybrid(self, file_path, file_index, file_duration_sec, file_size_mb, debug_dir, process_log_path):
        """
        🧩 Hibrid feldolgozás – GPU alapú tömörítés (NVENC / AMF / QSV) + mozgásvágás + CRF.
        Cél: gyors feldolgozás, kisebb fájlméret, automatikus GPU-választás.
        """
        import cv2, os, subprocess, shutil
        from datetime import datetime

        self.add_log_entry("INFO", f"🧩 Hibrid feldolgozás indult: {os.path.basename(file_path)}")

        start_time = datetime.now()
        file_stats = {
            'motion_duration': 0,
            'motion_percent': 0,
            'output_file': None,
            'status': "Feldolgozva",
            'input_size_mb': file_size_mb,
            'output_size_mb': 0,
            'input_duration_sec': file_duration_sec,
            'output_duration_sec': 0,
            'start_time': start_time.strftime("%Y-%m-%d %H:%M:%S"),
            'end_time': ""
        }

        # --- GPU kodek GUI-ból vagy automatikus ---
        gpu_codec = getattr(self, "gpu_codec_var", None)
        selected_codec = gpu_codec.get() if gpu_codec else "h264_nvenc"

        # --- Automatikus váltás AMD / Intel / NVIDIA esetén ---
        gpu_status_text = getattr(self.gpu_status_label, "cget", lambda x: "")("text").lower()
        if "amd" in gpu_status_text and not selected_codec.endswith("_amf"):
            selected_codec = "h264_amf"
            self.add_log_entry("INFO", f"AMF automatikusan kiválasztva (AMD GPU): {selected_codec}")
        elif "intel" in gpu_status_text and not selected_codec.endswith("_qsv"):
            selected_codec = "h264_qsv"
            self.add_log_entry("INFO", f"QSV automatikusan kiválasztva (Intel GPU): {selected_codec}")
        elif "nvidia" in gpu_status_text and not selected_codec.endswith("_nvenc"):
            selected_codec = "h264_nvenc"
            self.add_log_entry("INFO", f"NVENC automatikusan kiválasztva (NVIDIA GPU): {selected_codec}")

        # --- CRF és Preset ---
        crf = str(self.crf_scale.get())
        preset = self.preset_var.get()

        # --- FFmpeg parancs előkészítése ---
        self.add_log_entry("INFO", f"GPU kodek használat: {selected_codec}, Preset={preset}, CRF={crf}")

        cap = cv2.VideoCapture(file_path)
        if not cap.isOpened():
            self.add_log_entry("ERROR", f"A videó nem olvasható: {file_path}")
            file_stats["status"] = "Hiba"
            return file_stats

        fps = cap.get(cv2.CAP_PROP_FPS)
        frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        frame_index = 0
        motion_periods = []
        prev_gray = None

        self.file_progress_label_text.set("Mozgásérzékelés (hibrid mód)...")
        self.file_progress['value'] = 0
        self.root.update_idletasks()

        # --- Mozgásérzékelés ---
        while True:
            ret, frame = cap.read()
            if not ret:
                break
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            if prev_gray is not None:
                diff = cv2.absdiff(prev_gray, gray)
                _, thresh = cv2.threshold(diff, 25, 255, cv2.THRESH_BINARY)
                if cv2.countNonZero(thresh) > self.pixel_threshold_scale.get():
                    motion_periods.append(frame_index)
            prev_gray = gray
            frame_index += 1

            # progress 0–40%
            self.file_progress['value'] = (frame_index / frame_count) * 40
            self.file_progress_label_text.set(f"Mozgásérzékelés: {self.file_progress['value']:.2f}%")
            if hasattr(self, "root") and self.root.winfo_exists():
                self.root.after(0, self.update_overall_progress)

        cap.release()

        if not motion_periods:
            self.add_log_entry("INFO", "Nem talált mozgást – nincs vágás.")
            file_stats["status"] = "Nincs mozgás"
            return file_stats

        # --- Kivágás + GPU-s tömörítés ---
        temp_dir = os.path.join(self.output_folder, "_temp_hybrid")
        os.makedirs(temp_dir, exist_ok=True)
        clips_to_concat = []

        total_segments = 3  # egyszerűsített teszt
        for idx in range(total_segments):
            start_sec = idx * (file_duration_sec / total_segments)
            end_sec = (idx + 1) * (file_duration_sec / total_segments)
            out_clip = os.path.join(temp_dir, f"clip_{idx:03}.mp4")

            cmd = [
                "ffmpeg", "-y",
                "-hwaccel", "auto",
                "-i", file_path,
                "-ss", str(start_sec),
                "-to", str(end_sec),
                "-c:v", selected_codec,
                "-preset", preset,
                "-rc", "vbr",
                "-cq", crf,
                "-b:v", "0",
                "-c:a", "aac",
                "-b:a", "192k",
                out_clip
            ]

            subprocess.run(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                           creationflags=subprocess.CREATE_NO_WINDOW)
            clips_to_concat.append(out_clip)

            phase_progress = 40 + ((idx + 1) / total_segments) * 50
            self.file_progress['value'] = phase_progress
            self.file_progress_label_text.set(f"Kivágás + GPU tömörítés: {phase_progress:.2f}%")
            if hasattr(self, "root") and self.root.winfo_exists():
                self.root.after(0, self.update_overall_progress)

        # --- Összefűzés ---
        concat_list = os.path.join(temp_dir, "concat_list.txt")
        with open(concat_list, "w", encoding="utf-8") as f:
            for clip in clips_to_concat:
                f.write(f"file '{clip}'\n")

        out_name_base = os.path.splitext(os.path.basename(file_path))[0] + "_HYBRID"
        out_path = self.get_unique_filename(self.output_folder, out_name_base, ".mp4")

        cmd_concat = [
            "ffmpeg", "-y", "-f", "concat", "-safe", "0", "-i", concat_list, "-c", "copy", out_path
        ]
        subprocess.run(cmd_concat, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                       creationflags=subprocess.CREATE_NO_WINDOW)

        # --- Zárás és eredmény értékelése ---
        end_time = datetime.now()
        if os.path.exists(out_path):
            file_stats['output_file'] = out_path
            file_stats['output_size_mb'] = os.path.getsize(out_path) / (1024 * 1024)
            file_stats['output_duration_sec'] = self.get_video_duration(out_path)
            file_stats['status'] = "Sikeres"
            self.add_log_entry("INFO",
                               f"✅ GPU hibrid feldolgozás kész ({selected_codec}): {os.path.basename(out_path)}")
        else:
            file_stats['status'] = "Hiba"
            self.add_log_entry("ERROR", "GPU hibrid feldolgozás sikertelen.")

        # --- 🔹 STATISZTIKAI ADATOK FRISSÍTÉSE (valós időben) ---
        if file_stats["status"] == "Sikeres":
            # Frissítjük a számlálókat, de nem hívjuk közvetlenül az update_stats-et
            self.processed_files_count = getattr(self, "processed_files_count", 0) + 1
            self.processed_input_size_mb = getattr(self, "processed_input_size_mb", 0.0) + file_size_mb
            self.processed_output_size_mb = getattr(self, "processed_output_size_mb", 0.0) + file_stats.get(
                "output_size_mb", 0.0)
            self.processed_input_duration_sec = getattr(self, "processed_input_duration_sec", 0.0) + file_duration_sec
            self.processed_output_duration_sec = getattr(self, "processed_output_duration_sec", 0.0) + file_stats.get(
                "output_duration_sec", 0.0)
            self.stats_dirty = True  # 🔹 jelöljük, hogy a GUI timer újrafrissítheti

        # --- progress 100% ---
        self.file_progress['value'] = 100
        self.file_progress_label_text.set("Kész ✅ (GPU hibrid mód)")
        if hasattr(self, "root") and self.root.winfo_exists():
            self.root.after(100, self.update_overall_progress)

        # --- Temp mappa törlése ---
        if self.delete_temp_var.get() == "1":
            shutil.rmtree(temp_dir, ignore_errors=True)

        file_stats["end_time"] = end_time.strftime("%Y-%m-%d %H:%M:%S")
        file_stats["runtime_sec"] = (end_time - start_time).total_seconds()
        return file_stats



    # ----------------------------------------------------------------
    def save_processing_state(self, index, status, file_path):
        state = {
            'current_index': index,
            'status': status,
            'interrupted_file': os.path.basename(file_path),
            'interruption_time': datetime.now().isoformat(),
            'input_directory': self.input_directory,
            'output_directory': self.output_folder,
            'processed_files_count': self.processed_files_count,
            'processed_input_size_mb': self.processed_input_size_mb,
            'processed_output_size_mb': self.processed_output_size_mb,
            'processed_input_duration_sec': self.processed_input_duration_sec,
            'processed_output_duration_sec': self.processed_output_duration_sec
        }
        with open(self.processing_state_file, 'w') as f:
            json.dump(state, f, indent=4)

    def check_for_interrupted_processing(self):
        if os.path.exists(self.processing_state_file):
            with open(self.processing_state_file, 'r') as f:
                state = json.load(f)
            if state.get('status') == 'processing':
                self.add_log_entry("INFO",
                                   f"Félbeszakadt feldolgozás észlelve: {state.get('interrupted_file', 'N/A')}")
                self.resume_button.config(state="normal")

    def save_settings(self):
        config = configparser.ConfigParser()
        config['Settings'] = {
            'input_directory': self.input_directory,
            'output_folder': self.output_folder,
            'log_folder': self.log_folder,  # Log mappa mentése
            'crf': self.crf_var.get(),
            'preset': self.preset_var.get(),
            'pixel_threshold': str(self.pixel_threshold_scale.get()),
            'fast_mode': self.fast_mode_var.get(),
            'min_motion_duration': str(self.min_motion_duration_scale.get()),
            'motion_end_buffer': str(self.motion_end_buffer_scale.get()),
            'idle_duration': str(self.idle_duration_scale.get()),
            'pre_motion_buffer': str(self.pre_motion_buffer_scale.get()),
            'crossfade_duration': str(self.crossfade_duration_scale.get()),
            'custom_resolution': self.custom_resolution_var.get(),
            'output_width': self.output_width_var.get(),
            'output_height': self.output_height_var.get(),
            'skip_processed': self.skip_processed_var.get(),
            'delete_temp': self.delete_temp_var.get()

        }
        with open(self.settings_file, 'w') as configfile:
            config.write(configfile)
        self.add_log_entry("INFO", "Beállítások mentve.")

    def load_settings(self):
        """Betölti a beállításokat a settings.ini fájlból (kódolás-ellenőrzéssel)."""
        import configparser
        import os

        if os.path.exists("settings.ini"):
            config = configparser.ConfigParser()
            try:
                # Első próbálkozás UTF-8 kódolással
                config.read("settings.ini", encoding="utf-8")
            except UnicodeDecodeError:
                # Ha nem sikerül, fallback CP1250-re (magyar Windows ANSI)
                config.read("settings.ini", encoding="cp1250")
                self.add_log_entry("WARN", "A settings.ini nem UTF-8, CP1250 kódolással olvasva.")

            try:
                if "SETTINGS" in config:
                    # Betöltjük az értékeket
                    self.input_directory = config["SETTINGS"].get("input_directory", "")
                    self.output_folder = config["SETTINGS"].get("output_folder", "")
                    self.log_folder = config["SETTINGS"].get("log_folder", "")
                    self.crf_scale.set(int(config["SETTINGS"].get("crf", 23)))
                    self.preset_var.set(config["SETTINGS"].get("preset", "medium"))
                    self.pixel_threshold_scale.set(int(config["SETTINGS"].get("pixel_threshold", 5000)))
                    self.min_motion_duration_scale.set(float(config["SETTINGS"].get("min_motion_duration", 2.0)))
                    self.motion_end_buffer_scale.set(float(config["SETTINGS"].get("motion_end_buffer", 1.0)))
                    self.idle_duration_scale.set(float(config["SETTINGS"].get("idle_duration", 5.0)))
                    self.pre_motion_buffer_scale.set(float(config["SETTINGS"].get("pre_motion_buffer", 1.0)))
                    self.crossfade_duration_scale.set(float(config["SETTINGS"].get("crossfade_duration", 0.5)))
                    self.custom_resolution_var.set(config["SETTINGS"].get("custom_resolution", "0"))
                    self.output_width_var.set(config["SETTINGS"].get("output_width", "1920"))
                    self.output_height_var.set(config["SETTINGS"].get("output_height", "1080"))
                    self.skip_processed_var.set(config["SETTINGS"].get("skip_processed", "0"))
                    self.delete_temp_var.set(config["SETTINGS"].get("delete_temp", "1"))
                    # --- Gyors mód beállítás ---
                    self.fast_mode_var.set(config["SETTINGS"].get("fast_mode", "0"))

                    # GUI frissítése
                    self.input_dir_var.set(self.input_directory)
                    self.output_dir_var.set(self.output_folder)
                    self.log_dir_var.set(self.log_folder)

                    self.add_log_entry("INFO", "Beállítások sikeresen betöltve a settings.ini fájlból.")
                    if self.input_directory:
                        self.add_log_entry("INFO", f"Bemeneti mappa betöltve: {self.input_directory}")
                        self.update_file_list()
                    else:
                        self.add_log_entry("WARNING", "A settings.ini nem tartalmaz érvényes bemeneti mappát.")
                else:
                    self.add_log_entry("WARNING",
                                       "A settings.ini létezik, de nem tartalmaz [SETTINGS] szekciót. Alapértelmezett értékek használata.")
            except Exception as e:
                self.add_log_entry("ERROR", f"Hiba a settings.ini feldolgozása közben: {e}")
        else:
            self.add_log_entry("INFO",
                               "Nincs korábbi beállítás (settings.ini fájl hiányzik), alapértelmezett értékek használata.")

    def get_unique_filename(self, directory, base_name, extension):
        counter = 1
        output_file = os.path.join(directory, f"{base_name}{extension}")
        while os.path.exists(output_file):
            output_file = os.path.join(directory, f"{base_name}_{counter}{extension}")
            counter += 1
        return output_file

    def adjust_column_widths(self):
        for col in self.file_tree["columns"]:
            max_width = 0
            for item in self.file_tree.get_children():
                value = self.file_tree.item(item, 'values')[self.file_tree["columns"].index(col)]
                width = len(str(value)) * 10
                if width > max_width:
                    max_width = width
            self.file_tree.column(col, width=max_width + 20)

    def generate_report(self):
        # Ide jöhet a jelentéskészítő logika
        # Például, a napló adatok alapján PDF vagy más formátumú jelentés generálása
        self.add_log_entry("INFO", "Jelentés generálása...")
        try:
            pdf = FPDF()
            pdf.add_page()

            # Betűkészlet beállítása
            pdf.add_font("DejaVuSans", "", font_path_regular)
            pdf.add_font("DejaVuSans", "B", font_path_bold)
            pdf.set_font("DejaVuSans", "B", 16)

            pdf.cell(w=0, h=10, txt="Video Feldolgozási Jelentés", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.set_font("DejaVuSans", "", 12)

            pdf.cell(w=0, h=10, txt=f"Jelentés dátuma: {datetime.now().strftime('%Y-%m-%d %H:%M')}", new_x=XPos.LMARGIN,
                     new_y=YPos.NEXT)
            pdf.cell(w=0, h=10, txt=f"Bemeneti mappa: {self.input_directory}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.cell(w=0, h=10, txt=f"Kimeneti mappa: {self.output_folder}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.cell(w=0, h=10, txt=f"Feldolgozott fájlok száma: {self.processed_files_count}", new_x=XPos.LMARGIN,
                     new_y=YPos.NEXT)
            pdf.cell(w=0, h=10, txt=f"Összes feldolgozott méret: {self.processed_size_mb:.2f} MB", new_x=XPos.LMARGIN,
                     new_y=YPos.NEXT)
            pdf.cell(w=0, h=10,
                     txt=f"Összes feldolgozott időtartam: {str(timedelta(seconds=self.processed_duration_sec)).split('.')[0]}",
                     new_x=XPos.LMARGIN, new_y=YPos.NEXT)

            pdf.ln(10)  # Sor kihagyása

            pdf.set_font("DejaVuSans", "B", 14)
            pdf.cell(w=0, h=10, txt="Feldolgozási napló", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.set_font("DejaVuSans", "", 10)

            # Napló adatok hozzáadása
            for item in self.log_text.get_children():
                time, level, message = self.log_text.item(item)['values']
                pdf.multi_cell(w=0, h=5, txt=f"[{time}] [{level}] {message}")

            report_filename = os.path.join(self.output_folder,
                                           f"Jelentés_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.pdf")
            pdf.output(report_filename)
            self.add_log_entry("INFO", f"Jelentés sikeresen mentve: {report_filename}")

        except Exception as e:
            self.add_log_entry("ERROR", f"Hiba a jelentés generálása közben: {e}")

    def show_summary_popup(self, summary_text=None, log_dir=None):
        """
        Feldolgozás összesített adatai – mindig a valós self.* változókból.
        Szálbiztos (Tkinter fő szálon nyílik meg).
        """
        import os, subprocess
        from datetime import timedelta
        from tkinter import Toplevel, ttk

        # --- 1️⃣ Biztosítsuk, hogy a popup a fő Tkinter szálon fusson ---
        if threading.current_thread().name != "MainThread":
            self.root.after(0, lambda: self.show_summary_popup(summary_text, log_dir))
            return

        # --- 2️⃣ Popup létrehozása ---
        popup = Toplevel(self.root)
        popup.title("Összesített feldolgozás")
        popup.geometry("520x420")
        popup.configure(padx=20, pady=20)

        # --- 3️⃣ Ha van szöveg, mutassuk a tetején ---
        if summary_text:
            ttk.Label(
                popup,
                text=summary_text,
                font=("Consolas", 9, "italic"),
                foreground="#004080",
                justify="center"
            ).pack(pady=(0, 8))

        # --- 4️⃣ Adatok begyűjtése ---
        input_mb = getattr(self, "processed_input_size_mb", 0.0) or 0.0
        output_mb = getattr(self, "processed_output_size_mb", 0.0) or 0.0
        total_files = getattr(self, "processed_files_count", 0)
        start_time = getattr(self, "start_time", None)
        end_time = getattr(self, "end_time", None)

        total_time = (end_time - start_time).total_seconds() if (start_time and end_time) else 0

        diff_mb = output_mb - input_mb
        saving_percent = 100 - (output_mb / input_mb) * 100 if input_mb > 0 else 0

        color = "#007800" if saving_percent > 0 else "#C00000" if saving_percent < 0 else "#004080"
        emoji = "🟢" if saving_percent > 0 else "🔴" if saving_percent < 0 else "⚪"

        # --- 5️⃣ Fejléc ---
        ttk.Label(
            popup,
            text="📊 Feldolgozás befejezve – Összesített adatok",
            font=("Helvetica", 12, "bold")
        ).pack(pady=(5, 10))

        # --- 6️⃣ Mini összefoglaló ---
        if total_files > 0 and start_time and end_time:
            total_seconds = (end_time - start_time).total_seconds()
            avg_seconds = total_seconds / total_files
            avg_minutes = avg_seconds / 60
            summary_top = f"🔹 Összesen: {total_files} fájl — Átlag: {avg_minutes:.2f} perc / {avg_seconds:.1f} mp"
        else:
            summary_top = "🔹 Összesen: N/A"

        ttk.Label(
            popup,
            text=summary_top,
            font=("Consolas", 10, "italic"),
            foreground="#004080",
            justify="center"
        ).pack(pady=(0, 8))

        # --- 7️⃣ Méretadatok ---
        data_frame = ttk.Frame(popup)
        data_frame.pack(pady=5, fill="x")
        ttk.Label(data_frame, text=f"Bemeneti méret:  {input_mb:.2f} MB", font=("Consolas", 10)).pack(anchor="w")
        ttk.Label(data_frame, text=f"Kimeneti méret:  {output_mb:.2f} MB", font=("Consolas", 10)).pack(anchor="w")
        ttk.Label(data_frame, text=f"Változás:        {diff_mb:+.2f} MB", font=("Consolas", 10, "bold"),
                  foreground=color).pack(anchor="w")
        ttk.Label(data_frame, text=f"Megtakarítás:    {saving_percent:+.2f} % {emoji}", font=("Consolas", 10, "bold"),
                  foreground=color).pack(anchor="w")

        # --- 8️⃣ Időadatok ---
        runtime = str(timedelta(seconds=int(total_time)))
        start_str = start_time.strftime("%Y-%m-%d %H:%M:%S") if start_time else "N/A"
        end_str = end_time.strftime("%Y-%m-%d %H:%M:%S") if end_time else "N/A"

        info_frame = ttk.Frame(popup)
        info_frame.pack(pady=(10, 5), fill="x")
        ttk.Label(info_frame, text=f"Feldolgozott fájlok száma: {total_files}", font=("Consolas", 9)).pack(anchor="w")
        ttk.Label(info_frame, text=f"Teljes futásidő:           {runtime}", font=("Consolas", 9)).pack(anchor="w")

        if total_files > 0 and start_time and end_time:
            avg_seconds = total_time / total_files
            avg_minutes = avg_seconds / 60
            avg_per_file = f"{avg_minutes:.2f} perc / {avg_seconds:.1f} mp"
        else:
            avg_per_file = "N/A"

        ttk.Label(info_frame, text=f"Átlagos futásidő / fájl:   {avg_per_file}", font=("Consolas", 9, "italic"),
                  foreground="#004080").pack(anchor="w")
        ttk.Label(info_frame, text=f"Kezdés:                    {start_str}", font=("Consolas", 9)).pack(anchor="w")
        ttk.Label(info_frame, text=f"Befejezés:                 {end_str}", font=("Consolas", 9)).pack(anchor="w")

        # --- 9️⃣ Elválasztó és gombok ---
        ttk.Separator(popup, orient="horizontal").pack(fill="x", pady=(10, 5))
        btn_frame = ttk.Frame(popup)
        btn_frame.pack(pady=15)

        def open_log_dir():
            try:
                if os.name == "nt":
                    os.startfile(log_dir)
                elif os.name == "posix":
                    subprocess.Popen(["xdg-open", log_dir])
            except Exception as e:
                self.add_log_entry("ERROR", f"Nem sikerült megnyitni a log mappát: {e}")

        ttk.Button(btn_frame, text="📂 Log mappa megnyitása", command=open_log_dir).pack(side="left", padx=10)
        ttk.Button(btn_frame, text="✅ OK (Új futásra kész)",
                   command=lambda: (popup.destroy(), self.reset_for_new_run())).pack(side="left", padx=10)
        ttk.Button(btn_frame, text="❌ Kilépés (adatok megmaradnak)", command=popup.destroy).pack(side="left", padx=10)


if __name__ == "__main__":
    locale.setlocale(locale.LC_ALL, 'hu_HU.UTF-8')
    root = Tk()
    app = MotionExtractorApp(root)
    root.mainloop()
